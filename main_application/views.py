"""
Wajir County Government ERP System - Enhanced Authentication Views
Includes 2FA, account locking, login attempt tracking
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from datetime import datetime, timedelta
from decimal import Decimal
import logging
import random
import string

from .models import *

logger = logging.getLogger(__name__)


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def get_client_ip(request):
    """Get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def get_session_key(request):
    """Get or create session key"""
    if not hasattr(request.session, 'session_key') or not request.session.session_key:
        request.session.create()
    return request.session.session_key or 'no-session'


def send_security_email(user, subject, message):
    """Send security notification email"""
    try:
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [user.email],
            fail_silently=False,
        )
        return True
    except Exception as e:
        logger.error(f"Failed to send security email to {user.email}: {str(e)}")
        return False


def create_security_notification(user, notification_type, ip_address, message):
    """Create a security notification record"""
    notification = SecurityNotification.objects.create(
        user=user,
        notification_type=notification_type,
        ip_address=ip_address,
        message=message
    )
    
    # Subject mapping
    subject_map = {
        'failed_login': 'Security Alert: Failed Login Attempt',
        'account_locked': 'Security Alert: Account Locked',
        'tfa_code': 'Your 2FA Verification Code',
        'successful_login': 'Security: Successful Login',
        'account_unlocked': 'Security: Account Unlocked'
    }
    
    subject = subject_map.get(notification_type, 'Security Notification')
    email_sent = send_security_email(user, subject, message)
    
    if email_sent:
        notification.email_sent = True
        notification.email_sent_at = timezone.now()
        notification.save()


def check_account_lock(username, ip_address=None):
    """Check if account is locked and return lock status"""
    try:
        user = User.objects.get(username=username)
        account_lock, created = AccountLock.objects.get_or_create(
            user=user,
            defaults={
                'failed_attempts': 0,
                'is_locked': False,
                'last_attempt_ip': ip_address or '127.0.0.1'
            }
        )
        
        if account_lock.is_account_locked():
            return True, account_lock, user
        return False, account_lock, user
    except User.DoesNotExist:
        return False, None, None


def handle_failed_login(username, ip_address, user_agent):
    """Handle failed login attempt"""
    # Log the attempt
    LoginAttempt.objects.create(
        username=username,
        ip_address=ip_address,
        success=False,
        user_agent=user_agent
    )
    
    try:
        user = User.objects.get(username=username)
        account_lock, created = AccountLock.objects.get_or_create(
            user=user,
            defaults={'failed_attempts': 0, 'is_locked': False, 'last_attempt_ip': ip_address}
        )
        
        account_lock.failed_attempts += 1
        account_lock.last_attempt_ip = ip_address
        
        if account_lock.failed_attempts >= 5:  # Lock after 5 failed attempts
            account_lock.is_locked = True
            account_lock.unlock_time = timezone.now() + timedelta(minutes=30)  # Lock for 30 minutes
            account_lock.save()
            
            # Send security notification
            message = f"""
Security Alert: Your account has been locked due to multiple failed login attempts.

Details:
- Time: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}
- IP Address: {ip_address}
- Failed Attempts: {account_lock.failed_attempts}

Your account will be automatically unlocked after 30 minutes, or contact the system administrator.

If this wasn't you, please contact support immediately.
            """.strip()
            
            create_security_notification(user, 'account_locked', ip_address, message)
            return True  # Account was locked
        else:
            account_lock.save()
            
            # Send failed attempt notification
            attempts_left = 5 - account_lock.failed_attempts
            message = f"""
Security Alert: Failed login attempt detected on your account.

Details:
- Time: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}
- IP Address: {ip_address}
- Attempts remaining: {attempts_left}

If this wasn't you, please contact support immediately.
            """.strip()
            
            create_security_notification(user, 'failed_login', ip_address, message)
    
    except User.DoesNotExist:
        pass
    
    return False


def generate_tfa_code(user, ip_address, session_key):
    """Generate and send 2FA code"""
    # Invalidate any existing unused codes for this user
    TwoFactorCode.objects.filter(user=user, used=False).update(used=True)
    
    # Create new 2FA code
    tfa_code = TwoFactorCode.objects.create(
        user=user,
        ip_address=ip_address,
        session_key=session_key
    )
    
    # Send code via email
    message = f"""
Your verification code for Wajir County ERP System:

{tfa_code.code}

This code will expire in 2 minutes at {tfa_code.expires_at.strftime('%H:%M:%S')}.

Time: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}
IP Address: {ip_address}

If you didn't request this code, please contact support immediately.
    """.strip()
    
    create_security_notification(user, 'tfa_code', ip_address, message)
    
    return tfa_code


def get_user_dashboard_url(user):
    """Get appropriate dashboard URL based on user's role"""
    # Check if superuser
    if user.is_superuser:
        return 'admin_dashboard'
    
    # Get user's primary active role
    user_role = UserRole.objects.filter(user=user, is_active=True).first()
    
    if user_role:
        role_name = user_role.role.name.lower()
        
        # Role-based dashboard routing
        role_dashboard_map = {
            'system administrator': 'admin_dashboard',
            'revenue officer': 'revenue_dashboard',
            'health worker': 'health_dashboard',
            'fleet manager': 'fleet_dashboard',
            'hr manager': 'hr_dashboard',
            'finance officer': 'finance_dashboard',
            'lands officer': 'lands_dashboard',
            'asset manager': 'asset_dashboard',
            'inventory manager': 'inventory_dashboard',
        }
        
        for key, dashboard in role_dashboard_map.items():
            if key in role_name:
                return dashboard
    
    # Default dashboard
    return 'general_dashboard'


# ============================================================================
# AUTHENTICATION VIEWS
# ============================================================================

def login_view(request):
    """Enhanced login view with 2FA and security features"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    # Check if session expired
    session_expired = request.session.pop('session_expired', False)
    redirect_after_login = request.session.get('redirect_after_login')
    
    # Get the 'next' parameter from URL
    next_url = request.GET.get('next', redirect_after_login)
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        tfa_code = request.POST.get('tfa_code', '').strip()
        
        ip_address = get_client_ip(request)
        user_agent = request.META.get('HTTP_USER_AGENT', '')
        
        # Check if account is locked
        is_locked, account_lock, user_obj = check_account_lock(username, ip_address)
        if is_locked:
            minutes_left = int((account_lock.unlock_time - timezone.now()).total_seconds() / 60)
            messages.error(
                request, 
                f'Account is temporarily locked. Please try again in {minutes_left} minutes.'
            )
            return render(request, 'auth/login.html', {'next': next_url})
        
        # If 2FA code is provided, verify it
        if tfa_code:
            return handle_tfa_verification(request, username, tfa_code, ip_address, next_url)
        
        # Regular authentication
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Log successful authentication
            LoginAttempt.objects.create(
                username=username,
                ip_address=ip_address,
                success=True,
                user_agent=user_agent
            )
            
            # Reset failed attempts on successful authentication
            if hasattr(user, 'account_lock'):
                account_lock = user.account_lock
                account_lock.failed_attempts = 0
                account_lock.is_locked = False
                account_lock.unlock_time = None
                account_lock.save()
            
            # Check if user requires 2FA (all staff members)
            if user.is_active_staff or user.is_superuser:
                # Require 2FA for staff and admin users
                session_key = get_session_key(request)
                tfa_code_obj = generate_tfa_code(user, ip_address, session_key)
                
                # Store pending login data in session
                request.session['pending_login_user_id'] = user.id
                request.session['pending_login_time'] = timezone.now().isoformat()
                request.session['tfa_code_id'] = tfa_code_obj.id
                request.session['pending_next_url'] = next_url
                
                messages.info(request, 'Verification code sent to your email. Please check and enter the code.')
                return render(request, 'auth/login.html', {
                    'show_tfa': True,
                    'username': username,
                    'expires_at': tfa_code_obj.expires_at.isoformat(),
                    'next': next_url,
                })
            else:
                # Direct login for citizen portal users
                login(request, user)
                
                # Initialize session activity tracking
                request.session['last_activity'] = timezone.now().isoformat()
                
                # Clear the redirect flag
                request.session.pop('redirect_after_login', None)
                
                messages.success(request, f'Welcome back, {user.get_full_name()}!')
                
                # Redirect to next URL or default dashboard
                if next_url and next_url != '/':
                    return redirect(next_url)
                else:
                    return redirect('citizen_dashboard')
        else:
            # Handle failed login
            was_locked = handle_failed_login(username, ip_address, user_agent)
            if was_locked:
                messages.error(request, 'Account locked due to multiple failed attempts. Check your email for details.')
            else:
                messages.error(request, 'Invalid username or password')
    
    # Show session expired message if applicable
    if session_expired:
        messages.warning(request, 'Your session expired due to inactivity. Please log in again.')
    
    return render(request, 'auth/login.html', {'next': next_url})


def handle_tfa_verification(request, username, tfa_code, ip_address, next_url=None):
    """Handle 2FA code verification with redirect support"""
    try:
        # Get pending login data from session
        pending_user_id = request.session.get('pending_login_user_id')
        tfa_code_id = request.session.get('tfa_code_id')
        stored_next_url = request.session.get('pending_next_url', next_url)
        
        if not pending_user_id or not tfa_code_id:
            messages.error(request, 'Session expired. Please login again.')
            return redirect('login')
        
        user = get_object_or_404(User, id=pending_user_id, username=username)
        code_obj = get_object_or_404(TwoFactorCode, id=tfa_code_id, user=user)
        
        # Check if code is valid
        if not code_obj.is_valid():
            messages.error(request, 'Verification code has expired or already been used.')
            return render(request, 'auth/login.html', {
                'show_tfa': True,
                'username': username,
                'code_expired': True,
                'next': stored_next_url,
            })
        
        # Verify the code
        if code_obj.code == tfa_code:
            # Mark code as used
            code_obj.mark_as_used()
            
            # Clear session data
            request.session.pop('pending_login_user_id', None)
            request.session.pop('pending_login_time', None)
            request.session.pop('tfa_code_id', None)
            request.session.pop('pending_next_url', None)
            request.session.pop('redirect_after_login', None)
            
            # Log the user in
            login(request, user)
            
            # Initialize session activity tracking
            request.session['last_activity'] = timezone.now().isoformat()
            
            # Send successful login notification
            message = f"""
Successful login to your account:

Time: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}
IP Address: {ip_address}
Department: {user.department.name if user.department else 'N/A'}

If this wasn't you, please contact support immediately.
            """.strip()
            
            create_security_notification(user, 'successful_login', ip_address, message)
            
            messages.success(request, f'Welcome back, {user.get_full_name()}!')
            
            # Redirect based on stored URL or user role
            if stored_next_url and stored_next_url != '/':
                return redirect(stored_next_url)
            else:
                dashboard_url = get_user_dashboard_url(user)
                return redirect(dashboard_url)
        else:
            messages.error(request, 'Invalid verification code.')
            return render(request, 'auth/login.html', {
                'show_tfa': True,
                'username': username,
                'expires_at': code_obj.expires_at.isoformat(),
                'next': stored_next_url,
            })
    
    except Exception as e:
        logger.error(f"2FA verification error: {str(e)}")
        messages.error(request, 'An error occurred during verification. Please try again.')
        return redirect('login')


def resend_tfa_code(request):
    """Resend 2FA code via AJAX"""
    if request.method == 'POST':
        try:
            username = request.POST.get('username')
            pending_user_id = request.session.get('pending_login_user_id')
            
            if not pending_user_id:
                return JsonResponse({'success': False, 'message': 'Session expired'})
            
            user = get_object_or_404(User, id=pending_user_id, username=username)
            ip_address = get_client_ip(request)
            
            # Generate new code
            session_key = get_session_key(request)
            tfa_code_obj = generate_tfa_code(user, ip_address, session_key)
            request.session['tfa_code_id'] = tfa_code_obj.id
            
            return JsonResponse({
                'success': True,
                'message': 'New verification code sent to your email.',
                'expires_at': tfa_code_obj.expires_at.isoformat(),
            })
        
        except Exception as e:
            logger.error(f"Resend 2FA code error: {str(e)}")
            return JsonResponse({'success': False, 'message': 'Failed to send code'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})


def logout_view(request):
    """Logout user and clear session"""
    # Clear any pending login data
    request.session.pop('pending_login_user_id', None)
    request.session.pop('pending_login_time', None)
    request.session.pop('tfa_code_id', None)
    
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')


@login_required(login_url='login')
def dashboard(request):
    """Route users to appropriate dashboard based on their role"""
    dashboard_url = get_user_dashboard_url(request.user)
    return redirect(dashboard_url)

# ============================================================================
# SYSTEM ADMINISTRATOR DASHBOARD
# ============================================================================

@login_required(login_url='login')
def admin_dashboard(request):
    """Admin dashboard with comprehensive statistics and analytics"""
    from django.db.models import Sum, Count, Q, Avg
    from datetime import datetime, timedelta
    import json
    
    # Get current date and time ranges
    today = timezone.now().date()
    thirty_days_ago = today - timedelta(days=30)
    six_months_ago = today - timedelta(days=180)
    current_year = today.year
    
    # ============================================================================
    # REVENUE STATISTICS
    # ============================================================================
    
    # Total revenue collected
    total_revenue = Payment.objects.filter(
        status='completed'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Revenue this month
    revenue_this_month = Payment.objects.filter(
        status='completed',
        payment_date__year=today.year,
        payment_date__month=today.month
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Revenue today
    revenue_today = Payment.objects.filter(
        status='completed',
        payment_date__date=today
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Total bills
    total_bills = Bill.objects.count()
    pending_bills = Bill.objects.filter(status__in=['issued', 'overdue']).count()
    paid_bills = Bill.objects.filter(status='paid').count()
    
    # Outstanding amount
    outstanding_amount = Bill.objects.filter(
        status__in=['issued', 'overdue', 'partially_paid']
    ).aggregate(total=Sum('balance'))['total'] or 0
    
    # ============================================================================
    # CITIZEN & SERVICE STATISTICS
    # ============================================================================
    
    total_citizens = Citizen.objects.filter(is_active=True).count()
    total_licenses = License.objects.count()
    active_licenses = License.objects.filter(status='active').count()
    expired_licenses = License.objects.filter(status='expired').count()
    
    total_properties = Property.objects.filter(status='active').count()
    total_vehicles = Vehicle.objects.filter(is_active=True).count()
    
    # ============================================================================
    # HEALTH STATISTICS
    # ============================================================================
    
    total_patients = Patient.objects.filter(is_active=True).count()
    visits_today = Visit.objects.filter(visit_date__date=today).count()
    visits_this_month = Visit.objects.filter(
        visit_date__year=today.year,
        visit_date__month=today.month
    ).count()
    
    pending_lab_tests = LabTest.objects.filter(status='pending').count()
    active_admissions = Admission.objects.filter(status='admitted').count()
    
    # ============================================================================
    # FLEET & ASSETS STATISTICS
    # ============================================================================
    
    total_vehicles_fleet = FleetVehicle.objects.filter(status='active').count()
    vehicles_maintenance = FleetVehicle.objects.filter(status='maintenance').count()
    
    fuel_cost_month = FuelTransaction.objects.filter(
        transaction_date__year=today.year,
        transaction_date__month=today.month
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    total_assets = Asset.objects.filter(status='active').count()
    
    # ============================================================================
    # HR STATISTICS
    # ============================================================================
    
    total_employees = User.objects.filter(is_active_staff=True, is_active=True).count()
    present_today = Attendance.objects.filter(
        attendance_date=today,
        attendance_type='check_in'
    ).values('employee').distinct().count()
    
    pending_leaves = LeaveApplication.objects.filter(status='pending').count()
    
    # ============================================================================
    # CHART DATA - Revenue Trends
    # ============================================================================
    
    # Monthly revenue for last 6 months
    monthly_revenue_data = []
    monthly_labels = []
    
    for i in range(5, -1, -1):
        date = today - timedelta(days=i*30)
        month_start = date.replace(day=1)
        if i == 0:
            month_end = today
        else:
            next_month = month_start.replace(day=28) + timedelta(days=4)
            month_end = next_month - timedelta(days=next_month.day)
        
        revenue = Payment.objects.filter(
            status='completed',
            payment_date__range=[month_start, month_end]
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        monthly_revenue_data.append(float(revenue))
        monthly_labels.append(month_start.strftime('%b %Y'))
    
    revenue_chart_data = json.dumps({
        'labels': monthly_labels,
        'data': monthly_revenue_data
    })
    
    # ============================================================================
    # CHART DATA - Revenue by Stream (Donut Chart)
    # ============================================================================
    
    revenue_by_stream = Payment.objects.filter(
        status='completed'
    ).values('revenue_stream__name').annotate(
        total=Sum('amount')
    ).order_by('-total')[:6]
    
    stream_labels = [item['revenue_stream__name'] for item in revenue_by_stream]
    stream_data = [float(item['total']) for item in revenue_by_stream]
    stream_colors = ['#139145', '#C18B5A', '#CD4F27', '#3B82F6', '#8B5CF6', '#06B6D4']
    
    revenue_stream_chart = json.dumps({
        'labels': stream_labels,
        'data': stream_data,
        'colors': stream_colors
    })
    
    # ============================================================================
    # CHART DATA - Bills Status (Pie Chart)
    # ============================================================================
    
    bill_status_data = Bill.objects.values('status').annotate(
        count=Count('id')
    )
    
    bill_labels = [item['status'].replace('_', ' ').title() for item in bill_status_data]
    bill_counts = [item['count'] for item in bill_status_data]
    bill_colors = ['#139145', '#C18B5A', '#CD4F27', '#EF4444', '#F59E0B']
    
    bill_status_chart = json.dumps({
        'labels': bill_labels,
        'data': bill_counts,
        'colors': bill_colors
    })
    
    # ============================================================================
    # CHART DATA - Daily Revenue (Last 30 Days - Area Chart)
    # ============================================================================
    
    daily_revenue_data = []
    daily_labels = []
    
    for i in range(29, -1, -1):
        date = today - timedelta(days=i)
        revenue = Payment.objects.filter(
            status='completed',
            payment_date__date=date
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        daily_revenue_data.append(float(revenue))
        daily_labels.append(date.strftime('%b %d'))
    
    daily_revenue_chart = json.dumps({
        'labels': daily_labels,
        'data': daily_revenue_data
    })
    
    # ============================================================================
    # CHART DATA - Revenue by Sub-County (Bar Chart)
    # ============================================================================
    
    subcounty_revenue = Payment.objects.filter(
        status='completed'
    ).values('sub_county__name').annotate(
        total=Sum('amount')
    ).order_by('-total')[:10]
    
    subcounty_labels = [item['sub_county__name'] or 'Unknown' for item in subcounty_revenue]
    subcounty_data = [float(item['total']) for item in subcounty_revenue]
    
    subcounty_chart = json.dumps({
        'labels': subcounty_labels,
        'data': subcounty_data
    })
    
    # ============================================================================
    # CHART DATA - Patients by Month (Line Chart)
    # ============================================================================
    
    patient_monthly_data = []
    patient_labels = []
    
    for i in range(5, -1, -1):
        date = today - timedelta(days=i*30)
        month_start = date.replace(day=1)
        if i == 0:
            month_end = today
        else:
            next_month = month_start.replace(day=28) + timedelta(days=4)
            month_end = next_month - timedelta(days=next_month.day)
        
        visits = Visit.objects.filter(
            visit_date__range=[month_start, month_end]
        ).count()
        
        patient_monthly_data.append(visits)
        patient_labels.append(month_start.strftime('%b'))
    
    patient_chart = json.dumps({
        'labels': patient_labels,
        'data': patient_monthly_data
    })
    
    # ============================================================================
    # CHART DATA - License Status (Donut Chart)
    # ============================================================================
    
    license_status = License.objects.values('status').annotate(
        count=Count('id')
    )
    
    license_labels = [item['status'].replace('_', ' ').title() for item in license_status]
    license_counts = [item['count'] for item in license_status]
    license_colors = ['#139145', '#C18B5A', '#CD4F27', '#EF4444', '#F59E0B', '#3B82F6']
    
    license_chart = json.dumps({
        'labels': license_labels,
        'data': license_counts,
        'colors': license_colors
    })
    
    # ============================================================================
    # CHART DATA - Fleet Fuel Consumption (Bar Chart)
    # ============================================================================
    
    fuel_by_vehicle = FuelTransaction.objects.filter(
        transaction_date__gte=thirty_days_ago
    ).values('vehicle__registration_number').annotate(
        total_liters=Sum('quantity_liters'),
        total_cost=Sum('total_amount')
    ).order_by('-total_cost')[:10]
    
    fuel_labels = [item['vehicle__registration_number'] for item in fuel_by_vehicle]
    fuel_data = [float(item['total_cost']) for item in fuel_by_vehicle]
    
    fuel_chart = json.dumps({
        'labels': fuel_labels,
        'data': fuel_data
    })
    
    # ============================================================================
    # CHART DATA - Payment Methods Distribution (Pie Chart)
    # ============================================================================
    
    payment_methods = Payment.objects.filter(
        status='completed'
    ).values('payment_method__name').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('-total')
    
    method_labels = [item['payment_method__name'] for item in payment_methods]
    method_data = [float(item['total']) for item in payment_methods]
    method_colors = ['#139145', '#C18B5A', '#CD4F27', '#3B82F6', '#8B5CF6']
    
    payment_method_chart = json.dumps({
        'labels': method_labels,
        'data': method_data,
        'colors': method_colors
    })
    
    # ============================================================================
    # Recent Activities
    # ============================================================================
    
    recent_payments = Payment.objects.filter(
        status='completed'
    ).select_related('citizen', 'revenue_stream', 'payment_method').order_by('-payment_date')[:10]
    
    recent_bills = Bill.objects.select_related(
        'citizen', 'revenue_stream'
    ).order_by('-created_at')[:10]
    
    # ============================================================================
    # Alerts & Notifications
    # ============================================================================
    
    overdue_bills_count = Bill.objects.filter(
        status='overdue',
        due_date__lt=today
    ).count()
    
    expiring_licenses = License.objects.filter(
        status='active',
        expiry_date__range=[today, today + timedelta(days=30)]
    ).count()
    
    context = {
        # Revenue Stats
        'total_revenue': total_revenue,
        'revenue_this_month': revenue_this_month,
        'revenue_today': revenue_today,
        'total_bills': total_bills,
        'pending_bills': pending_bills,
        'paid_bills': paid_bills,
        'outstanding_amount': outstanding_amount,
        
        # Citizen Stats
        'total_citizens': total_citizens,
        'total_licenses': total_licenses,
        'active_licenses': active_licenses,
        'expired_licenses': expired_licenses,
        'total_properties': total_properties,
        'total_vehicles': total_vehicles,
        
        # Health Stats
        'total_patients': total_patients,
        'visits_today': visits_today,
        'visits_this_month': visits_this_month,
        'pending_lab_tests': pending_lab_tests,
        'active_admissions': active_admissions,
        
        # Fleet & Assets
        'total_vehicles_fleet': total_vehicles_fleet,
        'vehicles_maintenance': vehicles_maintenance,
        'fuel_cost_month': fuel_cost_month,
        'total_assets': total_assets,
        
        # HR Stats
        'total_employees': total_employees,
        'present_today': present_today,
        'pending_leaves': pending_leaves,
        
        # Chart Data
        'revenue_chart_data': revenue_chart_data,
        'revenue_stream_chart': revenue_stream_chart,
        'bill_status_chart': bill_status_chart,
        'daily_revenue_chart': daily_revenue_chart,
        'subcounty_chart': subcounty_chart,
        'patient_chart': patient_chart,
        'license_chart': license_chart,
        'fuel_chart': fuel_chart,
        'payment_method_chart': payment_method_chart,
        
        # Recent Data
        'recent_payments': recent_payments,
        'recent_bills': recent_bills,
        
        # Alerts
        'overdue_bills_count': overdue_bills_count,
        'expiring_licenses': expiring_licenses,
    }
    
    return render(request, 'dashboards/admin_dashboard.html', context)


# ============================================================================
# REVENUE OFFICER DASHBOARD
# ============================================================================

@login_required(login_url='login')
def revenue_dashboard(request):
    """Revenue Officer Dashboard"""
    today = timezone.now().date()
    
    # Revenue statistics
    total_bills = Bill.objects.filter(status__in=['issued', 'partially_paid', 'overdue'])
    total_bills_amount = total_bills.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_collected = Payment.objects.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Today's collections
    today_collections = Payment.objects.filter(
        payment_date__date=today,
        status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # This month's collections
    month_start = today.replace(day=1)
    month_collections = Payment.objects.filter(
        payment_date__date__gte=month_start,
        status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Revenue by stream
    revenue_by_stream = Payment.objects.filter(
        status='completed',
        payment_date__date__gte=month_start
    ).values('revenue_stream__name').annotate(
        total=Sum('amount')
    ).order_by('-total')[:5]
    
    context = {
        'total_bills': total_bills.count(),
        'total_bills_amount': total_bills_amount,
        'total_collected': total_collected,
        'collection_rate': (total_collected / total_bills_amount * 100) if total_bills_amount > 0 else 0,
        
        'today_collections': today_collections,
        'month_collections': month_collections,
        
        'pending_bills': Bill.objects.filter(status='issued').count(),
        'overdue_bills': Bill.objects.filter(status='overdue').count(),
        
        'revenue_by_stream': revenue_by_stream,
        'recent_payments': Payment.objects.filter(status='completed').order_by('-payment_date')[:10],
        'recent_bills': Bill.objects.order_by('-created_at')[:10],
        
        # Sub-county breakdown
        'sub_county': request.user.sub_county,
    }
    
    return render(request, 'dashboards/revenue_dashboard.html', context)


# ============================================================================
# HEALTH WORKER DASHBOARD
# ============================================================================

@login_required(login_url='login')
def health_dashboard(request):
    """Health Worker Dashboard"""
    today = timezone.now().date()
    
    # Get user's facility if assigned
    user_facility = None
    if hasattr(request.user, 'department'):
        # Assuming health workers are assigned to facilities
        user_facility = HealthFacility.objects.filter(is_active=True).first()
    
    # Patient statistics
    total_patients = Patient.objects.filter(is_active=True).count()
    
    # Today's visits
    today_visits = Visit.objects.filter(visit_date__date=today)
    
    # Current admissions
    current_admissions = Admission.objects.filter(status='admitted')
    
    # Pending lab tests
    pending_lab_tests = LabTest.objects.filter(status='pending')
    
    # Pending prescriptions
    pending_prescriptions = Prescription.objects.filter(status='pending')
    
    context = {
        'total_patients': total_patients,
        'today_visits': today_visits.count(),
        'current_admissions': current_admissions.count(),
        'pending_lab_tests': pending_lab_tests.count(),
        'pending_prescriptions': pending_prescriptions.count(),
        
        'user_facility': user_facility,
        'recent_visits': Visit.objects.order_by('-visit_date')[:10],
        'recent_admissions': Admission.objects.order_by('-admission_date')[:5],
        'pending_tests': pending_lab_tests[:10],
        
        # Facility statistics
        'facilities': HealthFacility.objects.filter(is_active=True),
        'total_beds': HealthFacility.objects.aggregate(Sum('bed_capacity'))['bed_capacity__sum'] or 0,
    }
    
    return render(request, 'dashboards/health_dashboard.html', context)


# ============================================================================
# FLEET MANAGER DASHBOARD
# ============================================================================

@login_required(login_url='login')
def fleet_dashboard(request):
    """Fleet Manager Dashboard"""
    today = timezone.now().date()
    
    # Fleet statistics
    total_vehicles = FleetVehicle.objects.all()
    active_vehicles = total_vehicles.filter(status='active')
    maintenance_vehicles = total_vehicles.filter(status='maintenance')
    
    # Fuel consumption today
    today_fuel = FuelTransaction.objects.filter(
        transaction_date__date=today
    ).aggregate(
        total_liters=Sum('quantity_liters'),
        total_cost=Sum('total_amount')
    )
    
    # This month's fuel
    month_start = today.replace(day=1)
    month_fuel = FuelTransaction.objects.filter(
        transaction_date__date__gte=month_start
    ).aggregate(
        total_liters=Sum('quantity_liters'),
        total_cost=Sum('total_amount')
    )
    
    # Upcoming maintenance
    upcoming_maintenance = VehicleMaintenance.objects.filter(
        status='scheduled',
        scheduled_date__gte=today
    ).order_by('scheduled_date')[:5]
    
    # Active trips
    active_trips = VehicleTrip.objects.filter(status='in_progress')
    
    context = {
        'total_vehicles': total_vehicles.count(),
        'active_vehicles': active_vehicles.count(),
        'maintenance_vehicles': maintenance_vehicles.count(),
        'inactive_vehicles': total_vehicles.filter(status='inactive').count(),
        
        'today_fuel_liters': today_fuel['total_liters'] or 0,
        'today_fuel_cost': today_fuel['total_cost'] or 0,
        'month_fuel_liters': month_fuel['total_liters'] or 0,
        'month_fuel_cost': month_fuel['total_cost'] or 0,
        
        'upcoming_maintenance': upcoming_maintenance,
        'active_trips': active_trips,
        
        'recent_fuel_transactions': FuelTransaction.objects.order_by('-transaction_date')[:10],
        'vehicles': active_vehicles[:10],
        'vehicle_types': total_vehicles.values('vehicle_type').annotate(count=Count('id')),
    }
    
    return render(request, 'dashboards/fleet_dashboard.html', context)


# ============================================================================
# HR MANAGER DASHBOARD
# ============================================================================

@login_required(login_url='login')
def hr_dashboard(request):
    """HR Manager Dashboard"""
    today = timezone.now().date()
    
    # Staff statistics
    total_staff = User.objects.filter(is_active=True, is_active_staff=True)
    
    # Leave statistics
    pending_leaves = LeaveApplication.objects.filter(status='pending')
    approved_leaves = LeaveApplication.objects.filter(
        status='approved',
        start_date__lte=today,
        end_date__gte=today
    )
    
    # Pending transfers
    pending_transfers = Transfer.objects.filter(status='pending')
    
    # Attendance today
    today_attendance = Attendance.objects.filter(attendance_date=today)
    checked_in = today_attendance.filter(attendance_type='check_in').values('employee').distinct().count()
    
    # Upcoming trainings
    upcoming_trainings = TrainingProgram.objects.filter(
        start_date__gte=today,
        is_active=True
    ).order_by('start_date')[:5]
    
    # Open disciplinary cases
    open_cases = DisciplinaryCase.objects.exclude(status='closed')
    
    context = {
        'total_staff': total_staff.count(),
        'pending_leaves': pending_leaves.count(),
        'approved_leaves': approved_leaves.count(),
        'pending_transfers': pending_transfers.count(),
        
        'checked_in_today': checked_in,
        'attendance_rate': (checked_in / total_staff.count() * 100) if total_staff.count() > 0 else 0,
        
        'upcoming_trainings': upcoming_trainings,
        'open_cases': open_cases.count(),
        
        'recent_leaves': LeaveApplication.objects.order_by('-created_at')[:10],
        'recent_transfers': Transfer.objects.order_by('-created_at')[:5],
        'recent_attendance': today_attendance.order_by('-attendance_time')[:10],
        
        # Department breakdown
        'staff_by_department': total_staff.values('department__name').annotate(count=Count('id')).order_by('-count')[:5],
    }
    
    return render(request, 'dashboards/hr_dashboard.html', context)


# ============================================================================
# FINANCE OFFICER DASHBOARD
# ============================================================================

@login_required(login_url='login')
def finance_dashboard(request):
    """Finance Officer Dashboard"""
    today = timezone.now().date()
    month_start = today.replace(day=1)
    
    # Financial statistics
    total_revenue = Payment.objects.filter(
        status='completed',
        payment_date__date__gte=month_start
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Pending payments
    pending_payments = Payment.objects.filter(status='pending')
    
    # Bills statistics
    total_bills_amount = Bill.objects.filter(
        bill_date__gte=month_start
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    outstanding_amount = Bill.objects.filter(
        status__in=['issued', 'partially_paid', 'overdue']
    ).aggregate(Sum('balance'))['balance__sum'] or 0
    
    # Revenue by stream
    revenue_by_stream = Payment.objects.filter(
        status='completed',
        payment_date__date__gte=month_start
    ).values('revenue_stream__name').annotate(
        total=Sum('amount')
    ).order_by('-total')[:10]
    
    # Budget vs actual
    budgets = RevenueBudget.objects.filter(
        period_start__lte=today,
        period_end__gte=today
    )
    
    # Bank reconciliation status
    pending_reconciliation = BankReconciliation.objects.filter(is_reconciled=False).count()
    
    context = {
        'total_revenue': total_revenue,
        'total_bills_amount': total_bills_amount,
        'outstanding_amount': outstanding_amount,
        'collection_rate': (total_revenue / total_bills_amount * 100) if total_bills_amount > 0 else 0,
        
        'pending_payments': pending_payments.count(),
        'pending_reconciliation': pending_reconciliation,
        
        'revenue_by_stream': revenue_by_stream,
        'budgets': budgets,
        
        'recent_payments': Payment.objects.filter(status='completed').order_by('-payment_date')[:15],
        'recent_reversals': PaymentReversal.objects.order_by('-created_at')[:5],
        
        # Payment methods breakdown
        'payment_by_method': Payment.objects.filter(
            status='completed',
            payment_date__date__gte=month_start
        ).values('payment_method__name').annotate(
            total=Sum('amount'),
            count=Count('id')
        ).order_by('-total'),
    }
    
    return render(request, 'dashboards/finance_dashboard.html', context)


# ============================================================================
# LANDS OFFICER DASHBOARD
# ============================================================================

@login_required(login_url='login')
def lands_dashboard(request):
    """Lands Officer Dashboard"""
    today = timezone.now().date()
    
    # Property statistics
    total_properties = Property.objects.filter(status='active')
    
    # Land rates collection
    land_rate_stream = RevenueStream.objects.filter(code='LND-RT').first()
    if land_rate_stream:
        land_rates_collected = Payment.objects.filter(
            revenue_stream=land_rate_stream,
            status='completed'
        ).aggregate(Sum('amount'))['amount__sum'] or 0
    else:
        land_rates_collected = 0
    
    # Development applications
    pending_applications = DevelopmentApplication.objects.filter(status='submitted')
    under_review = DevelopmentApplication.objects.filter(status='under_review')
    
    # Recent property transactions
    recent_transfers = PropertyOwnershipHistory.objects.order_by('-transfer_date')[:10]
    
    # Active caveats
    active_caveats = PropertyCaveat.objects.filter(is_active=True)
    
    # Property valuations
    recent_valuations = PropertyValuation.objects.filter(is_current=True).order_by('-valuation_date')[:5]
    
    context = {
        'total_properties': total_properties.count(),
        'land_rates_collected': land_rates_collected,
        
        'pending_applications': pending_applications.count(),
        'under_review': under_review.count(),
        'active_caveats': active_caveats.count(),
        
        'recent_transfers': recent_transfers,
        'recent_valuations': recent_valuations,
        'recent_applications': DevelopmentApplication.objects.order_by('-application_date')[:10],
        
        # Property type breakdown
        'property_by_type': total_properties.values('property_type__name').annotate(count=Count('id')).order_by('-count'),
        'property_by_subcounty': total_properties.values('sub_county__name').annotate(count=Count('id')).order_by('-count'),
        
        # Subdivisions and amalgamations
        'recent_subdivisions': PropertySubdivision.objects.order_by('-subdivision_date')[:5],
        'recent_amalgamations': PropertyAmalgamation.objects.order_by('-amalgamation_date')[:5],
    }
    
    return render(request, 'dashboards/lands_dashboard.html', context)


# ============================================================================
# GENERAL DASHBOARD (FOR USERS WITHOUT SPECIFIC ROLES)
# ============================================================================

@login_required(login_url='login')
def general_dashboard(request):
    """General Dashboard for users without specific roles"""
    context = {
        'user': request.user,
        'department': request.user.department,
        'sub_county': request.user.sub_county,
        'recent_notifications': Notification.objects.filter(
            recipient=request.user
        ).order_by('-created_at')[:10],
    }
    
    return render(request, 'dashboards/general_dashboard.html', context)


"""
Citizen Management Views
Handles CRUD operations, search, filtering, and export
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Count, Sum
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from main_application.models import (
    Citizen, SubCounty, Ward, User, Bill, Payment,
    CitizenDocument
)


@login_required
def citizen_list(request):
    """
    List all citizens with search, filter, and pagination
    """
    # Base queryset
    citizens = Citizen.objects.select_related(
        'sub_county', 'ward', 'created_by'
    ).prefetch_related('bills', 'payments')
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        citizens = citizens.filter(
            Q(unique_identifier__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(business_name__icontains=search_query) |
            Q(phone_primary__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(registration_number__icontains=search_query)
        )
    
    # Filters
    entity_type = request.GET.get('entity_type', '')
    sub_county_id = request.GET.get('sub_county', '')
    ward_id = request.GET.get('ward', '')
    status = request.GET.get('status', '')
    has_portal = request.GET.get('has_portal', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if entity_type:
        citizens = citizens.filter(entity_type=entity_type)
    
    if sub_county_id:
        citizens = citizens.filter(sub_county_id=sub_county_id)
    
    if ward_id:
        citizens = citizens.filter(ward_id=ward_id)
    
    if status:
        is_active = status == 'active'
        citizens = citizens.filter(is_active=is_active)
    
    if has_portal:
        has_access = has_portal == 'true'
        citizens = citizens.filter(has_portal_access=has_access)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            citizens = citizens.filter(created_at__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            citizens = citizens.filter(created_at__date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Order by most recent
    citizens = citizens.order_by('-created_at')
    
    # Statistics
    total_citizens = Citizen.objects.count()
    active_citizens = Citizen.objects.filter(is_active=True).count()
    individuals = Citizen.objects.filter(entity_type='individual').count()
    businesses = Citizen.objects.filter(entity_type='business').count()
    portal_users = Citizen.objects.filter(has_portal_access=True).count()
    
    # Recent registrations (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_registrations = Citizen.objects.filter(
        created_at__gte=thirty_days_ago
    ).count()
    
    stats = {
        'total': total_citizens,
        'active': active_citizens,
        'individuals': individuals,
        'businesses': businesses,
        'portal_users': portal_users,
        'recent_registrations': recent_registrations,
    }
    
    # Pagination
    paginator = Paginator(citizens, 25)  # 25 citizens per page
    page = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Get filter options
    sub_counties = SubCounty.objects.filter(is_active=True).order_by('name')
    wards = Ward.objects.filter(is_active=True).select_related('sub_county').order_by('name')
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'sub_counties': sub_counties,
        'wards': wards,
        'current_entity_type': entity_type,
        'current_sub_county': sub_county_id,
        'current_ward': ward_id,
        'current_status': status,
        'current_has_portal': has_portal,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'admin/citizens/citizen_list.html', context)


@login_required
def citizen_detail(request, unique_identifier):
    """
    View detailed information about a specific citizen
    Uses unique_identifier (ID number or business registration number)
    """
    citizen = get_object_or_404(
        Citizen.objects.select_related('sub_county', 'ward', 'created_by', 'portal_user'),
        unique_identifier=unique_identifier
    )
    
    # Get related records
    bills = citizen.bills.select_related('revenue_stream').order_by('-created_at')[:10]
    payments = citizen.payments.select_related('payment_method', 'revenue_stream').order_by('-created_at')[:10]
    documents = citizen.documents.select_related('uploaded_by').order_by('-uploaded_at')
    
    # Financial summary
    total_billed = citizen.bills.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_paid = citizen.payments.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
    outstanding_balance = citizen.bills.filter(
        status__in=['issued', 'partially_paid', 'overdue']
    ).aggregate(Sum('balance'))['balance__sum'] or 0
    
    # Bill statistics
    bill_stats = {
        'total': citizen.bills.count(),
        'paid': citizen.bills.filter(status='paid').count(),
        'pending': citizen.bills.filter(status__in=['issued', 'partially_paid']).count(),
        'overdue': citizen.bills.filter(status='overdue').count(),
    }
    
    context = {
        'citizen': citizen,
        'bills': bills,
        'payments': payments,
        'documents': documents,
        'total_billed': total_billed,
        'total_paid': total_paid,
        'outstanding_balance': outstanding_balance,
        'bill_stats': bill_stats,
    }
    
    return render(request, 'admin/citizens/citizen_detail.html', context)


@login_required
def citizen_create(request):
    """
    Create a new citizen record
    """
    if request.method == 'POST':
        entity_type = request.POST.get('entity_type')
        unique_identifier = request.POST.get('unique_identifier')
        
        # Check if unique identifier already exists
        if Citizen.objects.filter(unique_identifier=unique_identifier).exists():
            messages.error(request, f'A citizen with ID/Registration number {unique_identifier} already exists.')
            return render(request, 'admin/citizens/citizen_create.html', {
                'sub_counties': SubCounty.objects.filter(is_active=True),
                'wards': Ward.objects.filter(is_active=True),
                'post_data': request.POST,
            })
        
        # Check for duplicate phone
        phone_primary = request.POST.get('phone_primary')
        if Citizen.objects.filter(phone_primary=phone_primary).exists():
            messages.error(request, f'A citizen with phone number {phone_primary} already exists.')
            return render(request, 'admin/citizens/citizen_create.html', {
                'sub_counties': SubCounty.objects.filter(is_active=True),
                'wards': Ward.objects.filter(is_active=True),
                'post_data': request.POST,
            })
        
        # Check for duplicate email if provided
        email = request.POST.get('email', '').strip()
        if email and Citizen.objects.filter(email=email).exists():
            messages.error(request, f'A citizen with email {email} already exists.')
            return render(request, 'admin/citizens/citizen_create.html', {
                'sub_counties': SubCounty.objects.filter(is_active=True),
                'wards': Ward.objects.filter(is_active=True),
                'post_data': request.POST,
            })
        
        try:
            # Create citizen based on entity type
            citizen_data = {
                'entity_type': entity_type,
                'unique_identifier': unique_identifier,
                'phone_primary': phone_primary,
                'email': email if email else None,
                'phone_secondary': request.POST.get('phone_secondary', ''),
                'postal_address': request.POST.get('postal_address', ''),
                'physical_address': request.POST.get('physical_address', ''),
                'sub_county_id': request.POST.get('sub_county'),
                'ward_id': request.POST.get('ward'),
                'created_by': request.user,
                'is_active': True,
            }
            
            if entity_type == 'individual':
                citizen_data.update({
                    'first_name': request.POST.get('first_name'),
                    'middle_name': request.POST.get('middle_name', ''),
                    'last_name': request.POST.get('last_name'),
                    'gender': request.POST.get('gender'),
                })
                
                # Parse date of birth
                dob = request.POST.get('date_of_birth')
                if dob:
                    try:
                        citizen_data['date_of_birth'] = datetime.strptime(dob, '%Y-%m-%d').date()
                    except ValueError:
                        pass
            
            else:  # business or organization
                citizen_data.update({
                    'business_name': request.POST.get('business_name'),
                    'registration_number': request.POST.get('registration_number', ''),
                })
            
            citizen = Citizen.objects.create(**citizen_data)
            
            messages.success(
                request,
                f'Citizen {citizen} registered successfully with ID: {unique_identifier}'
            )
            return redirect('citizen_detail', unique_identifier=citizen.unique_identifier)
            
        except Exception as e:
            messages.error(request, f'Error creating citizen: {str(e)}')
    
    context = {
        'sub_counties': SubCounty.objects.filter(is_active=True).order_by('name'),
        'wards': Ward.objects.filter(is_active=True).select_related('sub_county').order_by('name'),
    }
    
    return render(request, 'admin/citizens/citizen_create.html', context)


@login_required
def citizen_update(request, unique_identifier):
    """
    Update citizen information
    """
    citizen = get_object_or_404(Citizen, unique_identifier=unique_identifier)
    
    if request.method == 'POST':
        try:
            # Update common fields
            citizen.phone_primary = request.POST.get('phone_primary')
            citizen.phone_secondary = request.POST.get('phone_secondary', '')
            citizen.email = request.POST.get('email', '') or None
            citizen.postal_address = request.POST.get('postal_address', '')
            citizen.physical_address = request.POST.get('physical_address', '')
            citizen.sub_county_id = request.POST.get('sub_county')
            citizen.ward_id = request.POST.get('ward')
            
            # Update type-specific fields
            if citizen.entity_type == 'individual':
                citizen.first_name = request.POST.get('first_name')
                citizen.middle_name = request.POST.get('middle_name', '')
                citizen.last_name = request.POST.get('last_name')
                citizen.gender = request.POST.get('gender')
                
                dob = request.POST.get('date_of_birth')
                if dob:
                    try:
                        citizen.date_of_birth = datetime.strptime(dob, '%Y-%m-%d').date()
                    except ValueError:
                        pass
            else:
                citizen.business_name = request.POST.get('business_name')
                citizen.registration_number = request.POST.get('registration_number', '')
            
            citizen.save()
            
            messages.success(request, f'Citizen {citizen} updated successfully.')
            return redirect('citizen_detail', unique_identifier=citizen.unique_identifier)
            
        except Exception as e:
            messages.error(request, f'Error updating citizen: {str(e)}')
    
    context = {
        'citizen': citizen,
        'sub_counties': SubCounty.objects.filter(is_active=True).order_by('name'),
        'wards': Ward.objects.filter(is_active=True).select_related('sub_county').order_by('name'),
    }
    
    return render(request, 'admin/citizens/citizen_update.html', context)


@login_required
def citizen_delete(request, unique_identifier):
    """
    Soft delete a citizen (mark as inactive)
    """
    if request.method == 'POST':
        citizen = get_object_or_404(Citizen, unique_identifier=unique_identifier)
        
        # Check if citizen has active bills
        active_bills = citizen.bills.filter(
            status__in=['issued', 'partially_paid']
        ).count()
        
        if active_bills > 0:
            messages.error(
                request,
                f'Cannot delete citizen with {active_bills} active bill(s). Please settle all bills first.'
            )
            return redirect('citizen_detail', unique_identifier=unique_identifier)
        
        # Soft delete
        citizen.is_active = False
        citizen.save()
        
        messages.success(request, f'Citizen {citizen} has been deactivated.')
        return redirect('citizen_list')
    
    return redirect('citizen_list')


@login_required
def citizen_export_excel(request):
    """
    Export filtered citizens to Excel
    """
    # Apply same filters as list view
    citizens = Citizen.objects.select_related('sub_county', 'ward', 'created_by')
    
    # Search
    search_query = request.GET.get('search', '').strip()
    if search_query:
        citizens = citizens.filter(
            Q(unique_identifier__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(business_name__icontains=search_query) |
            Q(phone_primary__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Filters
    entity_type = request.GET.get('entity_type', '')
    sub_county_id = request.GET.get('sub_county', '')
    ward_id = request.GET.get('ward', '')
    status = request.GET.get('status', '')
    has_portal = request.GET.get('has_portal', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if entity_type:
        citizens = citizens.filter(entity_type=entity_type)
    if sub_county_id:
        citizens = citizens.filter(sub_county_id=sub_county_id)
    if ward_id:
        citizens = citizens.filter(ward_id=ward_id)
    if status:
        is_active = status == 'active'
        citizens = citizens.filter(is_active=is_active)
    if has_portal:
        has_access = has_portal == 'true'
        citizens = citizens.filter(has_portal_access=has_access)
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            citizens = citizens.filter(created_at__date__gte=date_from_obj)
        except ValueError:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            citizens = citizens.filter(created_at__date__lte=date_to_obj)
        except ValueError:
            pass
    
    citizens = citizens.order_by('-created_at')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Citizens Export"
    
    # Styling
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'ID/Registration #',
        'Entity Type',
        'Full Name/Business Name',
        'Gender',
        'Date of Birth',
        'Phone Primary',
        'Phone Secondary',
        'Email',
        'Sub-County',
        'Ward',
        'Physical Address',
        'Postal Address',
        'Portal Access',
        'Status',
        'Registration Date',
        'Created By',
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_num, citizen in enumerate(citizens, 2):
        if citizen.entity_type == 'individual':
            full_name = f"{citizen.first_name} {citizen.middle_name} {citizen.last_name}".strip()
            gender = citizen.gender or ''
            dob = citizen.date_of_birth.strftime('%Y-%m-%d') if citizen.date_of_birth else ''
        else:
            full_name = citizen.business_name or ''
            gender = 'N/A'
            dob = 'N/A'
        
        row_data = [
            citizen.unique_identifier,
            citizen.get_entity_type_display(),
            full_name,
            gender,
            dob,
            citizen.phone_primary,
            citizen.phone_secondary or '',
            citizen.email or '',
            citizen.sub_county.name if citizen.sub_county else '',
            citizen.ward.name if citizen.ward else '',
            citizen.physical_address or '',
            citizen.postal_address or '',
            'Yes' if citizen.has_portal_access else 'No',
            'Active' if citizen.is_active else 'Inactive',
            citizen.created_at.strftime('%Y-%m-%d %H:%M'),
            citizen.created_by.get_full_name() if citizen.created_by else '',
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Alignment
            if col_num in [1, 5, 6, 15]:  # ID, phones, date
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='left', wrap_text=True)
    
    # Adjust column widths
    column_widths = {
        'A': 20,  # ID
        'B': 15,  # Entity Type
        'C': 30,  # Name
        'D': 10,  # Gender
        'E': 15,  # DOB
        'F': 15,  # Phone 1
        'G': 15,  # Phone 2
        'H': 25,  # Email
        'I': 20,  # Sub-County
        'J': 20,  # Ward
        'K': 35,  # Physical Address
        'L': 20,  # Postal Address
        'M': 12,  # Portal
        'N': 10,  # Status
        'O': 18,  # Reg Date
        'P': 20,  # Created By
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f'citizens_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def get_wards_by_subcounty(request):
    """
    AJAX endpoint to get wards by sub-county
    """
    sub_county_id = request.GET.get('sub_county_id')
    
    if not sub_county_id:
        return JsonResponse({'wards': []})
    
    wards = Ward.objects.filter(
        sub_county_id=sub_county_id,
        is_active=True
    ).values('id', 'name').order_by('name')
    
    return JsonResponse({'wards': list(wards)})


"""
Wajir County Analytics Dashboard View
Dynamic analytics with filtering, graphs, and export capabilities
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Avg, Q, F
from django.db.models.functions import TruncMonth, TruncWeek, TruncDay, TruncYear
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
import json
import csv
from io import BytesIO
import xlsxwriter

# Import your models
from .models import (
    Payment, Bill, Citizen, RevenueStream, License, 
    Patient, Visit, FleetVehicle, FuelTransaction,
    SubCounty, Ward, Property, Business, Fine,
    Attendance, LeaveApplication, Asset, Store,
    HealthFacility, Department
)


@login_required
def analytics_dashboard(request):
    """
    Main analytics dashboard view with dynamic filtering
    """
    # Get filter parameters
    date_range = request.GET.get('date_range', '30')  # days, or custom
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sub_county_id = request.GET.get('sub_county')
    ward_id = request.GET.get('ward')
    revenue_stream_id = request.GET.get('revenue_stream')
    department_id = request.GET.get('department')
    
    # Calculate date range
    today = timezone.now().date()
    
    if start_date and end_date:
        # Custom date range
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        # Predefined ranges
        if date_range == '7':
            start = today - timedelta(days=7)
        elif date_range == '30':
            start = today - timedelta(days=30)
        elif date_range == '90':
            start = today - timedelta(days=90)
        elif date_range == '180':
            start = today - timedelta(days=180)
        elif date_range == '365':
            start = today - timedelta(days=365)
        elif date_range == 'ytd':
            start = datetime(today.year, 1, 1).date()
        elif date_range == 'month':
            start = datetime(today.year, today.month, 1).date()
        else:
            start = today - timedelta(days=30)
        end = today
    
    # Build base querysets with filters
    payments_qs = Payment.objects.filter(
        payment_date__date__gte=start,
        payment_date__date__lte=end,
        status='completed'
    )
    
    bills_qs = Bill.objects.filter(
        bill_date__gte=start,
        bill_date__lte=end
    )
    
    # Apply location filters
    if sub_county_id:
        payments_qs = payments_qs.filter(sub_county_id=sub_county_id)
        bills_qs = bills_qs.filter(sub_county_id=sub_county_id)
    
    if ward_id:
        payments_qs = payments_qs.filter(ward_id=ward_id)
        bills_qs = bills_qs.filter(ward_id=ward_id)
    
    if revenue_stream_id:
        payments_qs = payments_qs.filter(revenue_stream_id=revenue_stream_id)
        bills_qs = bills_qs.filter(revenue_stream_id=revenue_stream_id)
    
    # =====================================================================
    # REVENUE ANALYTICS
    # =====================================================================
    
    # Total revenue collected
    total_revenue = payments_qs.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    
    # Total billed amount
    total_billed = bills_qs.aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0.00')
    
    # Outstanding amount
    outstanding = bills_qs.filter(
        status__in=['issued', 'partially_paid', 'overdue']
    ).aggregate(
        total=Sum('balance')
    )['total'] or Decimal('0.00')
    
    # Collection rate
    collection_rate = (total_revenue / total_billed * 100) if total_billed > 0 else 0
    
    # Revenue growth (compare with previous period)
    period_days = (end - start).days
    prev_start = start - timedelta(days=period_days)
    prev_end = start - timedelta(days=1)
    
    prev_revenue = Payment.objects.filter(
        payment_date__date__gte=prev_start,
        payment_date__date__lte=prev_end,
        status='completed'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    revenue_growth = ((total_revenue - prev_revenue) / prev_revenue * 100) if prev_revenue > 0 else 0
    
    # =====================================================================
    # REVENUE TREND CHARTS
    # =====================================================================
    
    # Determine grouping based on date range
    days_diff = (end - start).days
    
    if days_diff <= 31:
        # Daily grouping
        trunc_func = TruncDay
        date_format = '%b %d'
    elif days_diff <= 90:
        # Weekly grouping
        trunc_func = TruncWeek
        date_format = 'Week %W'
    elif days_diff <= 365:
        # Monthly grouping
        trunc_func = TruncMonth
        date_format = '%b %Y'
    else:
        # Yearly grouping
        trunc_func = TruncYear
        date_format = '%Y'
    
    # Revenue trend over time
    revenue_trend = payments_qs.annotate(
        period=trunc_func('payment_date')
    ).values('period').annotate(
        revenue=Sum('amount'),
        count=Count('id')
    ).order_by('period')
    
    revenue_trend_data = {
        'labels': [item['period'].strftime(date_format) for item in revenue_trend],
        'data': [float(item['revenue']) for item in revenue_trend],
        'counts': [item['count'] for item in revenue_trend]
    }
    
    # Revenue by stream
    revenue_by_stream = payments_qs.values(
        'revenue_stream__name'
    ).annotate(
        revenue=Sum('amount')
    ).order_by('-revenue')[:10]
    
    revenue_stream_data = {
        'labels': [item['revenue_stream__name'] for item in revenue_by_stream],
        'data': [float(item['revenue']) for item in revenue_by_stream],
        'colors': ['#139145', '#C18B5A', '#CD4F27', '#3B82F6', '#8B5CF6', 
                   '#06B6D4', '#F59E0B', '#EF4444', '#10B981', '#6366F1']
    }
    
    # Revenue by sub-county
    revenue_by_subcounty = payments_qs.values(
        'sub_county__name'
    ).annotate(
        revenue=Sum('amount')
    ).order_by('-revenue')[:10]
    
    subcounty_revenue_data = {
        'labels': [item['sub_county__name'] or 'Unknown' for item in revenue_by_subcounty],
        'data': [float(item['revenue']) for item in revenue_by_subcounty]
    }
    
    # Revenue by payment method
    revenue_by_method = payments_qs.values(
        'payment_method__name'
    ).annotate(
        revenue=Sum('amount'),
        count=Count('id')
    ).order_by('-revenue')
    
    payment_method_data = {
        'labels': [item['payment_method__name'] for item in revenue_by_method],
        'data': [float(item['revenue']) for item in revenue_by_method],
        'counts': [item['count'] for item in revenue_by_method],
        'colors': ['#139145', '#C18B5A', '#CD4F27', '#3B82F6', '#8B5CF6']
    }
    
    # =====================================================================
    # BILLS ANALYTICS
    # =====================================================================
    
    bills_by_status = bills_qs.values('status').annotate(
        count=Count('id'),
        amount=Sum('total_amount')
    ).order_by('-count')
    
    bills_status_data = {
        'labels': [item['status'].replace('_', ' ').title() for item in bills_by_status],
        'data': [item['count'] for item in bills_by_status],
        'amounts': [float(item['amount']) for item in bills_by_status],
        'colors': ['#139145', '#C18B5A', '#CD4F27', '#3B82F6', '#8B5CF6', '#06B6D4']
    }
    
    # Bill statistics
    total_bills = bills_qs.count()
    paid_bills = bills_qs.filter(status='paid').count()
    overdue_bills = bills_qs.filter(status='overdue').count()
    pending_bills = bills_qs.filter(status__in=['issued', 'partially_paid']).count()
    
    # =====================================================================
    # CITIZEN ANALYTICS
    # =====================================================================
    
    total_citizens = Citizen.objects.filter(is_active=True).count()
    new_citizens = Citizen.objects.filter(
        created_at__date__gte=start,
        created_at__date__lte=end
    ).count()
    
    citizens_by_type = Citizen.objects.values('entity_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    citizen_type_data = {
        'labels': [item['entity_type'].title() for item in citizens_by_type],
        'data': [item['count'] for item in citizens_by_type],
        'colors': ['#139145', '#C18B5A', '#CD4F27', '#3B82F6']
    }
    
    # =====================================================================
    # HEALTH SERVICES ANALYTICS
    # =====================================================================
    
    total_patients = Patient.objects.filter(is_active=True).count()
    
    visits_in_period = Visit.objects.filter(
        visit_date__date__gte=start,
        visit_date__date__lte=end
    )
    
    total_visits = visits_in_period.count()
    
    # Visits trend
    visits_trend = visits_in_period.annotate(
        period=trunc_func('visit_date')
    ).values('period').annotate(
        count=Count('id')
    ).order_by('period')
    
    visits_trend_data = {
        'labels': [item['period'].strftime(date_format) for item in visits_trend],
        'data': [item['count'] for item in visits_trend]
    }
    
    # Visits by facility
    visits_by_facility = visits_in_period.values(
        'facility__name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    facility_visits_data = {
        'labels': [item['facility__name'] for item in visits_by_facility],
        'data': [item['count'] for item in visits_by_facility]
    }
    
    # Visits by type
    visits_by_type = visits_in_period.values('visit_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    visit_type_data = {
        'labels': [item['visit_type'].title() for item in visits_by_type],
        'data': [item['count'] for item in visits_by_type],
        'colors': ['#139145', '#C18B5A', '#CD4F27']
    }
    
    # =====================================================================
    # LICENSE & BUSINESS ANALYTICS
    # =====================================================================
    
    total_licenses = License.objects.filter(status='active').count()
    expiring_licenses = License.objects.filter(
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30),
        status='active'
    ).count()
    
    licenses_issued = License.objects.filter(
        issue_date__gte=start,
        issue_date__lte=end
    ).count()
    
    licenses_by_status = License.objects.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    license_status_data = {
        'labels': [item['status'].replace('_', ' ').title() for item in licenses_by_status],
        'data': [item['count'] for item in licenses_by_status],
        'colors': ['#139145', '#C18B5A', '#CD4F27', '#3B82F6', '#8B5CF6']
    }
    
    # Businesses by category
    businesses_by_category = Business.objects.filter(
        is_active=True
    ).values(
        'business_category__name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    business_category_data = {
        'labels': [item['business_category__name'] for item in businesses_by_category],
        'data': [item['count'] for item in businesses_by_category]
    }
    
    # =====================================================================
    # FLEET & FUEL ANALYTICS
    # =====================================================================
    
    active_vehicles = FleetVehicle.objects.filter(status='active').count()
    
    fuel_transactions = FuelTransaction.objects.filter(
        transaction_date__date__gte=start,
        transaction_date__date__lte=end
    )
    
    total_fuel_cost = fuel_transactions.aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0.00')
    
    total_fuel_liters = fuel_transactions.aggregate(
        total=Sum('quantity_liters')
    )['total'] or Decimal('0.00')
    
    # Fuel consumption by vehicle
    fuel_by_vehicle = fuel_transactions.values(
        'vehicle__registration_number'
    ).annotate(
        cost=Sum('total_amount'),
        liters=Sum('quantity_liters')
    ).order_by('-cost')[:10]
    
    fuel_vehicle_data = {
        'labels': [item['vehicle__registration_number'] for item in fuel_by_vehicle],
        'data': [float(item['cost']) for item in fuel_by_vehicle],
        'liters': [float(item['liters']) for item in fuel_by_vehicle]
    }
    
    # Fuel trend
    fuel_trend = fuel_transactions.annotate(
        period=trunc_func('transaction_date')
    ).values('period').annotate(
        cost=Sum('total_amount'),
        liters=Sum('quantity_liters')
    ).order_by('period')
    
    fuel_trend_data = {
        'labels': [item['period'].strftime(date_format) for item in fuel_trend],
        'data': [float(item['cost']) for item in fuel_trend],
        'liters': [float(item['liters']) for item in fuel_trend]
    }
    
    # =====================================================================
    # HR ANALYTICS
    # =====================================================================
    
    total_employees = User.objects.filter(is_active=True, is_staff=True).count()
    
    # Attendance for the period
    attendance_records = Attendance.objects.filter(
        attendance_date__gte=start,
        attendance_date__lte=end
    )
    
    total_attendance = attendance_records.values('employee').distinct().count()
    
    # Leave applications
    leave_applications = LeaveApplication.objects.filter(
        start_date__gte=start,
        start_date__lte=end
    )
    
    leaves_by_status = leave_applications.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    leave_status_data = {
        'labels': [item['status'].title() for item in leaves_by_status],
        'data': [item['count'] for item in leaves_by_status],
        'colors': ['#139145', '#C18B5A', '#CD4F27', '#3B82F6']
    }
    
    # =====================================================================
    # PROPERTY & FINES ANALYTICS
    # =====================================================================
    
    total_properties = Property.objects.filter(status='active').count()
    
    fines_in_period = Fine.objects.filter(
        issued_date__gte=start,
        issued_date__lte=end
    )
    
    total_fines_amount = fines_in_period.aggregate(
        total=Sum('fine_amount')
    )['total'] or Decimal('0.00')
    
    fines_collected = fines_in_period.filter(
        status='paid'
    ).aggregate(
        total=Sum('fine_amount')
    )['total'] or Decimal('0.00')
    
    # Fines by category
    fines_by_category = fines_in_period.values(
        'category__name'
    ).annotate(
        count=Count('id'),
        amount=Sum('fine_amount')
    ).order_by('-amount')[:10]
    
    fines_category_data = {
        'labels': [item['category__name'] for item in fines_by_category],
        'data': [float(item['amount']) for item in fines_by_category],
        'counts': [item['count'] for item in fines_by_category]
    }
    
    # =====================================================================
    # PERFORMANCE METRICS
    # =====================================================================
    
    # Revenue per citizen
    revenue_per_citizen = total_revenue / total_citizens if total_citizens > 0 else 0
    
    # Average bill amount
    avg_bill_amount = bills_qs.aggregate(avg=Avg('total_amount'))['avg'] or Decimal('0.00')
    
    # Average payment amount
    avg_payment_amount = payments_qs.aggregate(avg=Avg('amount'))['avg'] or Decimal('0.00')
    
    # Top revenue streams
    top_revenue_streams = payments_qs.values(
        'revenue_stream__name',
        'revenue_stream__code'
    ).annotate(
        revenue=Sum('amount'),
        count=Count('id')
    ).order_by('-revenue')[:5]
    
    # =====================================================================
    # FILTER OPTIONS FOR DROPDOWNS
    # =====================================================================
    
    sub_counties = SubCounty.objects.filter(is_active=True).order_by('name')
    wards = Ward.objects.filter(is_active=True).order_by('name')
    if sub_county_id:
        wards = wards.filter(sub_county_id=sub_county_id)
    
    revenue_streams = RevenueStream.objects.filter(is_active=True).order_by('name')
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    # =====================================================================
    # CONTEXT DATA
    # =====================================================================
    
    context = {
        # Filter parameters
        'date_range': date_range,
        'start_date': start.isoformat(),
        'end_date': end.isoformat(),
        'selected_sub_county': sub_county_id,
        'selected_ward': ward_id,
        'selected_revenue_stream': revenue_stream_id,
        'selected_department': department_id,
        
        # Filter options
        'sub_counties': sub_counties,
        'wards': wards,
        'revenue_streams': revenue_streams,
        'departments': departments,
        
        # Summary metrics
        'total_revenue': total_revenue,
        'total_billed': total_billed,
        'outstanding': outstanding,
        'collection_rate': round(collection_rate, 2),
        'revenue_growth': round(revenue_growth, 2),
        'total_bills': total_bills,
        'paid_bills': paid_bills,
        'overdue_bills': overdue_bills,
        'pending_bills': pending_bills,
        'total_citizens': total_citizens,
        'new_citizens': new_citizens,
        'total_licenses': total_licenses,
        'expiring_licenses': expiring_licenses,
        'licenses_issued': licenses_issued,
        'total_patients': total_patients,
        'total_visits': total_visits,
        'active_vehicles': active_vehicles,
        'total_fuel_cost': total_fuel_cost,
        'total_fuel_liters': total_fuel_liters,
        'total_employees': total_employees,
        'total_attendance': total_attendance,
        'total_properties': total_properties,
        'total_fines_amount': total_fines_amount,
        'fines_collected': fines_collected,
        'revenue_per_citizen': round(revenue_per_citizen, 2),
        'avg_bill_amount': avg_bill_amount,
        'avg_payment_amount': avg_payment_amount,
        
        # Top performers
        'top_revenue_streams': top_revenue_streams,
        
        # Chart data (JSON serialized)
        'revenue_trend_data': json.dumps(revenue_trend_data),
        'revenue_stream_data': json.dumps(revenue_stream_data),
        'subcounty_revenue_data': json.dumps(subcounty_revenue_data),
        'payment_method_data': json.dumps(payment_method_data),
        'bills_status_data': json.dumps(bills_status_data),
        'citizen_type_data': json.dumps(citizen_type_data),
        'visits_trend_data': json.dumps(visits_trend_data),
        'facility_visits_data': json.dumps(facility_visits_data),
        'visit_type_data': json.dumps(visit_type_data),
        'license_status_data': json.dumps(license_status_data),
        'business_category_data': json.dumps(business_category_data),
        'fuel_vehicle_data': json.dumps(fuel_vehicle_data),
        'fuel_trend_data': json.dumps(fuel_trend_data),
        'leave_status_data': json.dumps(leave_status_data),
        'fines_category_data': json.dumps(fines_category_data),
    }
    
    return render(request, 'admin/analytics/analytics_dashboard.html', context)


@login_required
def export_analytics_data(request):
    """
    Export analytics data to CSV or Excel
    """
    export_format = request.GET.get('format', 'csv')
    report_type = request.GET.get('type', 'revenue')
    
    # Get same filters as dashboard
    date_range = request.GET.get('date_range', '30')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Calculate dates (same logic as above)
    today = timezone.now().date()
    if start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        if date_range == '7':
            start = today - timedelta(days=7)
        elif date_range == '30':
            start = today - timedelta(days=30)
        elif date_range == '90':
            start = today - timedelta(days=90)
        else:
            start = today - timedelta(days=30)
        end = today
    
    if export_format == 'csv':
        return export_to_csv(report_type, start, end)
    else:
        return export_to_excel(report_type, start, end)


def export_to_csv(report_type, start_date, end_date):
    """Export data to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="wajir_{report_type}_{start_date}_{end_date}.csv"'
    
    writer = csv.writer(response)
    
    if report_type == 'revenue':
        # Revenue report
        writer.writerow(['Receipt Number', 'Date', 'Payer', 'Revenue Stream', 'Amount', 'Method', 'Status'])
        
        payments = Payment.objects.filter(
            payment_date__date__gte=start_date,
            payment_date__date__lte=end_date
        ).select_related('revenue_stream', 'payment_method')
        
        for payment in payments:
            writer.writerow([
                payment.receipt_number,
                payment.payment_date.strftime('%Y-%m-%d %H:%M'),
                payment.payer_name,
                payment.revenue_stream.name,
                payment.amount,
                payment.payment_method.name,
                payment.get_status_display()
            ])
    
    elif report_type == 'bills':
        # Bills report
        writer.writerow(['Bill Number', 'Date', 'Citizen', 'Revenue Stream', 'Amount', 'Paid', 'Balance', 'Status'])
        
        bills = Bill.objects.filter(
            bill_date__gte=start_date,
            bill_date__lte=end_date
        ).select_related('citizen', 'revenue_stream')
        
        for bill in bills:
            citizen_name = f"{bill.citizen.first_name} {bill.citizen.last_name}" if bill.citizen.entity_type == 'individual' else bill.citizen.business_name
            writer.writerow([
                bill.bill_number,
                bill.bill_date.strftime('%Y-%m-%d'),
                citizen_name,
                bill.revenue_stream.name,
                bill.total_amount,
                bill.amount_paid,
                bill.balance,
                bill.get_status_display()
            ])
    
    return response


def export_to_excel(report_type, start_date, end_date):
    """Export data to Excel"""
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet(report_type.title())
    
    # Formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#139145',
        'font_color': 'white',
        'border': 1
    })
    
    money_format = workbook.add_format({'num_format': '#,##0.00'})
    
    if report_type == 'revenue':
        # Headers
        headers = ['Receipt Number', 'Date', 'Payer', 'Revenue Stream', 'Amount', 'Method', 'Status']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        # Data
        payments = Payment.objects.filter(
            payment_date__date__gte=start_date,
            payment_date__date__lte=end_date
        ).select_related('revenue_stream', 'payment_method')
        
        for row, payment in enumerate(payments, start=1):
            worksheet.write(row, 0, payment.receipt_number)
            worksheet.write(row, 1, payment.payment_date.strftime('%Y-%m-%d %H:%M'))
            worksheet.write(row, 2, payment.payer_name)
            worksheet.write(row, 3, payment.revenue_stream.name)
            worksheet.write(row, 4, float(payment.amount), money_format)
            worksheet.write(row, 5, payment.payment_method.name)
            worksheet.write(row, 6, payment.get_status_display())
    
    workbook.close()
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="wajir_{report_type}_{start_date}_{end_date}.xlsx"'
    
    return response

# Optional API View for dynamic ward loading
from django.http import JsonResponse
from .models import Ward

def get_wards_api(request):
    """
    API endpoint to get wards by sub-county
    """
    sub_county_id = request.GET.get('sub_county')
    
    if sub_county_id:
        wards = Ward.objects.filter(
            sub_county_id=sub_county_id,
            is_active=True
        ).values('id', 'name').order_by('name')
        return JsonResponse(list(wards), safe=False)
    
    return JsonResponse([], safe=False)

"""
Revenue Management Views
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, Avg
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    RevenueStream, ChargeRate, RevenueBudget, Department,
    SubCounty, Ward, Payment, Bill, User
)


# ============================================================================
# REVENUE STREAMS
# ============================================================================

@login_required
def revenue_stream_list(request):
    """Revenue streams list with filtering and search"""
    streams = RevenueStream.objects.all().select_related('department', 'parent_stream')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        streams = streams.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(department__name__icontains=search_query)
        )
    
    # Filters
    department_id = request.GET.get('department')
    if department_id:
        streams = streams.filter(department_id=department_id)
    
    is_recurring = request.GET.get('is_recurring')
    if is_recurring:
        streams = streams.filter(is_recurring=is_recurring == 'true')
    
    status = request.GET.get('status')
    if status:
        streams = streams.filter(is_active=status == 'active')
    
    # Statistics
    stats = {
        'total': RevenueStream.objects.count(),
        'active': RevenueStream.objects.filter(is_active=True).count(),
        'recurring': RevenueStream.objects.filter(is_recurring=True).count(),
        'non_recurring': RevenueStream.objects.filter(is_recurring=False).count(),
    }
    
    # Get departments for filter
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    context = {
        'streams': streams,
        'stats': stats,
        'departments': departments,
        'search_query': search_query,
        'current_department': department_id,
        'current_recurring': is_recurring,
        'current_status': status,
    }
    
    return render(request, 'revenue/stream_list.html', context)


@login_required
def revenue_stream_detail(request, stream_id):
    """Revenue stream detail view"""
    stream = get_object_or_404(RevenueStream, id=stream_id)
    
    # Get related data
    charge_rates = stream.charge_rates.all().order_by('-effective_from')
    penalty_rules = stream.penalty_rules.all().order_by('-effective_from')
    budgets = stream.budgets.all().order_by('-period_start')
    
    # Collection statistics (last 12 months)
    twelve_months_ago = timezone.now().date() - timedelta(days=365)
    collections = Payment.objects.filter(
        revenue_stream=stream,
        payment_date__gte=twelve_months_ago,
        status='completed'
    ).aggregate(
        total_collected=Sum('amount'),
        transaction_count=Count('id'),
        avg_transaction=Avg('amount')
    )
    
    # Bills statistics
    bills_stats = Bill.objects.filter(revenue_stream=stream).aggregate(
        total_bills=Count('id'),
        total_billed=Sum('bill_amount'),
        total_paid=Sum('amount_paid'),
        outstanding=Sum('balance')
    )
    
    context = {
        'stream': stream,
        'charge_rates': charge_rates,
        'penalty_rules': penalty_rules,
        'budgets': budgets,
        'collections': collections,
        'bills_stats': bills_stats,
    }
    
    return render(request, 'revenue/stream_detail.html', context)


@login_required
def revenue_stream_export(request):
    """Export revenue streams to Excel"""
    streams = RevenueStream.objects.all().select_related('department', 'parent_stream')
    
    # Apply same filters as list view
    search_query = request.GET.get('search', '')
    if search_query:
        streams = streams.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    department_id = request.GET.get('department')
    if department_id:
        streams = streams.filter(department_id=department_id)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue Streams"
    
    # Styling
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Code', 'Name', 'Department', 'Description', 'Is Recurring', 
               'Billing Frequency', 'Status', 'Created At']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for stream in streams:
        ws.append([
            stream.code,
            stream.name,
            stream.department.name,
            stream.description,
            'Yes' if stream.is_recurring else 'No',
            stream.billing_frequency or 'N/A',
            'Active' if stream.is_active else 'Inactive',
            stream.created_at.strftime('%Y-%m-%d %H:%M')
        ])
    
    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Adjust column widths
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=revenue_streams_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response


# ============================================================================
# CHARGE RATES
# ============================================================================

@login_required
def charge_rate_list(request):
    """Charge rates list with filtering"""
    rates = ChargeRate.objects.all().select_related('revenue_stream')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        rates = rates.filter(
            Q(name__icontains=search_query) |
            Q(revenue_stream__name__icontains=search_query) |
            Q(revenue_stream__code__icontains=search_query)
        )
    
    # Filters
    stream_id = request.GET.get('revenue_stream')
    if stream_id:
        rates = rates.filter(revenue_stream_id=stream_id)
    
    rate_type = request.GET.get('rate_type')
    if rate_type:
        rates = rates.filter(rate_type=rate_type)
    
    status = request.GET.get('status')
    if status == 'active':
        today = timezone.now().date()
        rates = rates.filter(
            is_active=True,
            effective_from__lte=today
        ).filter(
            Q(effective_to__isnull=True) | Q(effective_to__gte=today)
        )
    elif status == 'expired':
        today = timezone.now().date()
        rates = rates.filter(effective_to__lt=today)
    elif status == 'future':
        today = timezone.now().date()
        rates = rates.filter(effective_from__gt=today)
    
    # Statistics
    today = timezone.now().date()
    stats = {
        'total': ChargeRate.objects.count(),
        'active': ChargeRate.objects.filter(
            is_active=True,
            effective_from__lte=today
        ).filter(
            Q(effective_to__isnull=True) | Q(effective_to__gte=today)
        ).count(),
        'expired': ChargeRate.objects.filter(effective_to__lt=today).count(),
        'future': ChargeRate.objects.filter(effective_from__gt=today).count(),
    }
    
    # Get revenue streams for filter
    revenue_streams = RevenueStream.objects.filter(is_active=True).order_by('name')
    
    # Get unique rate types
    rate_types = ChargeRate.objects.values_list('rate_type', flat=True).distinct()
    
    context = {
        'rates': rates,
        'stats': stats,
        'revenue_streams': revenue_streams,
        'rate_types': rate_types,
        'search_query': search_query,
        'current_stream': stream_id,
        'current_rate_type': rate_type,
        'current_status': status,
    }
    
    return render(request, 'revenue/charge_rate_list.html', context)


@login_required
def charge_rate_export(request):
    """Export charge rates to Excel"""
    rates = ChargeRate.objects.all().select_related('revenue_stream')
    
    # Apply filters
    stream_id = request.GET.get('revenue_stream')
    if stream_id:
        rates = rates.filter(revenue_stream_id=stream_id)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Charge Rates"
    
    # Styling
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Revenue Stream', 'Rate Name', 'Rate Type', 'Amount', 
               'Min Amount', 'Max Amount', 'Effective From', 'Effective To', 'Status']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    today = timezone.now().date()
    for rate in rates:
        is_active = (
            rate.is_active and 
            rate.effective_from <= today and 
            (rate.effective_to is None or rate.effective_to >= today)
        )
        
        ws.append([
            f"{rate.revenue_stream.code} - {rate.revenue_stream.name}",
            rate.name,
            rate.rate_type,
            float(rate.amount),
            float(rate.min_amount) if rate.min_amount else 'N/A',
            float(rate.max_amount) if rate.max_amount else 'N/A',
            rate.effective_from.strftime('%Y-%m-%d'),
            rate.effective_to.strftime('%Y-%m-%d') if rate.effective_to else 'N/A',
            'Active' if is_active else 'Inactive'
        ])
    
    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    # Adjust column widths
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=charge_rates_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response


# ============================================================================
# REVENUE BUDGETS
# ============================================================================

@login_required
def revenue_budget_list(request):
    """Revenue budgets list with filtering"""
    budgets = RevenueBudget.objects.all().select_related(
        'revenue_stream', 'sub_county', 'ward', 'created_by'
    )
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        budgets = budgets.filter(
            Q(revenue_stream__name__icontains=search_query) |
            Q(revenue_stream__code__icontains=search_query) |
            Q(financial_year__icontains=search_query)
        )
    
    # Filters
    stream_id = request.GET.get('revenue_stream')
    if stream_id:
        budgets = budgets.filter(revenue_stream_id=stream_id)
    
    financial_year = request.GET.get('financial_year')
    if financial_year:
        budgets = budgets.filter(financial_year=financial_year)
    
    period_type = request.GET.get('period_type')
    if period_type:
        budgets = budgets.filter(period_type=period_type)
    
    sub_county_id = request.GET.get('sub_county')
    if sub_county_id:
        budgets = budgets.filter(sub_county_id=sub_county_id)
    
    # Calculate actual collections for each budget
    budgets_with_actual = []
    for budget in budgets:
        actual = Payment.objects.filter(
            revenue_stream=budget.revenue_stream,
            payment_date__gte=budget.period_start,
            payment_date__lte=budget.period_end,
            status='completed'
        )
        
        if budget.sub_county:
            actual = actual.filter(sub_county=budget.sub_county)
        if budget.ward:
            actual = actual.filter(ward=budget.ward)
        
        actual_amount = actual.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        budget.actual_amount = actual_amount
        budget.variance = actual_amount - budget.target_amount
        budget.achievement_percentage = (
            (actual_amount / budget.target_amount * 100) 
            if budget.target_amount > 0 else 0
        )
        budgets_with_actual.append(budget)
    
    # Statistics
    total_target = budgets.aggregate(Sum('target_amount'))['target_amount__sum'] or Decimal('0')
    total_actual = sum(b.actual_amount for b in budgets_with_actual)
    
    stats = {
        'total_budgets': budgets.count(),
        'total_target': total_target,
        'total_actual': total_actual,
        'total_variance': total_actual - total_target,
        'achievement_rate': (total_actual / total_target * 100) if total_target > 0 else 0,
    }
    
    # Get filter options
    revenue_streams = RevenueStream.objects.filter(is_active=True).order_by('name')
    financial_years = RevenueBudget.objects.values_list(
        'financial_year', flat=True
    ).distinct().order_by('-financial_year')
    period_types = RevenueBudget.objects.values_list(
        'period_type', flat=True
    ).distinct()
    sub_counties = SubCounty.objects.filter(is_active=True).order_by('name')
    
    context = {
        'budgets': budgets_with_actual,
        'stats': stats,
        'revenue_streams': revenue_streams,
        'financial_years': financial_years,
        'period_types': period_types,
        'sub_counties': sub_counties,
        'search_query': search_query,
        'current_stream': stream_id,
        'current_year': financial_year,
        'current_period': period_type,
        'current_sub_county': sub_county_id,
    }
    
    return render(request, 'revenue/budget_list.html', context)


@login_required
def revenue_budget_export(request):
    """Export revenue budgets to Excel"""
    budgets = RevenueBudget.objects.all().select_related(
        'revenue_stream', 'sub_county', 'ward'
    )
    
    # Apply filters
    stream_id = request.GET.get('revenue_stream')
    if stream_id:
        budgets = budgets.filter(revenue_stream_id=stream_id)
    
    financial_year = request.GET.get('financial_year')
    if financial_year:
        budgets = budgets.filter(financial_year=financial_year)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue Budgets"
    
    # Styling
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Revenue Stream', 'Financial Year', 'Period Type', 'Period Start', 
               'Period End', 'Target Amount', 'Actual Amount', 'Variance', 
               'Achievement %', 'Sub-County', 'Ward']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for budget in budgets:
        # Calculate actual
        actual = Payment.objects.filter(
            revenue_stream=budget.revenue_stream,
            payment_date__gte=budget.period_start,
            payment_date__lte=budget.period_end,
            status='completed'
        )
        
        if budget.sub_county:
            actual = actual.filter(sub_county=budget.sub_county)
        if budget.ward:
            actual = actual.filter(ward=budget.ward)
        
        actual_amount = actual.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        variance = actual_amount - budget.target_amount
        achievement = (actual_amount / budget.target_amount * 100) if budget.target_amount > 0 else 0
        
        ws.append([
            f"{budget.revenue_stream.code} - {budget.revenue_stream.name}",
            budget.financial_year,
            budget.period_type,
            budget.period_start.strftime('%Y-%m-%d'),
            budget.period_end.strftime('%Y-%m-%d'),
            float(budget.target_amount),
            float(actual_amount),
            float(variance),
            f"{achievement:.2f}%",
            budget.sub_county.name if budget.sub_county else 'County-wide',
            budget.ward.name if budget.ward else 'N/A'
        ])
    
    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    # Adjust column widths
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=revenue_budgets_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response


# ============================================================================
# COLLECTION REPORTS
# ============================================================================

@login_required
def collection_report(request):
    """Comprehensive revenue collection reports"""
    # Date filters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Default to current month if no dates provided
    if not date_from or not date_to:
        today = timezone.now().date()
        date_from = today.replace(day=1).strftime('%Y-%m-%d')
        date_to = today.strftime('%Y-%m-%d')
    
    # Parse dates
    start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
    end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # Base queryset
    payments = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date,
        status='completed'
    ).select_related('revenue_stream', 'sub_county', 'ward', 'payment_method')
    
    # Filters
    stream_id = request.GET.get('revenue_stream')
    if stream_id:
        payments = payments.filter(revenue_stream_id=stream_id)
    
    sub_county_id = request.GET.get('sub_county')
    if sub_county_id:
        payments = payments.filter(sub_county_id=sub_county_id)
    
    ward_id = request.GET.get('ward')
    if ward_id:
        payments = payments.filter(ward_id=ward_id)
    
    payment_method_id = request.GET.get('payment_method')
    if payment_method_id:
        payments = payments.filter(payment_method_id=payment_method_id)
    
    # Overall statistics
    overall_stats = payments.aggregate(
        total_amount=Sum('amount'),
        transaction_count=Count('id'),
        average_transaction=Avg('amount')
    )
    
    # By revenue stream
    by_stream = payments.values(
        'revenue_stream__code',
        'revenue_stream__name'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # By sub-county
    by_sub_county = payments.values(
        'sub_county__name'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # By payment method
    by_payment_method = payments.values(
        'payment_method__name'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Daily collections (for chart)
    daily_collections = payments.extra(
        select={'day': 'DATE(payment_date)'}
    ).values('day').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('day')
    
    # Get filter options
    revenue_streams = RevenueStream.objects.filter(is_active=True).order_by('name')
    sub_counties = SubCounty.objects.filter(is_active=True).order_by('name')
    wards = Ward.objects.filter(is_active=True).order_by('name')
    from .models import PaymentMethod
    payment_methods = PaymentMethod.objects.filter(is_active=True).order_by('name')
    
    context = {
        'overall_stats': overall_stats,
        'by_stream': by_stream,
        'by_sub_county': by_sub_county,
        'by_payment_method': by_payment_method,
        'daily_collections': daily_collections,
        'revenue_streams': revenue_streams,
        'sub_counties': sub_counties,
        'wards': wards,
        'payment_methods': payment_methods,
        'date_from': date_from,
        'date_to': date_to,
        'current_stream': stream_id,
        'current_sub_county': sub_county_id,
        'current_ward': ward_id,
        'current_payment_method': payment_method_id,
    }
    
    return render(request, 'revenue/collection_report.html', context)


@login_required
def collection_report_export(request):
    """Export collection report to Excel"""
    # Get date filters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from or not date_to:
        today = timezone.now().date()
        date_from = today.replace(day=1).strftime('%Y-%m-%d')
        date_to = today.strftime('%Y-%m-%d')
    
    start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
    end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # Get payments
    payments = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date,
        status='completed'
    ).select_related('revenue_stream', 'sub_county', 'payment_method', 'citizen')
    
    # Apply filters
    stream_id = request.GET.get('revenue_stream')
    if stream_id:
        payments = payments.filter(revenue_stream_id=stream_id)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Styling
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # Title
    ws_summary['A1'] = 'REVENUE COLLECTION REPORT'
    ws_summary['A1'].font = title_font
    ws_summary['A2'] = f'Period: {date_from} to {date_to}'
    
    # Overall stats
    ws_summary['A4'] = 'Overall Statistics'
    ws_summary['A4'].font = Font(bold=True, size=12)
    
    overall = payments.aggregate(
        total=Sum('amount'),
        count=Count('id'),
        avg=Avg('amount')
    )
    
    ws_summary['A5'] = 'Total Collections:'
    ws_summary['B5'] = float(overall['total'] or 0)
    ws_summary['A6'] = 'Total Transactions:'
    ws_summary['B6'] = overall['count'] or 0
    ws_summary['A7'] = 'Average Transaction:'
    ws_summary['B7'] = float(overall['avg'] or 0)
    
    # Detailed transactions sheet
    ws_detail = wb.create_sheet("Transactions")
    
    headers = ['Receipt No.', 'Date', 'Citizen', 'Revenue Stream', 'Amount', 
               'Payment Method', 'Sub-County', 'Status']
    ws_detail.append(headers)
    
    # Style headers
    for cell in ws_detail[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for payment in payments:
        ws_detail.append([
            payment.receipt_number,
            payment.payment_date.strftime('%Y-%m-%d %H:%M'),
            str(payment.citizen),
            f"{payment.revenue_stream.code} - {payment.revenue_stream.name}",
            float(payment.amount),
            payment.payment_method.name,
            payment.sub_county.name if payment.sub_county else 'N/A',
            payment.status
        ])
    
    # Adjust column widths
    for ws in [ws_summary, ws_detail]:
        for idx, col in enumerate(ws.columns, 1):
            ws.column_dimensions[get_column_letter(idx)].width = 20
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=collection_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response


"""
Billing Management Views
Handles all billing operations including listing, generation, and reporting
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, Avg
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Import models (adjust based on your app structure)
from .models import (
    Bill, BillLineItem, Citizen, RevenueStream, 
    SubCounty, Ward, Payment, ChargeRate
)


@login_required
def bill_list(request):
    """
    Display all bills with filtering and search functionality
    """
    bills = Bill.objects.select_related(
        'citizen', 'revenue_stream', 'sub_county', 'ward', 'created_by'
    ).prefetch_related('line_items', 'payments')
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        bills = bills.filter(
            Q(bill_number__icontains=search_query) |
            Q(citizen__first_name__icontains=search_query) |
            Q(citizen__last_name__icontains=search_query) |
            Q(citizen__business_name__icontains=search_query) |
            Q(revenue_stream__name__icontains=search_query)
        )
    
    # Filters
    status_filter = request.GET.get('status', '')
    if status_filter:
        bills = bills.filter(status=status_filter)
    
    revenue_stream_filter = request.GET.get('revenue_stream', '')
    if revenue_stream_filter:
        bills = bills.filter(revenue_stream_id=revenue_stream_filter)
    
    sub_county_filter = request.GET.get('sub_county', '')
    if sub_county_filter:
        bills = bills.filter(sub_county_id=sub_county_filter)
    
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        bills = bills.filter(bill_date__gte=date_from)
    if date_to:
        bills = bills.filter(bill_date__lte=date_to)
    
    # Order by most recent
    bills = bills.order_by('-bill_date', '-created_at')
    
    # Statistics
    stats = {
        'total': bills.count(),
        'draft': bills.filter(status='draft').count(),
        'issued': bills.filter(status='issued').count(),
        'partially_paid': bills.filter(status='partially_paid').count(),
        'paid': bills.filter(status='paid').count(),
        'overdue': bills.filter(status='overdue').count(),
        'cancelled': bills.filter(status='cancelled').count(),
        'total_amount': bills.aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        'total_paid': bills.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
        'total_balance': bills.aggregate(Sum('balance'))['balance__sum'] or 0,
    }
    
    # Get filter options
    revenue_streams = RevenueStream.objects.filter(is_active=True).order_by('name')
    sub_counties = SubCounty.objects.filter(is_active=True).order_by('name')
    
    context = {
        'bills': bills,
        'stats': stats,
        'search_query': search_query,
        'current_status': status_filter,
        'current_revenue_stream': revenue_stream_filter,
        'current_sub_county': sub_county_filter,
        'date_from': date_from,
        'date_to': date_to,
        'revenue_streams': revenue_streams,
        'sub_counties': sub_counties,
        'bill_statuses': Bill.BILL_STATUS_CHOICES,
    }
    
    return render(request, 'billing/bill_list.html', context)


@login_required
def bill_detail(request, bill_id):
    """
    Display detailed information about a specific bill
    """
    bill = get_object_or_404(
        Bill.objects.select_related(
            'citizen', 'revenue_stream', 'sub_county', 
            'ward', 'created_by'
        ).prefetch_related('line_items', 'payments'),
        id=bill_id
    )
    
    # Get payment history
    payments = bill.payments.select_related('payment_method', 'collected_by').order_by('-payment_date')
    
    context = {
        'bill': bill,
        'payments': payments,
        'line_items': bill.line_items.all(),
    }
    
    return render(request, 'billing/bill_detail.html', context)


@login_required
def generate_bills(request):
    """
    Generate new bills - form and processing
    """
    if request.method == 'POST':
        try:
            # Get form data
            citizen_id = request.POST.get('citizen_id')
            revenue_stream_id = request.POST.get('revenue_stream_id')
            bill_date = request.POST.get('bill_date')
            due_date = request.POST.get('due_date')
            description = request.POST.get('description', '')
            
            citizen = get_object_or_404(Citizen, id=citizen_id)
            revenue_stream = get_object_or_404(RevenueStream, id=revenue_stream_id)
            
            # Generate bill number
            today = timezone.now()
            bill_count = Bill.objects.filter(
                bill_date__year=today.year,
                bill_date__month=today.month
            ).count()
            bill_number = f"BL{today.strftime('%Y%m')}{str(bill_count + 1).zfill(5)}"
            
            # Calculate totals
            bill_amount = Decimal(request.POST.get('bill_amount', '0'))
            penalty_amount = Decimal(request.POST.get('penalty_amount', '0'))
            total_amount = bill_amount + penalty_amount
            
            # Create bill
            bill = Bill.objects.create(
                bill_number=bill_number,
                citizen=citizen,
                revenue_stream=revenue_stream,
                bill_date=bill_date,
                due_date=due_date,
                bill_amount=bill_amount,
                penalty_amount=penalty_amount,
                total_amount=total_amount,
                balance=total_amount,
                status='issued',
                sub_county=citizen.sub_county,
                ward=citizen.ward,
                description=description,
                created_by=request.user
            )
            
            # Create line items
            line_item_descriptions = request.POST.getlist('line_item_description[]')
            line_item_quantities = request.POST.getlist('line_item_quantity[]')
            line_item_unit_prices = request.POST.getlist('line_item_unit_price[]')
            
            for i, desc in enumerate(line_item_descriptions):
                if desc.strip():
                    quantity = Decimal(line_item_quantities[i])
                    unit_price = Decimal(line_item_unit_prices[i])
                    amount = quantity * unit_price
                    
                    BillLineItem.objects.create(
                        bill=bill,
                        description=desc,
                        quantity=quantity,
                        unit_price=unit_price,
                        amount=amount
                    )
            
            messages.success(request, f'Bill {bill_number} generated successfully!')
            return redirect('bill_detail', bill_id=bill.id)
            
        except Exception as e:
            messages.error(request, f'Error generating bill: {str(e)}')
    
    # GET request - show form
    citizens = Citizen.objects.filter(is_active=True).order_by('first_name', 'business_name')
    revenue_streams = RevenueStream.objects.filter(is_active=True).order_by('name')
    charge_rates = ChargeRate.objects.filter(is_active=True).select_related('revenue_stream')
    
    context = {
        'citizens': citizens,
        'revenue_streams': revenue_streams,
        'charge_rates': charge_rates,
    }
    
    return render(request, 'billing/generate_bills.html', context)


@login_required
def overdue_bills(request):
    """
    Display all overdue bills
    """
    today = timezone.now().date()
    
    bills = Bill.objects.select_related(
        'citizen', 'revenue_stream', 'sub_county', 'ward'
    ).filter(
        Q(status='issued') | Q(status='partially_paid'),
        due_date__lt=today
    ).prefetch_related('payments')
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        bills = bills.filter(
            Q(bill_number__icontains=search_query) |
            Q(citizen__first_name__icontains=search_query) |
            Q(citizen__last_name__icontains=search_query) |
            Q(citizen__business_name__icontains=search_query)
        )
    
    # Filters
    revenue_stream_filter = request.GET.get('revenue_stream', '')
    if revenue_stream_filter:
        bills = bills.filter(revenue_stream_id=revenue_stream_filter)
    
    sub_county_filter = request.GET.get('sub_county', '')
    if sub_county_filter:
        bills = bills.filter(sub_county_id=sub_county_filter)
    
    # Calculate days overdue and categorize
    bills_with_overdue = []
    for bill in bills:
        days_overdue = (today - bill.due_date).days
        bill.days_overdue = days_overdue
        
        # Categorize by overdue period
        if days_overdue <= 30:
            bill.overdue_category = '1-30 days'
        elif days_overdue <= 60:
            bill.overdue_category = '31-60 days'
        elif days_overdue <= 90:
            bill.overdue_category = '61-90 days'
        else:
            bill.overdue_category = '90+ days'
        
        bills_with_overdue.append(bill)
    
    # Order by days overdue (most overdue first)
    bills_with_overdue.sort(key=lambda x: x.days_overdue, reverse=True)
    
    # Statistics
    stats = {
        'total': len(bills_with_overdue),
        'total_balance': sum(bill.balance for bill in bills_with_overdue),
        'avg_days_overdue': sum(bill.days_overdue for bill in bills_with_overdue) / len(bills_with_overdue) if bills_with_overdue else 0,
        'overdue_1_30': sum(1 for bill in bills_with_overdue if bill.days_overdue <= 30),
        'overdue_31_60': sum(1 for bill in bills_with_overdue if 31 <= bill.days_overdue <= 60),
        'overdue_61_90': sum(1 for bill in bills_with_overdue if 61 <= bill.days_overdue <= 90),
        'overdue_90_plus': sum(1 for bill in bills_with_overdue if bill.days_overdue > 90),
    }
    
    # Get filter options
    revenue_streams = RevenueStream.objects.filter(is_active=True).order_by('name')
    sub_counties = SubCounty.objects.filter(is_active=True).order_by('name')
    
    context = {
        'bills': bills_with_overdue,
        'stats': stats,
        'search_query': search_query,
        'current_revenue_stream': revenue_stream_filter,
        'current_sub_county': sub_county_filter,
        'revenue_streams': revenue_streams,
        'sub_counties': sub_counties,
    }
    
    return render(request, 'billing/overdue_bills.html', context)


@login_required
def bill_reports(request):
    """
    Display billing reports and analytics
    """
    # Date range filter
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if not date_from:
        date_from = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = timezone.now().strftime('%Y-%m-%d')
    
    bills = Bill.objects.filter(
        bill_date__gte=date_from,
        bill_date__lte=date_to
    ).select_related('revenue_stream', 'sub_county')
    
    # Overall statistics
    overall_stats = {
        'total_bills': bills.count(),
        'total_billed': bills.aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        'total_collected': bills.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
        'total_outstanding': bills.aggregate(Sum('balance'))['balance__sum'] or 0,
        'collection_rate': 0,
    }
    
    if overall_stats['total_billed'] > 0:
        overall_stats['collection_rate'] = (
            overall_stats['total_collected'] / overall_stats['total_billed'] * 100
        )
    
    # Status breakdown
    status_breakdown = []
    for status_code, status_name in Bill.BILL_STATUS_CHOICES:
        count = bills.filter(status=status_code).count()
        amount = bills.filter(status=status_code).aggregate(
            Sum('total_amount')
        )['total_amount__sum'] or 0
        
        status_breakdown.append({
            'status': status_name,
            'count': count,
            'amount': amount,
        })
    
    # Revenue stream breakdown
    stream_breakdown = bills.values(
        'revenue_stream__code',
        'revenue_stream__name'
    ).annotate(
        bill_count=Count('id'),
        total_billed=Sum('total_amount'),
        total_collected=Sum('amount_paid'),
        outstanding=Sum('balance')
    ).order_by('-total_billed')[:10]
    
    # Sub-county breakdown
    subcounty_breakdown = bills.values(
        'sub_county__name'
    ).annotate(
        bill_count=Count('id'),
        total_billed=Sum('total_amount'),
        total_collected=Sum('amount_paid'),
        outstanding=Sum('balance')
    ).order_by('-total_billed')[:10]
    
    # Monthly trend (last 6 months)
    monthly_trend = []
    for i in range(5, -1, -1):
        date = timezone.now() - timedelta(days=30*i)
        month_start = date.replace(day=1)
        if i == 0:
            month_end = timezone.now()
        else:
            next_month = month_start + timedelta(days=32)
            month_end = next_month.replace(day=1) - timedelta(days=1)
        
        month_bills = Bill.objects.filter(
            bill_date__gte=month_start,
            bill_date__lte=month_end
        )
        
        monthly_trend.append({
            'month': month_start.strftime('%b %Y'),
            'bills': month_bills.count(),
            'billed': month_bills.aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
            'collected': month_bills.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
        })
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'overall_stats': overall_stats,
        'status_breakdown': status_breakdown,
        'stream_breakdown': stream_breakdown,
        'subcounty_breakdown': subcounty_breakdown,
        'monthly_trend': monthly_trend,
    }
    
    return render(request, 'billing/bill_reports.html', context)


@login_required
def export_bills_excel(request):
    """
    Export bills to Excel with all filters applied
    """
    # Get filtered bills (same logic as bill_list)
    bills = Bill.objects.select_related(
        'citizen', 'revenue_stream', 'sub_county', 'ward', 'created_by'
    )
    
    # Apply filters
    search_query = request.GET.get('search', '')
    if search_query:
        bills = bills.filter(
            Q(bill_number__icontains=search_query) |
            Q(citizen__first_name__icontains=search_query) |
            Q(citizen__last_name__icontains=search_query) |
            Q(citizen__business_name__icontains=search_query)
        )
    
    status_filter = request.GET.get('status', '')
    if status_filter:
        bills = bills.filter(status=status_filter)
    
    revenue_stream_filter = request.GET.get('revenue_stream', '')
    if revenue_stream_filter:
        bills = bills.filter(revenue_stream_id=revenue_stream_filter)
    
    sub_county_filter = request.GET.get('sub_county', '')
    if sub_county_filter:
        bills = bills.filter(sub_county_id=sub_county_filter)
    
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        bills = bills.filter(bill_date__gte=date_from)
    if date_to:
        bills = bills.filter(bill_date__lte=date_to)
    
    bills = bills.order_by('-bill_date')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bills Report"
    
    # Styling
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Bill Number', 'Bill Date', 'Due Date', 'Citizen/Business', 
        'Revenue Stream', 'Sub County', 'Ward', 'Bill Amount', 
        'Penalty', 'Total Amount', 'Amount Paid', 'Balance', 
        'Status', 'Created By', 'Created At'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_num, bill in enumerate(bills, 2):
        citizen_name = bill.citizen.business_name if bill.citizen.entity_type == 'business' else f"{bill.citizen.first_name} {bill.citizen.last_name}"
        
        data = [
            bill.bill_number,
            bill.bill_date.strftime('%Y-%m-%d'),
            bill.due_date.strftime('%Y-%m-%d'),
            citizen_name,
            bill.revenue_stream.name,
            bill.sub_county.name if bill.sub_county else '',
            bill.ward.name if bill.ward else '',
            float(bill.bill_amount),
            float(bill.penalty_amount),
            float(bill.total_amount),
            float(bill.amount_paid),
            float(bill.balance),
            bill.get_status_display(),
            bill.created_by.get_full_name() if bill.created_by else '',
            bill.created_at.strftime('%Y-%m-%d %H:%M'),
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Format currency columns
            if col_num in [8, 9, 10, 11, 12]:
                cell.number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Add summary row
    summary_row = bills.count() + 3
    ws.cell(row=summary_row, column=1, value="TOTALS:").font = Font(bold=True)
    ws.cell(row=summary_row, column=8, value=float(sum(b.bill_amount for b in bills))).number_format = '#,##0.00'
    ws.cell(row=summary_row, column=9, value=float(sum(b.penalty_amount for b in bills))).number_format = '#,##0.00'
    ws.cell(row=summary_row, column=10, value=float(sum(b.total_amount for b in bills))).number_format = '#,##0.00'
    ws.cell(row=summary_row, column=11, value=float(sum(b.amount_paid for b in bills))).number_format = '#,##0.00'
    ws.cell(row=summary_row, column=12, value=float(sum(b.balance for b in bills))).number_format = '#,##0.00'
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Bills_Report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


@login_required
def export_overdue_bills_excel(request):
    """
    Export overdue bills to Excel
    """
    today = timezone.now().date()
    
    bills = Bill.objects.select_related(
        'citizen', 'revenue_stream', 'sub_county', 'ward'
    ).filter(
        Q(status='issued') | Q(status='partially_paid'),
        due_date__lt=today
    )
    
    # Apply filters
    search_query = request.GET.get('search', '')
    if search_query:
        bills = bills.filter(
            Q(bill_number__icontains=search_query) |
            Q(citizen__first_name__icontains=search_query) |
            Q(citizen__last_name__icontains=search_query) |
            Q(citizen__business_name__icontains=search_query)
        )
    
    revenue_stream_filter = request.GET.get('revenue_stream', '')
    if revenue_stream_filter:
        bills = bills.filter(revenue_stream_id=revenue_stream_filter)
    
    sub_county_filter = request.GET.get('sub_county', '')
    if sub_county_filter:
        bills = bills.filter(sub_county_id=sub_county_filter)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Overdue Bills"
    
    # Styling
    header_fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Bill Number', 'Citizen/Business', 'Revenue Stream', 
        'Bill Date', 'Due Date', 'Days Overdue', 'Total Amount', 
        'Amount Paid', 'Balance', 'Sub County', 'Contact Phone', 'Status'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_num, bill in enumerate(bills, 2):
        citizen_name = bill.citizen.business_name if bill.citizen.entity_type == 'business' else f"{bill.citizen.first_name} {bill.citizen.last_name}"
        days_overdue = (today - bill.due_date).days
        
        data = [
            bill.bill_number,
            citizen_name,
            bill.revenue_stream.name,
            bill.bill_date.strftime('%Y-%m-%d'),
            bill.due_date.strftime('%Y-%m-%d'),
            days_overdue,
            float(bill.total_amount),
            float(bill.amount_paid),
            float(bill.balance),
            bill.sub_county.name if bill.sub_county else '',
            bill.citizen.phone_primary,
            bill.get_status_display(),
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Format currency columns
            if col_num in [7, 8, 9]:
                cell.number_format = '#,##0.00'
            
            # Highlight severely overdue bills
            if col_num == 6 and days_overdue > 90:
                cell.fill = PatternFill(start_color="ffcccc", end_color="ffcccc", fill_type="solid")
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Overdue_Bills_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


# payments/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, Avg
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.core.paginator import Paginator
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal

from .models import (
    Payment, PaymentMethod, PaymentReversal, BankReconciliation,
    RevenueStream, SubCounty, Ward, Bill, Citizen, User
)


# ============================================================================
# PAYMENT LIST VIEW
# ============================================================================

@login_required
def payment_list(request):
    """Display list of all payments with search, filter, and export"""
    
    # Get all payments
    payments = Payment.objects.select_related(
        'citizen', 'payment_method', 'revenue_stream', 
        'sub_county', 'ward', 'collected_by', 'bill'
    ).order_by('-payment_date')
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        payments = payments.filter(
            Q(receipt_number__icontains=search_query) |
            Q(transaction_reference__icontains=search_query) |
            Q(payer_name__icontains=search_query) |
            Q(payer_phone__icontains=search_query) |
            Q(citizen__first_name__icontains=search_query) |
            Q(citizen__last_name__icontains=search_query) |
            Q(citizen__business_name__icontains=search_query)
        )
    
    # Filters
    status_filter = request.GET.get('status', '')
    if status_filter:
        payments = payments.filter(status=status_filter)
    
    method_filter = request.GET.get('payment_method', '')
    if method_filter:
        payments = payments.filter(payment_method_id=method_filter)
    
    revenue_stream_filter = request.GET.get('revenue_stream', '')
    if revenue_stream_filter:
        payments = payments.filter(revenue_stream_id=revenue_stream_filter)
    
    sub_county_filter = request.GET.get('sub_county', '')
    if sub_county_filter:
        payments = payments.filter(sub_county_id=sub_county_filter)
    
    date_from = request.GET.get('date_from', '')
    if date_from:
        payments = payments.filter(payment_date__gte=date_from)
    
    date_to = request.GET.get('date_to', '')
    if date_to:
        payments = payments.filter(payment_date__lte=date_to)
    
    amount_min = request.GET.get('amount_min', '')
    if amount_min:
        try:
            payments = payments.filter(amount__gte=Decimal(amount_min))
        except:
            pass
    
    amount_max = request.GET.get('amount_max', '')
    if amount_max:
        try:
            payments = payments.filter(amount__lte=Decimal(amount_max))
        except:
            pass
    
    # Calculate statistics
    stats = payments.aggregate(
        total_count=Count('id'),
        total_amount=Sum('amount'),
        avg_amount=Avg('amount'),
        completed_count=Count('id', filter=Q(status='completed')),
        pending_count=Count('id', filter=Q(status='pending')),
        failed_count=Count('id', filter=Q(status='failed')),
        completed_amount=Sum('amount', filter=Q(status='completed'))
    )
    
    # Status breakdown
    status_stats = {
        'completed': payments.filter(status='completed').count(),
        'pending': payments.filter(status='pending').count(),
        'processing': payments.filter(status='processing').count(),
        'failed': payments.filter(status='failed').count(),
        'reversed': payments.filter(status='reversed').count(),
        'cancelled': payments.filter(status='cancelled').count(),
    }
    
    # Export to Excel
    if request.GET.get('export') == 'excel':
        return export_payments_excel(payments, request.GET)
    
    # Pagination
    paginator = Paginator(payments, 50)
    page_number = request.GET.get('page', 1)
    payments_page = paginator.get_page(page_number)
    
    # Get filter options
    payment_methods = PaymentMethod.objects.filter(is_active=True)
    revenue_streams = RevenueStream.objects.filter(is_active=True).select_related('department')
    sub_counties = SubCounty.objects.filter(is_active=True)
    
    context = {
        'payments': payments_page,
        'stats': stats,
        'status_stats': status_stats,
        'search_query': search_query,
        'payment_methods': payment_methods,
        'revenue_streams': revenue_streams,
        'sub_counties': sub_counties,
        'current_status': status_filter,
        'current_method': method_filter,
        'current_stream': revenue_stream_filter,
        'current_sub_county': sub_county_filter,
        'date_from': date_from,
        'date_to': date_to,
        'amount_min': amount_min,
        'amount_max': amount_max,
        'total_records': paginator.count,
    }
    
    return render(request, 'payments/payment_list.html', context)


# ============================================================================
# PAYMENT DETAIL VIEW
# ============================================================================

@login_required
def payment_detail(request, payment_id):
    """Display detailed information about a payment"""
    
    payment = get_object_or_404(
        Payment.objects.select_related(
            'citizen', 'payment_method', 'revenue_stream', 
            'sub_county', 'ward', 'collected_by', 'bill'
        ),
        id=payment_id
    )
    
    # Get payment history/audit trail
    from .models import AuditLog
    audit_logs = AuditLog.objects.filter(
        model_name='Payment',
        object_id=str(payment.id)
    ).select_related('user').order_by('-timestamp')[:20]
    
    # Get reversal if exists
    reversal = PaymentReversal.objects.filter(
        payment=payment
    ).select_related('reversed_by', 'approved_by').first()
    
    # Get related payments (same citizen)
    related_payments = Payment.objects.filter(
        citizen=payment.citizen
    ).exclude(id=payment.id).order_by('-payment_date')[:10]
    
    context = {
        'payment': payment,
        'audit_logs': audit_logs,
        'reversal': reversal,
        'related_payments': related_payments,
    }
    
    return render(request, 'payments/payment_detail.html', context)


# ============================================================================
# PAYMENT UPDATE/EDIT VIEW
# ============================================================================

@login_required
def payment_update(request, payment_id):
    """Update payment information"""
    
    payment = get_object_or_404(Payment, id=payment_id)
    
    # Only allow editing of pending/processing payments
    if payment.status not in ['pending', 'processing']:
        messages.error(request, 'Cannot edit completed or cancelled payments.')
        return redirect('payment_detail', payment_id=payment.id)
    
    if request.method == 'POST':
        # Update payment details
        payment.payer_name = request.POST.get('payer_name', payment.payer_name)
        payment.payer_phone = request.POST.get('payer_phone', payment.payer_phone)
        payment.payer_reference = request.POST.get('payer_reference', payment.payer_reference)
        payment.notes = request.POST.get('notes', payment.notes)
        
        # Update status if changed
        new_status = request.POST.get('status')
        if new_status and new_status != payment.status:
            old_status = payment.status
            payment.status = new_status
            
            # Create audit log
            from .models import AuditLog
            AuditLog.objects.create(
                user=request.user,
                action='update',
                model_name='Payment',
                object_id=str(payment.id),
                object_repr=payment.receipt_number,
                changes={
                    'status': {'old': old_status, 'new': new_status}
                },
                ip_address=request.META.get('REMOTE_ADDR')
            )
        
        payment.save()
        
        messages.success(request, 'Payment updated successfully.')
        return redirect('payment_detail', payment_id=payment.id)
    
    payment_methods = PaymentMethod.objects.filter(is_active=True)
    
    context = {
        'payment': payment,
        'payment_methods': payment_methods,
    }
    
    return render(request, 'payments/payment_update.html', context)


# ============================================================================
# PAYMENT DELETE/CANCEL VIEW
# ============================================================================

@login_required
def payment_delete(request, payment_id):
    """Cancel a payment (soft delete)"""
    
    payment = get_object_or_404(Payment, id=payment_id)
    
    # Only allow cancellation of pending payments
    if payment.status != 'pending':
        messages.error(request, 'Only pending payments can be cancelled.')
        return redirect('payment_detail', payment_id=payment.id)
    
    if request.method == 'POST':
        reason = request.POST.get('cancellation_reason', '')
        
        # Update payment status
        payment.status = 'cancelled'
        payment.notes = f"Cancelled: {reason}\n{payment.notes}"
        payment.save()
        
        # Create audit log
        from .models import AuditLog
        AuditLog.objects.create(
            user=request.user,
            action='delete',
            model_name='Payment',
            object_id=str(payment.id),
            object_repr=payment.receipt_number,
            changes={'reason': reason},
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        messages.success(request, f'Payment {payment.receipt_number} has been cancelled.')
        return redirect('payment_list')
    
    context = {
        'payment': payment,
    }
    
    return render(request, 'payments/payment_delete.html', context)


# ============================================================================
# PAYMENT METHODS MANAGEMENT
# ============================================================================

@login_required
def payment_method_list(request):
    """List all payment methods"""
    
    payment_methods = PaymentMethod.objects.all().order_by('-is_active', 'name')
    
    # Search
    search_query = request.GET.get('search', '').strip()
    if search_query:
        payment_methods = payment_methods.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(provider__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        is_active = status_filter == 'active'
        payment_methods = payment_methods.filter(is_active=is_active)
    
    # Get statistics for each method
    for method in payment_methods:
        method.total_transactions = Payment.objects.filter(
            payment_method=method,
            status='completed'
        ).count()
        
        method.total_amount = Payment.objects.filter(
            payment_method=method,
            status='completed'
        ).aggregate(total=Sum('amount'))['total'] or 0
    
    context = {
        'payment_methods': payment_methods,
        'search_query': search_query,
        'current_status': status_filter,
    }
    
    return render(request, 'payments/payment_method_list.html', context)


# ============================================================================
# PAYMENT RECONCILIATION
# ============================================================================

@login_required
def reconciliation_list(request):
    """List bank reconciliation records"""
    
    reconciliations = BankReconciliation.objects.select_related(
        'reconciled_by'
    ).order_by('-reconciliation_date')
    
    # Filter by date range
    date_from = request.GET.get('date_from', '')
    if date_from:
        reconciliations = reconciliations.filter(reconciliation_date__gte=date_from)
    
    date_to = request.GET.get('date_to', '')
    if date_to:
        reconciliations = reconciliations.filter(reconciliation_date__lte=date_to)
    
    # Filter by bank account
    bank_account = request.GET.get('bank_account', '')
    if bank_account:
        reconciliations = reconciliations.filter(bank_account=bank_account)
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        is_reconciled = status_filter == 'reconciled'
        reconciliations = reconciliations.filter(is_reconciled=is_reconciled)
    
    # Statistics
    stats = reconciliations.aggregate(
        total_count=Count('id'),
        reconciled_count=Count('id', filter=Q(is_reconciled=True)),
        pending_count=Count('id', filter=Q(is_reconciled=False)),
        total_variance=Sum('variance')
    )
    
    # Get unique bank accounts
    bank_accounts = BankReconciliation.objects.values_list(
        'bank_account', flat=True
    ).distinct()
    
    # Pagination
    paginator = Paginator(reconciliations, 20)
    page_number = request.GET.get('page', 1)
    reconciliations_page = paginator.get_page(page_number)
    
    context = {
        'reconciliations': reconciliations_page,
        'stats': stats,
        'bank_accounts': bank_accounts,
        'date_from': date_from,
        'date_to': date_to,
        'current_bank': bank_account,
        'current_status': status_filter,
    }
    
    return render(request, 'payments/reconciliation_list.html', context)


# ============================================================================
# PAYMENT REVERSALS
# ============================================================================

@login_required
def reversal_list(request):
    """List all payment reversals"""
    
    reversals = PaymentReversal.objects.select_related(
        'payment', 'payment__citizen', 'reversed_by', 'approved_by'
    ).order_by('-created_at')
    
    # Search
    search_query = request.GET.get('search', '').strip()
    if search_query:
        reversals = reversals.filter(
            Q(payment__receipt_number__icontains=search_query) |
            Q(reversal_reason__icontains=search_query)
        )
    
    # Filter by date range
    date_from = request.GET.get('date_from', '')
    if date_from:
        reversals = reversals.filter(created_at__gte=date_from)
    
    date_to = request.GET.get('date_to', '')
    if date_to:
        reversals = reversals.filter(created_at__lte=date_to)
    
    # Statistics
    stats = reversals.aggregate(
        total_count=Count('id'),
        total_amount=Sum('payment__amount')
    )
    
    # Pagination
    paginator = Paginator(reversals, 30)
    page_number = request.GET.get('page', 1)
    reversals_page = paginator.get_page(page_number)
    
    context = {
        'reversals': reversals_page,
        'stats': stats,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'payments/reversal_list.html', context)


@login_required
def payment_reverse(request, payment_id):
    """Reverse a payment"""
    
    payment = get_object_or_404(Payment, id=payment_id)
    
    # Check if payment can be reversed
    if payment.status != 'completed':
        messages.error(request, 'Only completed payments can be reversed.')
        return redirect('payment_detail', payment_id=payment.id)
    
    # Check if already reversed
    if PaymentReversal.objects.filter(payment=payment).exists():
        messages.error(request, 'This payment has already been reversed.')
        return redirect('payment_detail', payment_id=payment.id)
    
    if request.method == 'POST':
        reason = request.POST.get('reversal_reason', '')
        
        if not reason:
            messages.error(request, 'Please provide a reason for reversal.')
            return render(request, 'payments/payment_reverse.html', {'payment': payment})
        
        # Create reversal record
        reversal = PaymentReversal.objects.create(
            payment=payment,
            reversal_reason=reason,
            reversed_by=request.user,
            reversal_amount=payment.amount
        )
        
        # Update payment status
        payment.status = 'reversed'
        payment.save()
        
        # If payment was linked to a bill, update bill
        if payment.bill:
            bill = payment.bill
            bill.amount_paid -= payment.amount
            bill.balance += payment.amount
            
            if bill.balance > 0:
                if bill.status == 'paid':
                    bill.status = 'partially_paid'
            bill.save()
        
        # Create audit log
        from .models import AuditLog
        AuditLog.objects.create(
            user=request.user,
            action='reverse',
            model_name='Payment',
            object_id=str(payment.id),
            object_repr=payment.receipt_number,
            changes={'reason': reason, 'amount': str(payment.amount)},
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        messages.success(request, f'Payment {payment.receipt_number} has been reversed successfully.')
        return redirect('payment_detail', payment_id=payment.id)
    
    context = {
        'payment': payment,
    }
    
    return render(request, 'payments/payment_reverse.html', context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.http import require_http_methods
from .models import PaymentMethod

@login_required
#@permission_required('your_app.add_paymentmethod', raise_exception=True)
def payment_method_create(request):
    """
    View for creating new payment methods
    """
    if request.method == 'POST':
        try:
            # Extract form data
            name = request.POST.get('name')
            code = request.POST.get('code')
            provider = request.POST.get('provider')
            is_online = request.POST.get('is_online') == 'on'
            is_active = request.POST.get('is_active') == 'on'
            api_endpoint = request.POST.get('api_endpoint', '')
            
            # Basic validation
            if not name or not code or not provider:
                messages.error(request, "Name, Code, and Provider are required fields.")
                return render(request, 'payments/payment_method_create.html', {
                    'form_data': request.POST
                })
            
            # Check for duplicate code
            if PaymentMethod.objects.filter(code=code).exists():
                messages.error(request, f"A payment method with code '{code}' already exists.")
                return render(request, 'payments/payment_method_create.html', {
                    'form_data': request.POST
                })
            
            # Check for duplicate name
            if PaymentMethod.objects.filter(name=name).exists():
                messages.error(request, f"A payment method with name '{name}' already exists.")
                return render(request, 'payments/payment_method_create.html', {
                    'form_data': request.POST
                })
            
            # Create payment method
            payment_method = PaymentMethod.objects.create(
                name=name,
                code=code,
                provider=provider,
                is_online=is_online,
                is_active=is_active,
                api_endpoint=api_endpoint,
                configuration={}  # Empty configuration by default
            )
            
            messages.success(request, f"Payment method '{payment_method.name}' created successfully!")
            return redirect('payment_method_list')  # Redirect to payment method list view
            
        except Exception as e:
            messages.error(request, f"Error creating payment method: {str(e)}")
            return render(request, 'payments/payment_method_create.html', {
                'form_data': request.POST
            })
    
    # GET request - show empty form
    return render(request, 'payments/payment_method_create.html')


@login_required
@permission_required('your_app.view_paymentmethod', raise_exception=True)
def payment_method_detail(request, pk):
    """
    View for displaying payment method details
    """
    try:
        payment_method = get_object_or_404(PaymentMethod, pk=pk)
        
        # Get payment statistics for this method
        payment_stats = Payment.objects.filter(payment_method=payment_method).aggregate(
            total_payments=models.Count('id'),
            total_amount=models.Sum('amount'),
            successful_payments=models.Count('id', filter=models.Q(status='completed')),
            failed_payments=models.Count('id', filter=models.Q(status='failed'))
        )
        
        # Recent payments using this method
        recent_payments = Payment.objects.filter(
            payment_method=payment_method
        ).select_related('citizen', 'bill').order_by('-payment_date')[:10]
        
        context = {
            'method': payment_method,
            'payment_stats': payment_stats,
            'recent_payments': recent_payments,
        }
        
        return render(request, 'payments/payment_method_detail.html', context)
        
    except Exception as e:
        messages.error(request, f"Error loading payment method details: {str(e)}")
        return redirect('payment_method_list')

@login_required
@permission_required('your_app.change_paymentmethod', raise_exception=True)
def payment_method_update(request, pk):
    """
    View for updating payment methods
    """
    payment_method = get_object_or_404(PaymentMethod, pk=pk)
    
    if request.method == 'POST':
        try:
            # Extract form data
            name = request.POST.get('name')
            code = request.POST.get('code')
            provider = request.POST.get('provider')
            is_online = request.POST.get('is_online') == 'on'
            is_active = request.POST.get('is_active') == 'on'
            api_endpoint = request.POST.get('api_endpoint', '')
            
            # Basic validation
            if not name or not code or not provider:
                messages.error(request, "Name, Code, and Provider are required fields.")
                return render(request, 'payments/payment_method_update.html', {
                    'method': payment_method
                })
            
            # Check for duplicate code (excluding current instance)
            if PaymentMethod.objects.filter(code=code).exclude(pk=pk).exists():
                messages.error(request, f"A payment method with code '{code}' already exists.")
                return render(request, 'payments/payment_method_update.html', {
                    'method': payment_method
                })
            
            # Check for duplicate name (excluding current instance)
            if PaymentMethod.objects.filter(name=name).exclude(pk=pk).exists():
                messages.error(request, f"A payment method with name '{name}' already exists.")
                return render(request, 'payments/payment_method_update.html', {
                    'method': payment_method
                })
            
            # Update payment method
            payment_method.name = name
            payment_method.code = code
            payment_method.provider = provider
            payment_method.is_online = is_online
            payment_method.is_active = is_active
            payment_method.api_endpoint = api_endpoint
            payment_method.save()
            
            messages.success(request, f"Payment method '{payment_method.name}' updated successfully!")
            return redirect('payment_method_detail', pk=payment_method.pk)
            
        except Exception as e:
            messages.error(request, f"Error updating payment method: {str(e)}")
            return render(request, 'payments/payment_method_update.html', {
                'method': payment_method
            })
    
    # GET request - show form with current data
    return render(request, 'payments/payment_method_update.html', {
        'method': payment_method
    })

@login_required
@permission_required('your_app.delete_paymentmethod', raise_exception=True)
@require_http_methods(["POST"])
def payment_method_delete(request, pk):
    """
    View for deleting payment methods
    """
    payment_method = get_object_or_404(PaymentMethod, pk=pk)
    
    try:
        # Check if payment method is being used
        payment_count = Payment.objects.filter(payment_method=payment_method).count()
        
        if payment_count > 0:
            messages.error(
                request, 
                f"Cannot delete '{payment_method.name}'. It is being used by {payment_count} payment(s). "
                "Consider deactivating it instead."
            )
            return redirect('payment_method_detail', pk=pk)
        
        method_name = payment_method.name
        payment_method.delete()
        
        messages.success(request, f"Payment method '{method_name}' deleted successfully!")
        return redirect('payment_method_list')
        
    except Exception as e:
        messages.error(request, f"Error deleting payment method: {str(e)}")
        return redirect('payment_method_detail', pk=pk)

@login_required
@permission_required('your_app.change_paymentmethod', raise_exception=True)
@require_http_methods(["POST"])
def payment_method_toggle_active(request, pk):
    """
    View for toggling payment method active status
    """
    payment_method = get_object_or_404(PaymentMethod, pk=pk)
    
    try:
        payment_method.is_active = not payment_method.is_active
        payment_method.save()
        
        status = "activated" if payment_method.is_active else "deactivated"
        messages.success(request, f"Payment method '{payment_method.name}' {status} successfully!")
        
    except Exception as e:
        messages.error(request, f"Error updating payment method status: {str(e)}")
    
    return redirect('payment_method_detail', pk=pk)

# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================

def export_payments_excel(payments, filters):
    """Export payments to Excel"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Report header
    ws.merge_cells('A1:N1')
    title_cell = ws['A1']
    title_cell.value = "WAJIR COUNTY GOVERNMENT - PAYMENTS REPORT"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Report info
    ws.merge_cells('A2:N2')
    info_cell = ws['A2']
    info_cell.value = f"Generated on: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    info_cell.alignment = Alignment(horizontal="center")
    
    # Filter info
    filter_info = []
    if filters.get('date_from'):
        filter_info.append(f"From: {filters['date_from']}")
    if filters.get('date_to'):
        filter_info.append(f"To: {filters['date_to']}")
    if filters.get('status'):
        filter_info.append(f"Status: {filters['status']}")
    
    if filter_info:
        ws.merge_cells('A3:N3')
        filter_cell = ws['A3']
        filter_cell.value = f"Filters: {' | '.join(filter_info)}"
        filter_cell.alignment = Alignment(horizontal="center")
        header_row = 5
    else:
        header_row = 4
    
    # Column headers
    headers = [
        'Receipt No',
        'Transaction Ref',
        'Payment Date',
        'Payer Name',
        'Phone',
        'Payment Method',
        'Revenue Stream',
        'Amount (KES)',
        'Status',
        'Sub County',
        'Ward',
        'Collected By',
        'Bill Number',
        'Notes'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    row_num = header_row + 1
    total_amount = 0
    
    for payment in payments:
        ws.cell(row=row_num, column=1, value=payment.receipt_number).border = border
        ws.cell(row=row_num, column=2, value=payment.transaction_reference).border = border
        ws.cell(row=row_num, column=3, value=payment.payment_date.strftime('%d/%m/%Y %H:%M')).border = border
        ws.cell(row=row_num, column=4, value=payment.payer_name).border = border
        ws.cell(row=row_num, column=5, value=payment.payer_phone).border = border
        ws.cell(row=row_num, column=6, value=payment.payment_method.name).border = border
        ws.cell(row=row_num, column=7, value=payment.revenue_stream.name).border = border
        
        amount_cell = ws.cell(row=row_num, column=8, value=float(payment.amount))
        amount_cell.number_format = '#,##0.00'
        amount_cell.border = border
        
        ws.cell(row=row_num, column=9, value=payment.get_status_display()).border = border
        ws.cell(row=row_num, column=10, value=payment.sub_county.name if payment.sub_county else '').border = border
        ws.cell(row=row_num, column=11, value=payment.ward.name if payment.ward else '').border = border
        ws.cell(row=row_num, column=12, value=payment.collected_by.get_full_name() if payment.collected_by else '').border = border
        ws.cell(row=row_num, column=13, value=payment.bill.bill_number if payment.bill else '').border = border
        ws.cell(row=row_num, column=14, value=payment.notes).border = border
        
        total_amount += payment.amount
        row_num += 1
    
    # Total row
    ws.cell(row=row_num, column=7, value="TOTAL:").font = Font(bold=True)
    total_cell = ws.cell(row=row_num, column=8, value=float(total_amount))
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00'
    
    # Adjust column widths
    column_widths = [15, 20, 18, 25, 15, 18, 30, 15, 12, 15, 15, 20, 15, 30]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=payments_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.http import require_http_methods
from django.db import models
from django.utils import timezone
from datetime import timedelta
from .models import Payment, PaymentMethod, RevenueStream, Reconciliation

@login_required
#@permission_required('your_app.add_reconciliation', raise_exception=True)
def reconciliation_create(request):
    """
    View for creating new payment reconciliations
    """
    # Calculate default dates
    today = timezone.now().date()
    default_end_date = today
    default_start_date = today - timedelta(days=7)
    default_reconciliation_date = today
    
    # Get available payment methods and revenue streams
    payment_methods = PaymentMethod.objects.filter(is_active=True)
    revenue_streams = RevenueStream.objects.filter(is_active=True)
    
    if request.method == 'POST':
        try:
            # Extract form data
            reconciliation_date = request.POST.get('reconciliation_date')
            payment_method_id = request.POST.get('payment_method')
            revenue_stream_id = request.POST.get('revenue_stream')
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            notes = request.POST.get('notes', '')
            
            # Basic validation
            if not reconciliation_date or not payment_method_id or not start_date or not end_date:
                messages.error(request, "Please fill in all required fields.")
                return render(request, 'payments/reconciliation_create.html', {
                    'form_data': request.POST,
                    'payment_methods': payment_methods,
                    'revenue_streams': revenue_streams,
                    'default_start_date': default_start_date,
                    'default_end_date': default_end_date,
                    'default_reconciliation_date': default_reconciliation_date,
                })
            
            # Convert dates
            reconciliation_date = timezone.datetime.strptime(reconciliation_date, '%Y-%m-%d').date()
            start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
            
            if start_date > end_date:
                messages.error(request, "Start date cannot be after end date.")
                return render(request, 'payments/reconciliation_create.html', {
                    'form_data': request.POST,
                    'payment_methods': payment_methods,
                    'revenue_streams': revenue_streams,
                    'default_start_date': default_start_date,
                    'default_end_date': default_end_date,
                    'default_reconciliation_date': default_reconciliation_date,
                })
            
            if reconciliation_date < start_date or reconciliation_date > end_date:
                messages.error(request, "Reconciliation date must be within the reconciliation period.")
                return render(request, 'payments/reconciliation_create.html', {
                    'form_data': request.POST,
                    'payment_methods': payment_methods,
                    'revenue_streams': revenue_streams,
                    'default_start_date': default_start_date,
                    'default_end_date': default_end_date,
                    'default_reconciliation_date': default_reconciliation_date,
                })
            
            # Get payment method and revenue stream
            payment_method = get_object_or_404(PaymentMethod, id=payment_method_id, is_active=True)
            revenue_stream = None
            if revenue_stream_id:
                revenue_stream = get_object_or_404(RevenueStream, id=revenue_stream_id, is_active=True)
            
            # Get payments for reconciliation period
            payments = Payment.objects.filter(
                payment_method=payment_method,
                payment_date__date__range=[start_date, end_date],
                status='completed'
            )
            
            if revenue_stream:
                payments = payments.filter(revenue_stream=revenue_stream)
            
            # Calculate totals
            total_payments = payments.count()
            total_amount = payments.aggregate(total=models.Sum('amount'))['total'] or 0
            
            if total_payments == 0:
                messages.warning(request, f"No completed payments found for the selected period and criteria. Do you want to continue?")
                # You might want to add a confirmation step here
            
            # Generate reconciliation number
            last_reconciliation = Reconciliation.objects.order_by('-id').first()
            next_id = last_reconciliation.id + 1 if last_reconciliation else 1
            reconciliation_number = f"REC-{timezone.now().strftime('%Y%m%d')}-{next_id:04d}"
            
            # Create reconciliation
            reconciliation = Reconciliation.objects.create(
                reconciliation_number=reconciliation_number,
                reconciliation_date=reconciliation_date,
                payment_method=payment_method,
                revenue_stream=revenue_stream,
                start_date=start_date,
                end_date=end_date,
                total_payments=total_payments,
                total_amount=total_amount,
                reconciled_amount=0,  # Will be updated during reconciliation process
                discrepancy_amount=0,
                status='draft',
                notes=notes,
                created_by=request.user
            )
            
            # Associate payments with this reconciliation
            payments.update(reconciliation=reconciliation)
            
            messages.success(request, f"Reconciliation {reconciliation_number} created successfully!")
            return redirect('reconciliation_detail', pk=reconciliation.id)
            
        except Exception as e:
            messages.error(request, f"Error creating reconciliation: {str(e)}")
            return render(request, 'payments/reconciliation_create.html', {
                'form_data': request.POST,
                'payment_methods': payment_methods,
                'revenue_streams': revenue_streams,
                'default_start_date': default_start_date,
                'default_end_date': default_end_date,
                'default_reconciliation_date': default_reconciliation_date,
            })
    
    # GET request - show empty form
    context = {
        'payment_methods': payment_methods,
        'revenue_streams': revenue_streams,
        'default_start_date': default_start_date,
        'default_end_date': default_end_date,
        'default_reconciliation_date': default_reconciliation_date,
    }
    
    return render(request, 'payments/reconciliation_create.html', context)


"""
Fleet Management Views - Wajir County ERP
Handles vehicles, fuel, maintenance, and trips management
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, Avg
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import (
    FleetVehicle, FuelStation, FuelCard, FuelTransaction,
    VehicleMaintenance, VehicleTrip, Department, SubCounty, User
)


# ============================================================================
# VEHICLES MANAGEMENT
# ============================================================================

@login_required
def vehicle_list(request):
    """List all fleet vehicles with filtering and search"""
    vehicles = FleetVehicle.objects.select_related(
        'department', 'current_driver'
    ).all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        vehicles = vehicles.filter(
            Q(fleet_number__icontains=search_query) |
            Q(registration_number__icontains=search_query) |
            Q(make__icontains=search_query) |
            Q(model__icontains=search_query)
        )
    
    # Filters
    vehicle_type = request.GET.get('vehicle_type', '')
    department_id = request.GET.get('department', '')
    status = request.GET.get('status', '')
    fuel_type = request.GET.get('fuel_type', '')
    
    if vehicle_type:
        vehicles = vehicles.filter(vehicle_type=vehicle_type)
    if department_id:
        vehicles = vehicles.filter(department_id=department_id)
    if status:
        vehicles = vehicles.filter(status=status)
    if fuel_type:
        vehicles = vehicles.filter(fuel_type=fuel_type)
    
    # Statistics
    stats = {
        'total': FleetVehicle.objects.count(),
        'active': FleetVehicle.objects.filter(status='active').count(),
        'maintenance': FleetVehicle.objects.filter(status='maintenance').count(),
        'inactive': FleetVehicle.objects.filter(status='inactive').count(),
        'cars': FleetVehicle.objects.filter(vehicle_type='car').count(),
        'trucks': FleetVehicle.objects.filter(vehicle_type='truck').count(),
    }
    
    # Pagination
    paginator = Paginator(vehicles, 25)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    # Get departments for filter
    departments = Department.objects.filter(is_active=True)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'departments': departments,
        'current_vehicle_type': vehicle_type,
        'current_department': department_id,
        'current_status': status,
        'current_fuel_type': fuel_type,
    }
    
    return render(request, 'fleet/vehicle_list.html', context)


@login_required
def vehicle_detail(request, fleet_number):
    """View vehicle details"""
    vehicle = get_object_or_404(FleetVehicle, fleet_number=fleet_number)
    
    # Get related data
    fuel_transactions = FuelTransaction.objects.filter(
        vehicle=vehicle
    ).select_related('fuel_station', 'driver').order_by('-transaction_date')[:10]
    
    maintenance_records = VehicleMaintenance.objects.filter(
        vehicle=vehicle
    ).select_related('requested_by').order_by('-scheduled_date')[:10]
    
    trips = VehicleTrip.objects.filter(
        vehicle=vehicle
    ).select_related('driver', 'approved_by').order_by('-scheduled_departure')[:10]
    
    # Calculate statistics
    thirty_days_ago = timezone.now() - timedelta(days=30)
    
    fuel_stats = FuelTransaction.objects.filter(
        vehicle=vehicle,
        transaction_date__gte=thirty_days_ago
    ).aggregate(
        total_fuel=Sum('quantity_liters'),
        total_cost=Sum('total_amount'),
        avg_price=Avg('unit_price')
    )
    
    maintenance_stats = VehicleMaintenance.objects.filter(
        vehicle=vehicle,
        completed_date__gte=thirty_days_ago
    ).aggregate(
        total_cost=Sum('cost'),
        count=Count('id')
    )
    
    trip_stats = VehicleTrip.objects.filter(
        vehicle=vehicle,
        actual_departure__gte=thirty_days_ago
    ).count()
    
    context = {
        'vehicle': vehicle,
        'fuel_transactions': fuel_transactions,
        'maintenance_records': maintenance_records,
        'trips': trips,
        'fuel_stats': fuel_stats,
        'maintenance_stats': maintenance_stats,
        'trip_stats': trip_stats,
    }
    
    return render(request, 'fleet/vehicle_detail.html', context)


@login_required
def vehicle_create(request):
    """Create new vehicle"""
    if request.method == 'POST':
        try:
            vehicle = FleetVehicle.objects.create(
                fleet_number=request.POST.get('fleet_number'),
                registration_number=request.POST.get('registration_number'),
                vehicle_type=request.POST.get('vehicle_type'),
                make=request.POST.get('make'),
                model=request.POST.get('model'),
                year=request.POST.get('year'),
                department_id=request.POST.get('department'),
                fuel_type=request.POST.get('fuel_type'),
                engine_capacity=request.POST.get('engine_capacity', ''),
                purchase_date=request.POST.get('purchase_date'),
                purchase_cost=request.POST.get('purchase_cost'),
                insurance_expiry=request.POST.get('insurance_expiry'),
                inspection_due=request.POST.get('inspection_due'),
                current_mileage=request.POST.get('current_mileage', 0),
                has_gps=request.POST.get('has_gps') == 'on',
                gps_device_id=request.POST.get('gps_device_id', ''),
            )
            
            # Assign driver if provided
            driver_id = request.POST.get('current_driver')
            if driver_id:
                vehicle.current_driver_id = driver_id
                vehicle.save()
            
            messages.success(request, f'Vehicle {vehicle.fleet_number} registered successfully!')
            return redirect('vehicle_detail', fleet_number=vehicle.fleet_number)
            
        except Exception as e:
            messages.error(request, f'Error creating vehicle: {str(e)}')
    
    departments = Department.objects.filter(is_active=True)
    drivers = User.objects.filter(is_active=True, is_active_staff=True)
    
    context = {
        'departments': departments,
        'drivers': drivers,
    }
    
    return render(request, 'fleet/vehicle_form.html', context)


@login_required
def vehicle_update(request, fleet_number):
    """Update vehicle details"""
    vehicle = get_object_or_404(FleetVehicle, fleet_number=fleet_number)
    
    if request.method == 'POST':
        try:
            vehicle.registration_number = request.POST.get('registration_number')
            vehicle.vehicle_type = request.POST.get('vehicle_type')
            vehicle.make = request.POST.get('make')
            vehicle.model = request.POST.get('model')
            vehicle.year = request.POST.get('year')
            vehicle.department_id = request.POST.get('department')
            vehicle.fuel_type = request.POST.get('fuel_type')
            vehicle.engine_capacity = request.POST.get('engine_capacity', '')
            vehicle.purchase_date = request.POST.get('purchase_date')
            vehicle.purchase_cost = request.POST.get('purchase_cost')
            vehicle.insurance_expiry = request.POST.get('insurance_expiry')
            vehicle.inspection_due = request.POST.get('inspection_due')
            vehicle.current_mileage = request.POST.get('current_mileage')
            vehicle.status = request.POST.get('status')
            vehicle.has_gps = request.POST.get('has_gps') == 'on'
            vehicle.gps_device_id = request.POST.get('gps_device_id', '')
            
            driver_id = request.POST.get('current_driver')
            vehicle.current_driver_id = driver_id if driver_id else None
            
            vehicle.save()
            
            messages.success(request, 'Vehicle updated successfully!')
            return redirect('vehicle_detail', fleet_number=vehicle.fleet_number)
            
        except Exception as e:
            messages.error(request, f'Error updating vehicle: {str(e)}')
    
    departments = Department.objects.filter(is_active=True)
    drivers = User.objects.filter(is_active=True, is_active_staff=True)
    
    context = {
        'vehicle': vehicle,
        'departments': departments,
        'drivers': drivers,
    }
    
    return render(request, 'fleet/vehicle_form.html', context)


@login_required
def vehicle_delete(request, fleet_number):
    """Delete vehicle"""
    vehicle = get_object_or_404(FleetVehicle, fleet_number=fleet_number)
    
    if request.method == 'POST':
        try:
            vehicle.delete()
            messages.success(request, f'Vehicle {fleet_number} deleted successfully!')
            return redirect('vehicle_list')
        except Exception as e:
            messages.error(request, f'Error deleting vehicle: {str(e)}')
            return redirect('vehicle_detail', fleet_number=fleet_number)
    
    return redirect('vehicle_detail', fleet_number=fleet_number)


@login_required
def vehicle_export_excel(request):
    """Export vehicles to Excel"""
    # Get filtered vehicles
    vehicles = FleetVehicle.objects.select_related('department', 'current_driver').all()
    
    # Apply filters
    vehicle_type = request.GET.get('vehicle_type', '')
    department_id = request.GET.get('department', '')
    status = request.GET.get('status', '')
    
    if vehicle_type:
        vehicles = vehicles.filter(vehicle_type=vehicle_type)
    if department_id:
        vehicles = vehicles.filter(department_id=department_id)
    if status:
        vehicles = vehicles.filter(status=status)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fleet Vehicles'
    
    # Styles
    header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Fleet Number', 'Registration Number', 'Vehicle Type', 'Make', 'Model',
        'Year', 'Department', 'Current Driver', 'Fuel Type', 'Status',
        'Current Mileage', 'Insurance Expiry', 'Inspection Due', 'Purchase Cost'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data
    for row, vehicle in enumerate(vehicles, 2):
        data = [
            vehicle.fleet_number,
            vehicle.registration_number,
            vehicle.get_vehicle_type_display(),
            vehicle.make,
            vehicle.model,
            vehicle.year,
            vehicle.department.name if vehicle.department else '',
            vehicle.current_driver.get_full_name() if vehicle.current_driver else '',
            vehicle.fuel_type,
            vehicle.get_status_display(),
            vehicle.current_mileage,
            vehicle.insurance_expiry.strftime('%Y-%m-%d') if vehicle.insurance_expiry else '',
            vehicle.inspection_due.strftime('%Y-%m-%d') if vehicle.inspection_due else '',
            float(vehicle.purchase_cost),
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 14:  # Purchase cost
                cell.number_format = '#,##0.00'
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=fleet_vehicles_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response


# ============================================================================
# FUEL MANAGEMENT
# ============================================================================

@login_required
def fuel_transaction_list(request):
    """List fuel transactions"""
    transactions = FuelTransaction.objects.select_related(
        'vehicle', 'fuel_station', 'driver', 'fuel_card'
    ).all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        transactions = transactions.filter(
            Q(transaction_number__icontains=search_query) |
            Q(vehicle__registration_number__icontains=search_query) |
            Q(vehicle__fleet_number__icontains=search_query)
        )
    
    # Filters
    vehicle_id = request.GET.get('vehicle', '')
    fuel_station_id = request.GET.get('fuel_station', '')
    transaction_type = request.GET.get('transaction_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if vehicle_id:
        transactions = transactions.filter(vehicle_id=vehicle_id)
    if fuel_station_id:
        transactions = transactions.filter(fuel_station_id=fuel_station_id)
    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)
    if date_from:
        transactions = transactions.filter(transaction_date__gte=date_from)
    if date_to:
        transactions = transactions.filter(transaction_date__lte=date_to)
    
    # Statistics
    thirty_days_ago = timezone.now() - timedelta(days=30)
    stats = FuelTransaction.objects.filter(
        transaction_date__gte=thirty_days_ago
    ).aggregate(
        total_transactions=Count('id'),
        total_liters=Sum('quantity_liters'),
        total_amount=Sum('total_amount'),
        avg_price=Avg('unit_price')
    )
    
    # Pagination
    paginator = Paginator(transactions, 25)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    vehicles = FleetVehicle.objects.filter(status='active')
    fuel_stations = FuelStation.objects.filter(is_active=True)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'vehicles': vehicles,
        'fuel_stations': fuel_stations,
        'current_vehicle': vehicle_id,
        'current_fuel_station': fuel_station_id,
        'current_transaction_type': transaction_type,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'fleet/fuel_transaction_list.html', context)


@login_required
def fuel_transaction_create(request):
    """Create fuel transaction"""
    if request.method == 'POST':
        try:
            # Generate transaction number
            last_transaction = FuelTransaction.objects.order_by('-id').first()
            if last_transaction:
                last_num = int(last_transaction.transaction_number.split('-')[-1])
                transaction_number = f'FT-{timezone.now().year}-{last_num + 1:06d}'
            else:
                transaction_number = f'FT-{timezone.now().year}-000001'
            
            quantity = Decimal(request.POST.get('quantity_liters'))
            unit_price = Decimal(request.POST.get('unit_price'))
            total_amount = quantity * unit_price
            
            transaction = FuelTransaction.objects.create(
                transaction_number=transaction_number,
                vehicle_id=request.POST.get('vehicle'),
                fuel_station_id=request.POST.get('fuel_station'),
                transaction_type=request.POST.get('transaction_type'),
                transaction_date=request.POST.get('transaction_date'),
                quantity_liters=quantity,
                unit_price=unit_price,
                total_amount=total_amount,
                mileage=request.POST.get('mileage'),
                driver_id=request.POST.get('driver'),
                receipt_number=request.POST.get('receipt_number', ''),
                notes=request.POST.get('notes', ''),
            )
            
            # Update vehicle mileage
            vehicle = transaction.vehicle
            vehicle.current_mileage = transaction.mileage
            vehicle.save()
            
            messages.success(request, f'Fuel transaction {transaction_number} recorded successfully!')
            return redirect('fuel_transaction_list')
            
        except Exception as e:
            messages.error(request, f'Error creating transaction: {str(e)}')
    
    vehicles = FleetVehicle.objects.filter(status='active')
    fuel_stations = FuelStation.objects.filter(is_active=True)
    drivers = User.objects.filter(is_active=True, is_active_staff=True)
    
    context = {
        'vehicles': vehicles,
        'fuel_stations': fuel_stations,
        'drivers': drivers,
    }
    
    return render(request, 'fleet/fuel_transaction_form.html', context)


@login_required
def fuel_export_excel(request):
    """Export fuel transactions to Excel"""
    transactions = FuelTransaction.objects.select_related(
        'vehicle', 'fuel_station', 'driver'
    ).all()
    
    # Apply filters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if date_from:
        transactions = transactions.filter(transaction_date__gte=date_from)
    if date_to:
        transactions = transactions.filter(transaction_date__lte=date_to)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fuel Transactions'
    
    # Styles
    header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Transaction Number', 'Date', 'Vehicle', 'Registration', 'Fuel Station',
        'Type', 'Quantity (L)', 'Unit Price', 'Total Amount', 'Mileage',
        'Driver', 'Receipt Number'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data
    for row, trans in enumerate(transactions, 2):
        data = [
            trans.transaction_number,
            trans.transaction_date.strftime('%Y-%m-%d %H:%M'),
            trans.vehicle.fleet_number,
            trans.vehicle.registration_number,
            trans.fuel_station.name,
            trans.get_transaction_type_display(),
            float(trans.quantity_liters),
            float(trans.unit_price),
            float(trans.total_amount),
            trans.mileage,
            trans.driver.get_full_name() if trans.driver else '',
            trans.receipt_number,
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [7, 8, 9]:  # Numeric columns
                cell.number_format = '#,##0.00'
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=fuel_transactions_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response


# ============================================================================
# MAINTENANCE MANAGEMENT
# ============================================================================

@login_required
def maintenance_list(request):
    """List maintenance records"""
    maintenance = VehicleMaintenance.objects.select_related(
        'vehicle', 'requested_by'
    ).all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        maintenance = maintenance.filter(
            Q(maintenance_number__icontains=search_query) |
            Q(vehicle__registration_number__icontains=search_query) |
            Q(vehicle__fleet_number__icontains=search_query)
        )
    
    # Filters
    vehicle_id = request.GET.get('vehicle', '')
    maintenance_type = request.GET.get('maintenance_type', '')
    status = request.GET.get('status', '')
    
    if vehicle_id:
        maintenance = maintenance.filter(vehicle_id=vehicle_id)
    if maintenance_type:
        maintenance = maintenance.filter(maintenance_type=maintenance_type)
    if status:
        maintenance = maintenance.filter(status=status)
    
    # Statistics
    stats = {
        'total': VehicleMaintenance.objects.count(),
        'scheduled': VehicleMaintenance.objects.filter(status='scheduled').count(),
        'in_progress': VehicleMaintenance.objects.filter(status='in_progress').count(),
        'completed': VehicleMaintenance.objects.filter(status='completed').count(),
        'total_cost': VehicleMaintenance.objects.filter(
            status='completed'
        ).aggregate(Sum('cost'))['cost__sum'] or 0,
    }
    
    # Pagination
    paginator = Paginator(maintenance, 25)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    vehicles = FleetVehicle.objects.all()
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'vehicles': vehicles,
        'current_vehicle': vehicle_id,
        'current_maintenance_type': maintenance_type,
        'current_status': status,
    }
    
    return render(request, 'fleet/maintenance_list.html', context)


@login_required
def maintenance_create(request):
    """Create maintenance record"""
    if request.method == 'POST':
        try:
            # Generate maintenance number
            last_maintenance = VehicleMaintenance.objects.order_by('-id').first()
            if last_maintenance:
                last_num = int(last_maintenance.maintenance_number.split('-')[-1])
                maintenance_number = f'MT-{timezone.now().year}-{last_num + 1:06d}'
            else:
                maintenance_number = f'MT-{timezone.now().year}-000001'
            
            maintenance = VehicleMaintenance.objects.create(
                maintenance_number=maintenance_number,
                vehicle_id=request.POST.get('vehicle'),
                maintenance_type=request.POST.get('maintenance_type'),
                description=request.POST.get('description'),
                scheduled_date=request.POST.get('scheduled_date'),
                service_provider=request.POST.get('service_provider'),
                cost=request.POST.get('cost'),
                mileage=request.POST.get('mileage'),
                requested_by=request.user,
            )
            
            messages.success(request, f'Maintenance {maintenance_number} scheduled successfully!')
            return redirect('maintenance_list')
            
        except Exception as e:
            messages.error(request, f'Error creating maintenance record: {str(e)}')
    
    vehicles = FleetVehicle.objects.all()
    
    context = {
        'vehicles': vehicles,
    }
    
    return render(request, 'fleet/maintenance_form.html', context)


@login_required
def maintenance_update(request, maintenance_number):
    """Update maintenance record"""
    maintenance = get_object_or_404(VehicleMaintenance, maintenance_number=maintenance_number)
    
    if request.method == 'POST':
        try:
            maintenance.vehicle_id = request.POST.get('vehicle')
            maintenance.maintenance_type = request.POST.get('maintenance_type')
            maintenance.description = request.POST.get('description')
            maintenance.scheduled_date = request.POST.get('scheduled_date')
            maintenance.service_provider = request.POST.get('service_provider')
            maintenance.cost = request.POST.get('cost')
            maintenance.mileage = request.POST.get('mileage')
            maintenance.status = request.POST.get('status')
            
            if request.POST.get('completed_date'):
                maintenance.completed_date = request.POST.get('completed_date')
            
            maintenance.save()
            
            messages.success(request, 'Maintenance record updated successfully!')
            return redirect('maintenance_list')
            
        except Exception as e:
            messages.error(request, f'Error updating maintenance record: {str(e)}')
    
    vehicles = FleetVehicle.objects.all()
    
    context = {
        'maintenance': maintenance,
        'vehicles': vehicles,
    }
    
    return render(request, 'fleet/maintenance_form.html', context)


@login_required
def maintenance_delete(request, maintenance_number):
    """Delete maintenance record"""
    maintenance = get_object_or_404(VehicleMaintenance, maintenance_number=maintenance_number)
    
    if request.method == 'POST':
        try:
            maintenance.delete()
            messages.success(request, f'Maintenance record {maintenance_number} deleted successfully!')
            return redirect('maintenance_list')
        except Exception as e:
            messages.error(request, f'Error deleting maintenance record: {str(e)}')
    
    return redirect('maintenance_list')


# ============================================================================
# TRIPS & WORK TICKETS
# ============================================================================

@login_required
def trip_list(request):
    """List vehicle trips"""
    trips = VehicleTrip.objects.select_related(
        'vehicle', 'driver', 'approved_by'
    ).all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        trips = trips.filter(
            Q(trip_number__icontains=search_query) |
            Q(vehicle__registration_number__icontains=search_query) |
            Q(destination__icontains=search_query)
        )
    
    # Filters
    vehicle_id = request.GET.get('vehicle', '')
    status = request.GET.get('status', '')
    
    if vehicle_id:
        trips = trips.filter(vehicle_id=vehicle_id)
    if status:
        trips = trips.filter(status=status)
    
    # Statistics
    stats = {
        'total': VehicleTrip.objects.count(),
        'scheduled': VehicleTrip.objects.filter(status='scheduled').count(),
        'in_progress': VehicleTrip.objects.filter(status='in_progress').count(),
        'completed': VehicleTrip.objects.filter(status='completed').count(),
    }
    
    # Pagination
    paginator = Paginator(trips, 25)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    vehicles = FleetVehicle.objects.filter(status='active')
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'vehicles': vehicles,
        'current_vehicle': vehicle_id,
        'current_status': status,
    }
    
    return render(request, 'fleet/trip_list.html', context)


@login_required
def trip_create(request):
    """Create trip/work ticket"""
    if request.method == 'POST':
        try:
            # Generate trip number
            last_trip = VehicleTrip.objects.order_by('-id').first()
            if last_trip:
                last_num = int(last_trip.trip_number.split('-')[-1])
                trip_number = f'WT-{timezone.now().year}-{last_num + 1:06d}'
            else:
                trip_number = f'WT-{timezone.now().year}-000001'
            
            trip = VehicleTrip.objects.create(
                trip_number=trip_number,
                vehicle_id=request.POST.get('vehicle'),
                driver_id=request.POST.get('driver'),
                purpose=request.POST.get('purpose'),
                destination=request.POST.get('destination'),
                scheduled_departure=request.POST.get('scheduled_departure'),
                scheduled_return=request.POST.get('scheduled_return'),
                start_mileage=request.POST.get('start_mileage'),
                notes=request.POST.get('notes', ''),
            )
            
            messages.success(request, f'Work ticket {trip_number} created successfully!')
            return redirect('trip_list')
            
        except Exception as e:
            messages.error(request, f'Error creating work ticket: {str(e)}')
    
    vehicles = FleetVehicle.objects.filter(status='active')
    drivers = User.objects.filter(is_active=True, is_active_staff=True)
    
    context = {
        'vehicles': vehicles,
        'drivers': drivers,
    }
    
    return render(request, 'fleet/trip_form.html', context)


@login_required
def trip_update(request, trip_number):
    """Update trip/work ticket"""
    trip = get_object_or_404(VehicleTrip, trip_number=trip_number)
    
    if request.method == 'POST':
        try:
            trip.vehicle_id = request.POST.get('vehicle')
            trip.driver_id = request.POST.get('driver')
            trip.purpose = request.POST.get('purpose')
            trip.destination = request.POST.get('destination')
            trip.scheduled_departure = request.POST.get('scheduled_departure')
            trip.scheduled_return = request.POST.get('scheduled_return')
            trip.start_mileage = request.POST.get('start_mileage')
            trip.status = request.POST.get('status')
            trip.notes = request.POST.get('notes', '')
            
            if request.POST.get('actual_departure'):
                trip.actual_departure = request.POST.get('actual_departure')
            if request.POST.get('actual_return'):
                trip.actual_return = request.POST.get('actual_return')
            if request.POST.get('end_mileage'):
                trip.end_mileage = request.POST.get('end_mileage')
            
            trip.save()
            
            # Update vehicle mileage if trip completed
            if trip.status == 'completed' and trip.end_mileage:
                vehicle = trip.vehicle
                vehicle.current_mileage = trip.end_mileage
                vehicle.save()
            
            messages.success(request, 'Work ticket updated successfully!')
            return redirect('trip_list')
            
        except Exception as e:
            messages.error(request, f'Error updating work ticket: {str(e)}')
    
    vehicles = FleetVehicle.objects.filter(status='active')
    drivers = User.objects.filter(is_active=True, is_active_staff=True)
    
    context = {
        'trip': trip,
        'vehicles': vehicles,
        'drivers': drivers,
    }
    
    return render(request, 'fleet/trip_form.html', context)


@login_required
def trip_delete(request, trip_number):
    """Delete trip/work ticket"""
    trip = get_object_or_404(VehicleTrip, trip_number=trip_number)
    
    if request.method == 'POST':
        try:
            trip.delete()
            messages.success(request, f'Work ticket {trip_number} deleted successfully!')
            return redirect('trip_list')
        except Exception as e:
            messages.error(request, f'Error deleting work ticket: {str(e)}')
    
    return redirect('trip_list')


# facilities/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, Case, When, IntegerField
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from .models import (
    Facility, FacilityUnit, FacilityTenancy, FacilityBooking,
    FacilityCategory, SubCounty, Ward, Department, Citizen, User, Payment
)


# ============================================================================
# ALL FACILITIES
# ============================================================================

@login_required
def facility_list(request):
    """List all facilities with search and filters"""
    facilities = Facility.objects.select_related(
        'category', 'sub_county', 'ward', 'managed_by'
    ).all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        facilities = facilities.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(physical_address__icontains=search_query)
        )
    
    # Filters
    facility_type = request.GET.get('facility_type', '')
    category = request.GET.get('category', '')
    sub_county = request.GET.get('sub_county', '')
    status = request.GET.get('status', '')
    
    if facility_type:
        facilities = facilities.filter(facility_type=facility_type)
    if category:
        facilities = facilities.filter(category_id=category)
    if sub_county:
        facilities = facilities.filter(sub_county_id=sub_county)
    if status:
        is_active = status == 'active'
        facilities = facilities.filter(is_active=is_active)
    
    # Statistics
    stats = {
        'total': Facility.objects.count(),
        'active': Facility.objects.filter(is_active=True).count(),
        'inactive': Facility.objects.filter(is_active=False).count(),
        'markets': Facility.objects.filter(facility_type='market').count(),
        'stadiums': Facility.objects.filter(facility_type='stadium').count(),
        'housing': Facility.objects.filter(facility_type='housing').count(),
    }
    
    # Pagination
    paginator = Paginator(facilities, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'current_facility_type': facility_type,
        'current_category': category,
        'current_sub_county': sub_county,
        'current_status': status,
        'categories': FacilityCategory.objects.filter(is_active=True),
        'sub_counties': SubCounty.objects.filter(is_active=True),
    }
    
    return render(request, 'facilities/facility_list.html', context)


@login_required
def facility_detail(request, code):
    """View facility details"""
    facility = get_object_or_404(
        Facility.objects.select_related(
            'category', 'sub_county', 'ward', 'managed_by'
        ),
        code=code
    )
    
    # Get units statistics
    units_stats = FacilityUnit.objects.filter(facility=facility).aggregate(
        total_units=Count('id'),
        vacant=Count(Case(When(status='vacant', then=1), output_field=IntegerField())),
        occupied=Count(Case(When(status='occupied', then=1), output_field=IntegerField())),
        maintenance=Count(Case(When(status='maintenance', then=1), output_field=IntegerField())),
    )
    
    # Get recent bookings
    recent_bookings = FacilityBooking.objects.filter(
        facility=facility
    ).select_related('customer').order_by('-created_at')[:10]
    
    # Get active tenancies
    active_tenancies = FacilityTenancy.objects.filter(
        unit__facility=facility,
        status='active'
    ).select_related('unit', 'tenant').order_by('-start_date')[:10]
    
    context = {
        'facility': facility,
        'units_stats': units_stats,
        'recent_bookings': recent_bookings,
        'active_tenancies': active_tenancies,
    }
    
    return render(request, 'facilities/facility_detail.html', context)


@login_required
def facility_create(request):
    """Create new facility"""
    if request.method == 'POST':
        try:
            facility = Facility.objects.create(
                name=request.POST.get('name'),
                code=request.POST.get('code'),
                facility_type=request.POST.get('facility_type'),
                category_id=request.POST.get('category'),
                sub_county_id=request.POST.get('sub_county'),
                ward_id=request.POST.get('ward'),
                physical_address=request.POST.get('physical_address'),
                description=request.POST.get('description', ''),
                capacity=request.POST.get('capacity', 0),
                is_active=request.POST.get('is_active') == 'on',
                managed_by_id=request.POST.get('managed_by') if request.POST.get('managed_by') else None,
            )
            
            messages.success(request, f'Facility {facility.name} created successfully!')
            return redirect('facility_detail', code=facility.code)
            
        except Exception as e:
            messages.error(request, f'Error creating facility: {str(e)}')
    
    context = {
        'categories': FacilityCategory.objects.filter(is_active=True),
        'sub_counties': SubCounty.objects.filter(is_active=True),
        'wards': Ward.objects.filter(is_active=True),
        'managers': User.objects.filter(is_active=True),
        'form_title': 'Add New Facility',
        'submit_text': 'Create Facility',
    }
    
    return render(request, 'facilities/facility_form.html', context)


@login_required
def facility_update(request, code):
    """Update facility"""
    facility = get_object_or_404(Facility, code=code)
    
    if request.method == 'POST':
        try:
            facility.name = request.POST.get('name')
            facility.code = request.POST.get('code')
            facility.facility_type = request.POST.get('facility_type')
            facility.category_id = request.POST.get('category')
            facility.sub_county_id = request.POST.get('sub_county')
            facility.ward_id = request.POST.get('ward')
            facility.physical_address = request.POST.get('physical_address')
            facility.description = request.POST.get('description', '')
            facility.capacity = request.POST.get('capacity', 0)
            facility.is_active = request.POST.get('is_active') == 'on'
            
            if request.POST.get('managed_by'):
                facility.managed_by_id = request.POST.get('managed_by')
            else:
                facility.managed_by = None
                
            facility.save()
            
            messages.success(request, f'Facility {facility.name} updated successfully!')
            return redirect('facility_detail', code=facility.code)
            
        except Exception as e:
            messages.error(request, f'Error updating facility: {str(e)}')
    
    context = {
        'facility': facility,
        'categories': FacilityCategory.objects.filter(is_active=True),
        'sub_counties': SubCounty.objects.filter(is_active=True),
        'wards': Ward.objects.filter(is_active=True),
        'managers': User.objects.filter(is_active=True),
        'form_title': 'Edit Facility',
        'submit_text': 'Update Facility',
    }
    
    return render(request, 'facilities/facility_form.html', context)


@login_required
def facility_delete(request, code):
    """Delete facility"""
    facility = get_object_or_404(Facility, code=code)
    
    try:
        facility_name = facility.name
        facility.delete()
        messages.success(request, f'Facility {facility_name} deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error deleting facility: {str(e)}')
    
    return redirect('facility_list')


@login_required
def facility_export_excel(request):
    """Export facilities to Excel"""
    facilities = Facility.objects.select_related(
        'category', 'sub_county', 'ward', 'managed_by'
    ).all()
    
    # Apply same filters as list view
    search_query = request.GET.get('search', '')
    if search_query:
        facilities = facilities.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    facility_type = request.GET.get('facility_type', '')
    category = request.GET.get('category', '')
    sub_county = request.GET.get('sub_county', '')
    status = request.GET.get('status', '')
    
    if facility_type:
        facilities = facilities.filter(facility_type=facility_type)
    if category:
        facilities = facilities.filter(category_id=category)
    if sub_county:
        facilities = facilities.filter(sub_county_id=sub_county)
    if status:
        is_active = status == 'active'
        facilities = facilities.filter(is_active=is_active)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Facilities"
    
    # Styles
    header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Facility Code', 'Name', 'Type', 'Category', 'Sub County', 
        'Ward', 'Physical Address', 'Capacity', 'Managed By', 'Status'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data
    for row, facility in enumerate(facilities, 2):
        ws.cell(row=row, column=1).value = facility.code
        ws.cell(row=row, column=2).value = facility.name
        ws.cell(row=row, column=3).value = facility.get_facility_type_display()
        ws.cell(row=row, column=4).value = facility.category.name
        ws.cell(row=row, column=5).value = facility.sub_county.name
        ws.cell(row=row, column=6).value = facility.ward.name
        ws.cell(row=row, column=7).value = facility.physical_address
        ws.cell(row=row, column=8).value = facility.capacity
        ws.cell(row=row, column=9).value = facility.managed_by.get_full_name() if facility.managed_by else '-'
        ws.cell(row=row, column=10).value = 'Active' if facility.is_active else 'Inactive'
        
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = border
    
    # Adjust column widths
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=facilities_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response


# ============================================================================
# MARKETS & STALLS
# ============================================================================

from django.db.models import Count, Q

@login_required
def market_list(request):
    """List market facilities"""
    markets = Facility.objects.filter(
        facility_type='market'
    ).select_related('category', 'sub_county', 'ward', 'managed_by')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        markets = markets.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    # Filters
    sub_county = request.GET.get('sub_county', '')
    status = request.GET.get('status', '')
    
    if sub_county:
        markets = markets.filter(sub_county_id=sub_county)
    if status:
        is_active = status == 'active'
        markets = markets.filter(is_active=is_active)
    

    # ✅ Annotate units & occupancy
    markets = markets.annotate(
        total_units=Count('units'),
        occupied_units=Count('units', filter=Q(units__status='occupied')),
        vacant_units=Count('units', filter=Q(units__status='vacant')),
    )

    
    # Statistics
    stats = {
        'total_markets': markets.count(),
        'active': markets.filter(is_active=True).count(),
        'total_stalls': FacilityUnit.objects.filter(
            facility__facility_type='market',
            unit_type='stall'
        ).count(),
        'occupied_stalls': FacilityUnit.objects.filter(
            facility__facility_type='market',
            unit_type='stall',
            status='occupied'
        ).count(),
        'vacant_stalls': FacilityUnit.objects.filter(
            facility__facility_type='market',
            unit_type='stall',
            status='vacant'
        ).count(),
    }
    
    # Pagination
    paginator = Paginator(markets, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'current_sub_county': sub_county,
        'current_status': status,
        'sub_counties': SubCounty.objects.filter(is_active=True),
    }
    
    return render(request, 'facilities/market_list.html', context)

@login_required
def stall_list(request):
    """List all stalls"""
    stalls = FacilityUnit.objects.filter(
        unit_type='stall'
    ).select_related('facility', 'current_tenant')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        stalls = stalls.filter(
            Q(unit_number__icontains=search_query) |
            Q(facility__name__icontains=search_query)
        )
    
    # Filters
    facility = request.GET.get('facility', '')
    status = request.GET.get('status', '')
    
    if facility:
        stalls = stalls.filter(facility_id=facility)
    if status:
        stalls = stalls.filter(status=status)
    
    # Statistics
    stats = {
        'total': FacilityUnit.objects.filter(unit_type='stall').count(),
        'vacant': FacilityUnit.objects.filter(unit_type='stall', status='vacant').count(),
        'occupied': FacilityUnit.objects.filter(unit_type='stall', status='occupied').count(),
        'maintenance': FacilityUnit.objects.filter(unit_type='stall', status='maintenance').count(),
    }
    
    # Pagination
    paginator = Paginator(stalls, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'current_facility': facility,
        'current_status': status,
        'facilities': Facility.objects.filter(facility_type='market', is_active=True),
    }
    
    return render(request, 'facilities/stall_list.html', context)


# ============================================================================
# HOUSING
# ============================================================================

@login_required
def housing_list(request):
    """List housing facilities"""
    housing = Facility.objects.filter(
        facility_type='housing'
    ).select_related('category', 'sub_county', 'ward', 'managed_by')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        housing = housing.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    # Filters
    sub_county = request.GET.get('sub_county', '')
    status = request.GET.get('status', '')
    
    if sub_county:
        housing = housing.filter(sub_county_id=sub_county)
    if status:
        is_active = status == 'active'
        housing = housing.filter(is_active=is_active)
    
    # Statistics
    stats = {
        'total_estates': housing.count(),
        'total_houses': FacilityUnit.objects.filter(
            facility__facility_type='housing',
            unit_type='house'
        ).count(),
        'occupied': FacilityUnit.objects.filter(
            facility__facility_type='housing',
            unit_type='house',
            status='occupied'
        ).count(),
        'vacant': FacilityUnit.objects.filter(
            facility__facility_type='housing',
            unit_type='house',
            status='vacant'
        ).count(),
    }
    
    # Pagination
    paginator = Paginator(housing, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'current_sub_county': sub_county,
        'current_status': status,
        'sub_counties': SubCounty.objects.filter(is_active=True),
    }
    
    return render(request, 'facilities/housing_list.html', context)


# ============================================================================
# BOOKINGS
# ============================================================================

@login_required
def booking_list(request):
    """List facility bookings"""
    bookings = FacilityBooking.objects.select_related(
        'facility', 'customer', 'payment', 'approved_by'
    ).all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        bookings = bookings.filter(
            Q(booking_number__icontains=search_query) |
            Q(facility__name__icontains=search_query) |
            Q(customer__first_name__icontains=search_query) |
            Q(customer__last_name__icontains=search_query)
        )
    
    # Filters
    facility = request.GET.get('facility', '')
    status = request.GET.get('status', '')
    
    if facility:
        bookings = bookings.filter(facility_id=facility)
    if status:
        bookings = bookings.filter(status=status)
    
    # Statistics
    stats = {
        'total': FacilityBooking.objects.count(),
        'pending': FacilityBooking.objects.filter(status='pending').count(),
        'confirmed': FacilityBooking.objects.filter(status='confirmed').count(),
        'completed': FacilityBooking.objects.filter(status='completed').count(),
        'cancelled': FacilityBooking.objects.filter(status='cancelled').count(),
    }
    
    # Pagination
    paginator = Paginator(bookings, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'current_facility': facility,
        'current_status': status,
        'facilities': Facility.objects.filter(is_active=True),
    }
    
    return render(request, 'facilities/booking_list.html', context)


@login_required
def booking_detail(request, booking_number):
    """View booking details"""
    booking = get_object_or_404(
        FacilityBooking.objects.select_related(
            'facility', 'customer', 'payment', 'approved_by'
        ),
        booking_number=booking_number
    )
    
    context = {
        'booking': booking,
    }
    
    return render(request, 'facilities/booking_detail.html', context)


# ============================================================================
# TENANCIES
# ============================================================================

@login_required
def tenancy_list(request):
    """List facility tenancies"""
    tenancies = FacilityTenancy.objects.select_related(
        'unit', 'unit__facility', 'tenant', 'created_by'
    ).all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        tenancies = tenancies.filter(
            Q(tenancy_number__icontains=search_query) |
            Q(tenant__first_name__icontains=search_query) |
            Q(tenant__last_name__icontains=search_query) |
            Q(unit__unit_number__icontains=search_query)
        )
    
    # Filters
    facility = request.GET.get('facility', '')
    status = request.GET.get('status', '')
    
    if facility:
        tenancies = tenancies.filter(unit__facility_id=facility)
    if status:
        tenancies = tenancies.filter(status=status)
    
    # Statistics
    stats = {
        'total': FacilityTenancy.objects.count(),
        'active': FacilityTenancy.objects.filter(status='active').count(),
        'expired': FacilityTenancy.objects.filter(status='expired').count(),
        'terminated': FacilityTenancy.objects.filter(status='terminated').count(),
    }
    
    # Pagination
    paginator = Paginator(tenancies, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'current_facility': facility,
        'current_status': status,
        'facilities': Facility.objects.filter(is_active=True),
    }
    
    return render(request, 'facilities/tenancy_list.html', context)


@login_required
def tenancy_detail(request, tenancy_number):
    """View tenancy details"""
    tenancy = get_object_or_404(
        FacilityTenancy.objects.select_related(
            'unit', 'unit__facility', 'tenant', 'created_by'
        ),
        tenancy_number=tenancy_number
    )
    
    context = {
        'tenancy': tenancy,
    }
    
    return render(request, 'facilities/tenancy_detail.html', context)


@login_required
def tenancy_export_excel(request):
    """Export tenancies to Excel"""
    tenancies = FacilityTenancy.objects.select_related(
        'unit', 'unit__facility', 'tenant', 'created_by'
    ).all()
    
    # Apply filters
    facility = request.GET.get('facility', '')
    status = request.GET.get('status', '')
    
    if facility:
        tenancies = tenancies.filter(unit__facility_id=facility)
    if status:
        tenancies = tenancies.filter(status=status)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tenancies"
    
    # Styles
    header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Tenancy Number', 'Facility', 'Unit', 'Tenant', 'Phone',
        'Start Date', 'End Date', 'Rental Amount', 'Payment Frequency', 'Status'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data
    for row, tenancy in enumerate(tenancies, 2):
        ws.cell(row=row, column=1).value = tenancy.tenancy_number
        ws.cell(row=row, column=2).value = tenancy.unit.facility.name
        ws.cell(row=row, column=3).value = tenancy.unit.unit_number
        ws.cell(row=row, column=4).value = str(tenancy.tenant)
        ws.cell(row=row, column=5).value = tenancy.tenant.phone_primary
        ws.cell(row=row, column=6).value = tenancy.start_date.strftime('%Y-%m-%d')
        ws.cell(row=row, column=7).value = tenancy.end_date.strftime('%Y-%m-%d') if tenancy.end_date else '-'
        ws.cell(row=row, column=8).value = float(tenancy.rental_amount)
        ws.cell(row=row, column=9).value = tenancy.payment_frequency
        ws.cell(row=row, column=10).value = tenancy.get_status_display()
        
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = border
    
    # Adjust column widths
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=tenancies_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json

from .models import (
    User, Role, Permission, UserRole, Notification, 
    AuditLog, SystemConfiguration, Department, SubCounty
)
from .forms import (
    UserForm, RoleForm, PermissionForm, 
    SystemConfigurationForm
)


# ============================================================================
# USERS MANAGEMENT
# ============================================================================

@login_required
def user_list(request):
    """List all users with search and filters"""
    users = User.objects.select_related('department', 'sub_county').all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(employee_number__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )
    
    # Filters
    department = request.GET.get('department', '')
    if department:
        users = users.filter(department_id=department)
    
    sub_county = request.GET.get('sub_county', '')
    if sub_county:
        users = users.filter(sub_county_id=sub_county)
    
    is_active = request.GET.get('is_active', '')
    if is_active:
        users = users.filter(is_active=is_active == 'true')
    
    is_staff = request.GET.get('is_staff', '')
    if is_staff:
        users = users.filter(is_staff=is_staff == 'true')
    
    # Statistics
    stats = {
        'total': User.objects.count(),
        'active': User.objects.filter(is_active=True).count(),
        'inactive': User.objects.filter(is_active=False).count(),
        'staff': User.objects.filter(is_staff=True).count(),
        'superusers': User.objects.filter(is_superuser=True).count(),
    }
    
    # Pagination
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'departments': Department.objects.filter(is_active=True),
        'sub_counties': SubCounty.objects.filter(is_active=True),
        'current_department': department,
        'current_sub_county': sub_county,
        'current_is_active': is_active,
        'current_is_staff': is_staff,
    }
    
    return render(request, 'users/user_list.html', context)


@login_required
def user_detail(request, pk):
    """User detail view"""
    user = get_object_or_404(User, pk=pk)
    
    # Get user roles
    user_roles = UserRole.objects.filter(
        user=user, 
        is_active=True
    ).select_related('role')
    
    # Get recent activity
    recent_activity = AuditLog.objects.filter(
        user=user
    ).order_by('-timestamp')[:10]
    
    # Get login history
    login_attempts = LoginAttempt.objects.filter(
        username=user.username
    ).order_by('-timestamp')[:10]

    
    context = {
        'user': user,
        'user_roles': user_roles,
        'recent_activity': recent_activity,
        'login_attempts': login_attempts,
    }
    
    return render(request, 'users/user_detail.html', context)


@login_required
def user_create(request):
    """Create new user"""
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.save()
            
            # Assign roles if selected
            roles = request.POST.getlist('roles')
            for role_id in roles:
                UserRole.objects.create(
                    user=user,
                    role_id=role_id,
                    assigned_by=request.user
                )
            
            # Log action
            AuditLog.objects.create(
                user=request.user,
                action='create',
                model_name='User',
                object_id=str(user.id),
                object_repr=str(user),
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'User {user.username} created successfully!')
            return redirect('user_detail', pk=user.pk)
    else:
        form = UserForm()
    
    context = {
        'form': form,
        'roles': Role.objects.filter(is_active=True),
        'departments': Department.objects.filter(is_active=True),
        'sub_counties': SubCounty.objects.filter(is_active=True),
    }
    
    return render(request, 'users/user_form.html', context)


@login_required
def user_update(request, pk):
    """Update user"""
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            user = form.save()
            
            # Update roles
            UserRole.objects.filter(user=user).update(is_active=False)
            roles = request.POST.getlist('roles')
            for role_id in roles:
                UserRole.objects.update_or_create(
                    user=user,
                    role_id=role_id,
                    defaults={
                        'is_active': True,
                        'assigned_by': request.user
                    }
                )
            
            # Log action
            AuditLog.objects.create(
                user=request.user,
                action='update',
                model_name='User',
                object_id=str(user.id),
                object_repr=str(user),
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'User {user.username} updated successfully!')
            return redirect('user_detail', pk=user.pk)
    else:
        form = UserForm(instance=user)
    
    # Get current roles
    current_roles = user.userrole_set.filter(
        is_active=True
    ).values_list('role_id', flat=True)
    
    context = {
        'form': form,
        'user': user,
        'roles': Role.objects.filter(is_active=True),
        'current_roles': list(current_roles),
        'departments': Department.objects.filter(is_active=True),
        'sub_counties': SubCounty.objects.filter(is_active=True),
    }
    
    return render(request, 'users/user_form.html', context)


@login_required
def user_delete(request, pk):
    """Delete user"""
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        username = user.username
        
        # Log action before deletion
        AuditLog.objects.create(
            user=request.user,
            action='delete',
            model_name='User',
            object_id=str(user.id),
            object_repr=str(user),
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        user.delete()
        messages.success(request, f'User {username} deleted successfully!')
        return redirect('user_list')
    
    return render(request, 'users/user_confirm_delete.html', {'user': user})


@login_required
def user_export_excel(request):
    """Export users to Excel"""
    users = User.objects.select_related('department', 'sub_county').all()
    
    # Apply filters from request
    search_query = request.GET.get('search', '')
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    
    # Headers
    headers = [
        'Username', 'Full Name', 'Email', 'Phone Number',
        'Employee Number', 'Department', 'Sub County',
        'Is Active', 'Is Staff', 'Date Joined'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.username)
        ws.cell(row=row, column=2, value=user.get_full_name())
        ws.cell(row=row, column=3, value=user.email)
        ws.cell(row=row, column=4, value=user.phone_number)
        ws.cell(row=row, column=5, value=user.employee_number)
        ws.cell(row=row, column=6, value=user.department.name if user.department else '')
        ws.cell(row=row, column=7, value=user.sub_county.name if user.sub_county else '')
        ws.cell(row=row, column=8, value='Yes' if user.is_active else 'No')
        ws.cell(row=row, column=9, value='Yes' if user.is_staff else 'No')
        ws.cell(row=row, column=10, value=user.date_joined.strftime('%Y-%m-%d'))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=users_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


# ============================================================================
# ROLES MANAGEMENT
# ============================================================================

@login_required
def role_list(request):
    """List all roles"""
    roles = Role.objects.annotate(
        user_count=Count('userrole', filter=Q(userrole__is_active=True))
    ).prefetch_related('permissions')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        roles = roles.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Filter by status
    is_active = request.GET.get('is_active', '')
    if is_active:
        roles = roles.filter(is_active=is_active == 'true')
    
    # Statistics
    stats = {
        'total': Role.objects.count(),
        'active': Role.objects.filter(is_active=True).count(),
        'inactive': Role.objects.filter(is_active=False).count(),
    }
    
    # Pagination
    paginator = Paginator(roles, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'current_is_active': is_active,
    }
    
    return render(request, 'roles/role_list.html', context)


@login_required
def role_detail(request, pk):
    """Role detail view"""
    role = get_object_or_404(Role, pk=pk)
    
    # Get users with this role
    users = UserRole.objects.filter(
        role=role,
        is_active=True
    ).select_related('user')[:10]
    
    # Get permissions
    permissions = role.permissions.all()
    
    context = {
        'role': role,
        'users': users,
        'permissions': permissions,
        'user_count': UserRole.objects.filter(role=role, is_active=True).count(),
    }
    
    return render(request, 'roles/role_detail.html', context)


@login_required
def role_create(request):
    """Create new role"""
    if request.method == 'POST':
        form = RoleForm(request.POST)
        if form.is_valid():
            role = form.save()
            
            # Assign permissions
            permissions = request.POST.getlist('permissions')
            role.permissions.set(permissions)
            
            # Log action
            AuditLog.objects.create(
                user=request.user,
                action='create',
                model_name='Role',
                object_id=str(role.id),
                object_repr=str(role),
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'Role {role.name} created successfully!')
            return redirect('role_detail', pk=role.pk)
    else:
        form = RoleForm()
    
    context = {
        'form': form,
        'permissions': Permission.objects.all().order_by('module', 'name'),
    }
    
    return render(request, 'roles/role_form.html', context)


@login_required
def role_update(request, pk):
    """Update role"""
    role = get_object_or_404(Role, pk=pk)
    
    if request.method == 'POST':
        form = RoleForm(request.POST, instance=role)
        if form.is_valid():
            role = form.save()
            
            # Update permissions
            permissions = request.POST.getlist('permissions')
            role.permissions.set(permissions)
            
            # Log action
            AuditLog.objects.create(
                user=request.user,
                action='update',
                model_name='Role',
                object_id=str(role.id),
                object_repr=str(role),
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'Role {role.name} updated successfully!')
            return redirect('role_detail', pk=role.pk)
    else:
        form = RoleForm(instance=role)
    
    context = {
        'form': form,
        'role': role,
        'permissions': Permission.objects.all().order_by('module', 'name'),
        'current_permissions': list(role.permissions.values_list('id', flat=True)),
    }
    
    return render(request, 'roles/role_form.html', context)


@login_required
def role_delete(request, pk):
    """Delete role"""
    role = get_object_or_404(Role, pk=pk)
    
    if request.method == 'POST':
        role_name = role.name
        
        # Log action
        AuditLog.objects.create(
            user=request.user,
            action='delete',
            model_name='Role',
            object_id=str(role.id),
            object_repr=str(role),
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        role.delete()
        messages.success(request, f'Role {role_name} deleted successfully!')
        return redirect('role_list')
    
    return render(request, 'roles/role_confirm_delete.html', {'role': role})


# ============================================================================
# PERMISSIONS MANAGEMENT
# ============================================================================

@login_required
def permission_list(request):
    """List all permissions"""
    permissions = Permission.objects.all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        permissions = permissions.filter(
            Q(name__icontains=search_query) |
            Q(codename__icontains=search_query) |
            Q(module__icontains=search_query)
        )
    
    # Filter by module
    module = request.GET.get('module', '')
    if module:
        permissions = permissions.filter(module=module)
    
    # Get unique modules
    modules = Permission.objects.values_list('module', flat=True).distinct()
    
    # Statistics
    stats = {
        'total': Permission.objects.count(),
        'modules': len(modules),
    }
    
    # Pagination
    paginator = Paginator(permissions, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'modules': sorted(modules),
        'current_module': module,
    }
    
    return render(request, 'permissions/permission_list.html', context)


# Continue in next part due to length...
# Continuation of views.py

# ============================================================================
# NOTIFICATIONS
# ============================================================================

@login_required
def notification_list(request):
    """List user notifications"""
    notifications = Notification.objects.filter(
        recipient=request.user
    ).order_by('-created_at')
    
    # Filter
    status = request.GET.get('status', '')
    if status == 'unread':
        notifications = notifications.filter(read_at__isnull=True)
    elif status == 'read':
        notifications = notifications.filter(read_at__isnull=False)
    
    notification_type = request.GET.get('type', '')
    if notification_type:
        notifications = notifications.filter(notification_type=notification_type)
    
    # Statistics
    stats = {
        'total': Notification.objects.filter(recipient=request.user).count(),
        'unread': Notification.objects.filter(recipient=request.user, read_at__isnull=True).count(),
        'read': Notification.objects.filter(recipient=request.user, read_at__isnull=False).count(),
    }
    
    # Pagination
    paginator = Paginator(notifications, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'current_status': status,
        'current_type': notification_type,
    }
    
    return render(request, 'notifications/notification_list.html', context)


@login_required
def notification_mark_read(request, pk):
    """Mark notification as read"""
    notification = get_object_or_404(Notification, pk=pk, recipient=request.user)
    
    if not notification.read_at:
        notification.read_at = timezone.now()
        notification.save()
    
    return JsonResponse({'status': 'success'})


@login_required
def notification_mark_all_read(request):
    """Mark all notifications as read"""
    Notification.objects.filter(
        recipient=request.user,
        read_at__isnull=True
    ).update(read_at=timezone.now())
    
    messages.success(request, 'All notifications marked as read!')
    return redirect('notification_list')


# ============================================================================
# AUDIT TRAIL
# ============================================================================

@login_required
def audit_trail_list(request):
    """List audit trail logs"""
    logs = AuditLog.objects.select_related('user').order_by('-timestamp')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        logs = logs.filter(
            Q(user__username__icontains=search_query) |
            Q(model_name__icontains=search_query) |
            Q(object_repr__icontains=search_query)
        )
    
    # Filters
    action = request.GET.get('action', '')
    if action:
        logs = logs.filter(action=action)
    
    model_name = request.GET.get('model', '')
    if model_name:
        logs = logs.filter(model_name=model_name)
    
    user_id = request.GET.get('user', '')
    if user_id:
        logs = logs.filter(user_id=user_id)
    
    date_from = request.GET.get('date_from', '')
    if date_from:
        logs = logs.filter(timestamp__gte=date_from)
    
    date_to = request.GET.get('date_to', '')
    if date_to:
        logs = logs.filter(timestamp__lte=date_to)
    
    # Statistics
    stats = {
        'total': AuditLog.objects.count(),
        'today': AuditLog.objects.filter(
            timestamp__date=timezone.now().date()
        ).count(),
        'this_week': AuditLog.objects.filter(
            timestamp__gte=timezone.now() - timezone.timedelta(days=7)
        ).count(),
    }
    
    # Get unique models and actions
    models = AuditLog.objects.values_list('model_name', flat=True).distinct()
    actions = AuditLog.objects.values_list('action', flat=True).distinct()
    
    # Pagination
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'models': sorted(models),
        'actions': sorted(actions),
        'users': User.objects.filter(is_staff=True).order_by('username'),
        'current_action': action,
        'current_model': model_name,
        'current_user': user_id,
        'current_date_from': date_from,
        'current_date_to': date_to,
    }
    
    return render(request, 'audit/audit_trail_list.html', context)


@login_required
def audit_trail_export(request):
    """Export audit trail to Excel"""
    logs = AuditLog.objects.select_related('user').order_by('-timestamp')
    
    # Apply filters
    action = request.GET.get('action', '')
    if action:
        logs = logs.filter(action=action)
    
    model_name = request.GET.get('model', '')
    if model_name:
        logs = logs.filter(model_name=model_name)
    
    date_from = request.GET.get('date_from', '')
    if date_from:
        logs = logs.filter(timestamp__gte=date_from)
    
    date_to = request.GET.get('date_to', '')
    if date_to:
        logs = logs.filter(timestamp__lte=date_to)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Trail"
    
    # Headers
    headers = [
        'Timestamp', 'User', 'Action', 'Model', 
        'Object ID', 'Object', 'IP Address'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row, log in enumerate(logs[:1000], 2):  # Limit to 1000 rows
        ws.cell(row=row, column=1, value=log.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=2, value=log.user.username if log.user else 'System')
        ws.cell(row=row, column=3, value=log.action.upper())
        ws.cell(row=row, column=4, value=log.model_name)
        ws.cell(row=row, column=5, value=log.object_id)
        ws.cell(row=row, column=6, value=log.object_repr)
        ws.cell(row=row, column=7, value=log.ip_address or '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=audit_trail_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


# ============================================================================
# SYSTEM SETTINGS
# ============================================================================

@login_required
def system_settings(request):
    """System settings view"""
    settings = SystemConfiguration.objects.all().order_by('key')
    
    # Group settings by category
    grouped_settings = {}
    for setting in settings:
        category = setting.key.split('_')[0] if '_' in setting.key else 'general'
        if category not in grouped_settings:
            grouped_settings[category] = []
        grouped_settings[category].append(setting)
    
    context = {
        'grouped_settings': grouped_settings,
    }
    
    return render(request, 'settings/system_settings.html', context)


@login_required
def system_settings_update(request):
    """Update system settings"""
    if request.method == 'POST':
        for key, value in request.POST.items():
            if key != 'csrfmiddlewaretoken':
                try:
                    setting = SystemConfiguration.objects.get(key=key)
                    if setting.is_editable:
                        setting.value = value
                        setting.updated_by = request.user
                        setting.save()
                        
                        # Log action
                        AuditLog.objects.create(
                            user=request.user,
                            action='update',
                            model_name='SystemConfiguration',
                            object_id=str(setting.id),
                            object_repr=setting.key,
                            changes={'key': key, 'new_value': value},
                            ip_address=request.META.get('REMOTE_ADDR')
                        )
                except SystemConfiguration.DoesNotExist:
                    pass
        
        messages.success(request, 'Settings updated successfully!')
        return redirect('system_settings')


"""
Licenses & Permits Management Views
Handles business licenses, permits, and applications
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    Business, License, LicenseType, LicenseRequirement, 
    LicenseDocument, BusinessCategory, Citizen, SubCounty, 
    Ward, RevenueStream, User
)


# ============================================================================
# DASHBOARD & STATISTICS
# ============================================================================
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, Avg, Q
from django.utils import timezone
from datetime import timedelta
from collections import defaultdict
import json

@login_required
def licenses_dashboard(request):
    """Licenses & Permits dashboard with comprehensive statistics and analysis"""
    
    # Get date ranges
    today = timezone.now().date()
    this_month_start = today.replace(day=1)
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)
    this_year_start = today.replace(month=1, day=1)
    last_year_start = this_year_start.replace(year=this_year_start.year - 1)
    
    # ============================================================================
    # CORE LICENSE STATISTICS
    # ============================================================================
    
    total_licenses = License.objects.count()
    active_licenses = License.objects.filter(
        status='active',
        expiry_date__gte=today
    ).count()
    expired_licenses = License.objects.filter(
        expiry_date__lt=today
    ).count()
    pending_applications = License.objects.filter(
        status__in=['submitted', 'under_review']
    ).count()
    
    # Business statistics
    total_businesses = Business.objects.filter(is_active=True).count()
    businesses_with_licenses = Business.objects.filter(
        licenses__isnull=False
    ).distinct().count()
    
    # Monthly statistics
    licenses_this_month = License.objects.filter(
        issue_date__gte=this_month_start,
        issue_date__lte=today
    ).count()
    
    licenses_last_month = License.objects.filter(
        issue_date__gte=last_month_start,
        issue_date__lt=this_month_start
    ).count()
    
    # Calculate month-over-month growth
    if licenses_last_month > 0:
        mom_growth = ((licenses_this_month - licenses_last_month) / licenses_last_month) * 100
    else:
        mom_growth = 100 if licenses_this_month > 0 else 0
    
    # Expiring soon (next 30 days)
    expiring_soon = License.objects.filter(
        status='active',
        expiry_date__range=[today, today + timedelta(days=30)]
    ).count()
    
    # Renewals this month
    renewals_this_month = License.objects.filter(
        is_renewal=True,
        application_date__gte=this_month_start
    ).count()
    
    # ============================================================================
    # LICENSE TYPE DISTRIBUTION
    # ============================================================================
    
    license_type_distribution = list(
        License.objects.values('license_type__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )
    
    # Prepare data for pie chart
    license_type_labels = [item['license_type__name'] for item in license_type_distribution]
    license_type_data = [item['count'] for item in license_type_distribution]
    
    # ============================================================================
    # STATUS DISTRIBUTION
    # ============================================================================
    
    status_distribution = list(
        License.objects.values('status')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    
    status_labels = [item['status'].replace('_', ' ').title() for item in status_distribution]
    status_data = [item['count'] for item in status_distribution]
    
    # ============================================================================
    # MONTHLY TRENDS (Last 12 Months)
    # ============================================================================
    
    monthly_trends = []
    months_labels = []
    
    for i in range(11, -1, -1):
        month_date = today.replace(day=1) - timedelta(days=i*30)
        month_start = month_date.replace(day=1)
        if month_date.month == 12:
            month_end = month_date.replace(year=month_date.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = month_date.replace(month=month_date.month + 1, day=1) - timedelta(days=1)
        
        count = License.objects.filter(
            issue_date__gte=month_start,
            issue_date__lte=month_end
        ).count()
        
        monthly_trends.append(count)
        months_labels.append(month_start.strftime('%b %Y'))
    
    # ============================================================================
    # GEOGRAPHICAL DISTRIBUTION (By Sub-County)
    # ============================================================================
    
    subcounty_distribution = list(
        Business.objects.values('sub_county__name')
        .annotate(
            business_count=Count('id'),
            license_count=Count('licenses')
        )
        .order_by('-license_count')[:10]
    )
    
    subcounty_labels = [item['sub_county__name'] for item in subcounty_distribution]
    subcounty_business_data = [item['business_count'] for item in subcounty_distribution]
    subcounty_license_data = [item['license_count'] for item in subcounty_distribution]
    
    # ============================================================================
    # BUSINESS CATEGORY DISTRIBUTION
    # ============================================================================
    
    category_distribution = list(
        Business.objects.values('business_category__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:8]
    )
    
    category_labels = [item['business_category__name'] for item in category_distribution]
    category_data = [item['count'] for item in category_distribution]
    
    # ============================================================================
    # APPLICATION PROCESSING TIME ANALYSIS
    # ============================================================================
    
    # Average processing time for approved licenses
    approved_licenses = License.objects.filter(
        status__in=['approved', 'active', 'issued'],
        approval_date__isnull=False
    )
    
    processing_times = []
    for license in approved_licenses:
        if license.approval_date and license.application_date:
            days = (license.approval_date - license.application_date).days
            processing_times.append(days)
    
    avg_processing_time = sum(processing_times) / len(processing_times) if processing_times else 0
    
    # ============================================================================
    # RENEWAL RATE ANALYSIS
    # ============================================================================
    
    total_renewals = License.objects.filter(is_renewal=True).count()
    renewal_rate = (total_renewals / total_licenses * 100) if total_licenses > 0 else 0
    
    # ============================================================================
    # RECENT APPLICATIONS
    # ============================================================================
    
    recent_applications = License.objects.select_related(
        'business', 'license_type', 'business__citizen'
    ).order_by('-application_date')[:15]
    
    # ============================================================================
    # EXPIRING LICENSES (Next 90 Days)
    # ============================================================================
    
    expiring_licenses = License.objects.filter(
        status='active',
        expiry_date__range=[today, today + timedelta(days=90)]
    ).select_related('business', 'license_type').order_by('expiry_date')[:15]
    
    # ============================================================================
    # TOP REVENUE GENERATING LICENSE TYPES
    # ============================================================================
    
    # This assumes you have bill/payment data linked to licenses
    top_revenue_types = list(
        Bill.objects.filter(
            license_id__isnull=False
        ).values('revenue_stream__name')
        .annotate(total_revenue=Sum('amount_paid'))
        .order_by('-total_revenue')[:5]
    )
    
    revenue_type_labels = [item['revenue_stream__name'] for item in top_revenue_types]
    revenue_type_data = [float(item['total_revenue']) for item in top_revenue_types]
    
    # ============================================================================
    # COMPLIANCE RATE
    # ============================================================================
    
    total_active_businesses = Business.objects.filter(is_active=True).count()
    businesses_with_valid_licenses = Business.objects.filter(
        licenses__status='active',
        licenses__expiry_date__gte=today
    ).distinct().count()
    
    compliance_rate = (businesses_with_valid_licenses / total_active_businesses * 100) if total_active_businesses > 0 else 0
    
    # ============================================================================
    # QUARTERLY COMPARISON (This Year vs Last Year)
    # ============================================================================
    
    current_quarter = (today.month - 1) // 3 + 1
    quarters_this_year = []
    quarters_last_year = []
    quarter_labels = ['Q1', 'Q2', 'Q3', 'Q4']
    
    for q in range(1, 5):
        # This year
        q_start = this_year_start.replace(month=(q-1)*3+1)
        if q == 4:
            q_end = this_year_start.replace(year=this_year_start.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            q_end = this_year_start.replace(month=q*3+1) - timedelta(days=1)
        
        count_this_year = License.objects.filter(
            issue_date__gte=q_start,
            issue_date__lte=min(q_end, today)
        ).count()
        quarters_this_year.append(count_this_year)
        
        # Last year
        q_start_last = last_year_start.replace(month=(q-1)*3+1)
        if q == 4:
            q_end_last = last_year_start.replace(year=last_year_start.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            q_end_last = last_year_start.replace(month=q*3+1) - timedelta(days=1)
        
        count_last_year = License.objects.filter(
            issue_date__gte=q_start_last,
            issue_date__lte=q_end_last
        ).count()
        quarters_last_year.append(count_last_year)
    
    # ============================================================================
    # CONTEXT DATA
    # ============================================================================
    
    context = {
        # Core Statistics
        'total_licenses': total_licenses,
        'active_licenses': active_licenses,
        'expired_licenses': expired_licenses,
        'pending_applications': pending_applications,
        'total_businesses': total_businesses,
        'businesses_with_licenses': businesses_with_licenses,
        'licenses_this_month': licenses_this_month,
        'licenses_last_month': licenses_last_month,
        'mom_growth': round(mom_growth, 1),
        'expiring_soon': expiring_soon,
        'renewals_this_month': renewals_this_month,
        'avg_processing_time': round(avg_processing_time, 1),
        'renewal_rate': round(renewal_rate, 1),
        'compliance_rate': round(compliance_rate, 1),
        
        # Lists
        'recent_applications': recent_applications,
        'expiring_licenses': expiring_licenses,
        
        # Chart Data (JSON encoded for JavaScript)
        'license_type_labels': json.dumps(license_type_labels),
        'license_type_data': json.dumps(license_type_data),
        
        'status_labels': json.dumps(status_labels),
        'status_data': json.dumps(status_data),
        
        'months_labels': json.dumps(months_labels),
        'monthly_trends': json.dumps(monthly_trends),
        
        'subcounty_labels': json.dumps(subcounty_labels),
        'subcounty_business_data': json.dumps(subcounty_business_data),
        'subcounty_license_data': json.dumps(subcounty_license_data),
        
        'category_labels': json.dumps(category_labels),
        'category_data': json.dumps(category_data),
        
        'revenue_type_labels': json.dumps(revenue_type_labels),
        'revenue_type_data': json.dumps(revenue_type_data),
        
        'quarter_labels': json.dumps(quarter_labels),
        'quarters_this_year': json.dumps(quarters_this_year),
        'quarters_last_year': json.dumps(quarters_last_year),
    }
    
    return render(request, 'licenses/dashboard.html', context)

# ============================================================================
# BUSINESS MANAGEMENT
# ============================================================================

@login_required
def business_list(request):
    """List all businesses with filtering and search"""
    
    businesses = Business.objects.select_related(
        'citizen', 'business_category', 'sub_county', 'ward'
    ).order_by('-created_at')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        businesses = businesses.filter(
            Q(business_number__icontains=search_query) |
            Q(business_name__icontains=search_query) |
            Q(trading_name__icontains=search_query) |
            Q(registration_number__icontains=search_query) |
            Q(citizen__first_name__icontains=search_query) |
            Q(citizen__last_name__icontains=search_query)
        )
    
    # Filters
    category = request.GET.get('category', '')
    if category:
        businesses = businesses.filter(business_category_id=category)
    
    sub_county = request.GET.get('sub_county', '')
    if sub_county:
        businesses = businesses.filter(sub_county_id=sub_county)
    
    is_active = request.GET.get('is_active', '')
    if is_active:
        businesses = businesses.filter(is_active=(is_active == 'true'))
    
    # Statistics
    stats = {
        'total': Business.objects.count(),
        'active': Business.objects.filter(is_active=True).count(),
        'inactive': Business.objects.filter(is_active=False).count(),
    }
    
    # Pagination
    paginator = Paginator(businesses, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'categories': BusinessCategory.objects.filter(is_active=True),
        'sub_counties': SubCounty.objects.filter(is_active=True),
        'current_category': category,
        'current_sub_county': sub_county,
        'current_is_active': is_active,
    }
    
    return render(request, 'licenses/business_list.html', context)


@login_required
def business_detail(request, pk):
    """View business details"""
    
    business = get_object_or_404(
        Business.objects.select_related(
            'citizen', 'business_category', 'sub_county', 'ward', 'created_by'
        ),
        pk=pk
    )
    
    # Get all licenses for this business
    licenses = business.licenses.select_related('license_type').order_by('-issue_date')
    
    context = {
        'business': business,
        'licenses': licenses,
    }
    
    return render(request, 'licenses/business_detail.html', context)


@login_required
def business_create(request):
    """Create new business"""
    
    if request.method == 'POST':
        try:
            # Generate business number
            last_business = Business.objects.order_by('-id').first()
            if last_business and last_business.business_number:
                last_num = int(last_business.business_number.split('-')[-1])
                business_number = f"BUS-{str(last_num + 1).zfill(6)}"
            else:
                business_number = "BUS-000001"
            
            # Create business
            business = Business.objects.create(
                business_number=business_number,
                citizen_id=request.POST.get('citizen'),
                business_name=request.POST.get('business_name'),
                trading_name=request.POST.get('trading_name', ''),
                business_category_id=request.POST.get('business_category'),
                registration_number=request.POST.get('registration_number', ''),
                physical_address=request.POST.get('physical_address'),
                sub_county_id=request.POST.get('sub_county'),
                ward_id=request.POST.get('ward'),
                plot_number=request.POST.get('plot_number', ''),
                nature_of_business=request.POST.get('nature_of_business'),
                number_of_employees=request.POST.get('number_of_employees', 0),
                annual_turnover=request.POST.get('annual_turnover') or None,
                phone=request.POST.get('phone'),
                email=request.POST.get('email', ''),
                registration_date=request.POST.get('registration_date'),
                created_by=request.user
            )
            
            messages.success(request, f'Business {business.business_name} created successfully!')
            return redirect('business_detail', pk=business.pk)
            
        except Exception as e:
            messages.error(request, f'Error creating business: {str(e)}')
    
    context = {
        'citizens': Citizen.objects.filter(is_active=True).order_by('first_name', 'business_name'),
        'categories': BusinessCategory.objects.filter(is_active=True),
        'sub_counties': SubCounty.objects.filter(is_active=True),
    }
    
    return render(request, 'licenses/business_form.html', context)


@login_required
def business_update(request, pk):
    """Update business information"""
    
    business = get_object_or_404(Business, pk=pk)
    
    if request.method == 'POST':
        try:
            business.citizen_id = request.POST.get('citizen')
            business.business_name = request.POST.get('business_name')
            business.trading_name = request.POST.get('trading_name', '')
            business.business_category_id = request.POST.get('business_category')
            business.registration_number = request.POST.get('registration_number', '')
            business.physical_address = request.POST.get('physical_address')
            business.sub_county_id = request.POST.get('sub_county')
            business.ward_id = request.POST.get('ward')
            business.plot_number = request.POST.get('plot_number', '')
            business.nature_of_business = request.POST.get('nature_of_business')
            business.number_of_employees = request.POST.get('number_of_employees', 0)
            business.annual_turnover = request.POST.get('annual_turnover') or None
            business.phone = request.POST.get('phone')
            business.email = request.POST.get('email', '')
            business.registration_date = request.POST.get('registration_date')
            business.is_active = request.POST.get('is_active') == 'true'
            
            business.save()
            
            messages.success(request, f'Business {business.business_name} updated successfully!')
            return redirect('business_detail', pk=business.pk)
            
        except Exception as e:
            messages.error(request, f'Error updating business: {str(e)}')
    
    # Get wards for the business's sub_county
    wards = Ward.objects.filter(sub_county=business.sub_county, is_active=True) if business.sub_county else []
    
    context = {
        'business': business,
        'citizens': Citizen.objects.filter(is_active=True).order_by('first_name', 'business_name'),
        'categories': BusinessCategory.objects.filter(is_active=True),
        'sub_counties': SubCounty.objects.filter(is_active=True),
        'wards': wards,
    }
    
    return render(request, 'licenses/business_form.html', context)


@login_required
def business_delete(request, pk):
    """Delete business"""
    
    business = get_object_or_404(Business, pk=pk)
    
    if request.method == 'POST':
        try:
            business_name = business.business_name
            business.delete()
            messages.success(request, f'Business {business_name} deleted successfully!')
            return redirect('business_list')
        except Exception as e:
            messages.error(request, f'Error deleting business: {str(e)}')
            return redirect('business_detail', pk=pk)
    
    return redirect('business_detail', pk=pk)


@login_required
def business_export_excel(request):
    """Export businesses to Excel"""
    
    # Get filtered businesses
    businesses = Business.objects.select_related(
        'citizen', 'business_category', 'sub_county', 'ward'
    ).order_by('-created_at')
    
    # Apply same filters as list view
    search_query = request.GET.get('search', '')
    if search_query:
        businesses = businesses.filter(
            Q(business_number__icontains=search_query) |
            Q(business_name__icontains=search_query) |
            Q(trading_name__icontains=search_query)
        )
    
    category = request.GET.get('category', '')
    if category:
        businesses = businesses.filter(business_category_id=category)
    
    sub_county = request.GET.get('sub_county', '')
    if sub_county:
        businesses = businesses.filter(sub_county_id=sub_county)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Businesses"
    
    # Styling
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Business Number', 'Business Name', 'Trading Name', 'Owner',
        'Category', 'Registration Number', 'Phone', 'Email',
        'Sub County', 'Ward', 'Plot Number', 'Nature of Business',
        'Employees', 'Annual Turnover', 'Status', 'Registration Date'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data
    for row_num, business in enumerate(businesses, 2):
        owner_name = business.citizen.first_name if business.citizen.entity_type == 'individual' else business.citizen.business_name
        
        data = [
            business.business_number,
            business.business_name,
            business.trading_name,
            owner_name,
            business.business_category.name,
            business.registration_number,
            business.phone,
            business.email,
            business.sub_county.name,
            business.ward.name,
            business.plot_number,
            business.nature_of_business,
            business.number_of_employees,
            float(business.annual_turnover) if business.annual_turnover else 0,
            'Active' if business.is_active else 'Inactive',
            business.registration_date.strftime('%Y-%m-%d') if business.registration_date else ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    # Adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 18
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=businesses_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


# ============================================================================
# LICENSE MANAGEMENT
# ============================================================================

@login_required
def license_list(request):
    """List all licenses with filtering and search"""
    
    licenses = License.objects.select_related(
        'business', 'license_type', 'business__citizen', 'created_by'
    ).order_by('-application_date')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        licenses = licenses.filter(
            Q(license_number__icontains=search_query) |
            Q(business__business_name__icontains=search_query) |
            Q(business__business_number__icontains=search_query) |
            Q(license_type__name__icontains=search_query)
        )
    
    # Filters
    status = request.GET.get('status', '')
    if status:
        licenses = licenses.filter(status=status)
    
    license_type = request.GET.get('license_type', '')
    if license_type:
        licenses = licenses.filter(license_type_id=license_type)
    
    # Date range filter
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        licenses = licenses.filter(application_date__gte=date_from)
    if date_to:
        licenses = licenses.filter(application_date__lte=date_to)
    
    # Statistics
    today = timezone.now().date()
    stats = {
        'total': License.objects.count(),
        'active': License.objects.filter(status='active', expiry_date__gte=today).count(),
        'pending': License.objects.filter(status__in=['submitted', 'under_review']).count(),
        'expired': License.objects.filter(expiry_date__lt=today).count(),
        'expiring_soon': License.objects.filter(
            status='active',
            expiry_date__range=[today, today + timedelta(days=30)]
        ).count(),
    }
    
    # Pagination
    paginator = Paginator(licenses, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'license_types': LicenseType.objects.filter(is_active=True),
        'current_status': status,
        'current_license_type': license_type,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'licenses/license_list.html', context)


@login_required
def license_detail(request, pk):
    """View license details"""
    
    license = get_object_or_404(
        License.objects.select_related(
            'business', 'license_type', 'business__citizen',
            'created_by', 'reviewed_by', 'approved_by'
        ),
        pk=pk
    )
    
    # Get documents
    documents = license.documents.all()
    
    # Calculate days until expiry
    days_to_expiry = None
    if license.expiry_date:
        days_to_expiry = (license.expiry_date - timezone.now().date()).days
    
    context = {
        'license': license,
        'documents': documents,
        'days_to_expiry': days_to_expiry,
    }
    
    return render(request, 'licenses/license_detail.html', context)


@login_required
def license_create(request):
    """Create new license application"""
    
    if request.method == 'POST':
        try:
            # Generate license number
            last_license = License.objects.order_by('-id').first()
            if last_license and last_license.license_number:
                last_num = int(last_license.license_number.split('-')[-1])
                license_number = f"LIC-{str(last_num + 1).zfill(6)}"
            else:
                license_number = "LIC-000001"
            
            # Create license
            license = License.objects.create(
                license_number=license_number,
                business_id=request.POST.get('business'),
                license_type_id=request.POST.get('license_type'),
                application_date=request.POST.get('application_date'),
                is_renewal=request.POST.get('is_renewal') == 'true',
                notes=request.POST.get('notes', ''),
                created_by=request.user,
                status='draft'
            )
            
            messages.success(request, f'License application {license.license_number} created successfully!')
            return redirect('license_detail', pk=license.pk)
            
        except Exception as e:
            messages.error(request, f'Error creating license: {str(e)}')
    
    context = {
        'businesses': Business.objects.filter(is_active=True).order_by('business_name'),
        'license_types': LicenseType.objects.filter(is_active=True),
    }
    
    return render(request, 'licenses/license_form.html', context)


@login_required
def license_update(request, pk):
    """Update license application"""
    
    license = get_object_or_404(License, pk=pk)
    
    if request.method == 'POST':
        try:
            license.business_id = request.POST.get('business')
            license.license_type_id = request.POST.get('license_type')
            license.application_date = request.POST.get('application_date')
            license.is_renewal = request.POST.get('is_renewal') == 'true'
            license.notes = request.POST.get('notes', '')
            
            # Update dates if provided
            if request.POST.get('approval_date'):
                license.approval_date = request.POST.get('approval_date')
            if request.POST.get('issue_date'):
                license.issue_date = request.POST.get('issue_date')
            if request.POST.get('expiry_date'):
                license.expiry_date = request.POST.get('expiry_date')
            
            # Update status if provided
            if request.POST.get('status'):
                license.status = request.POST.get('status')
            
            license.save()
            
            messages.success(request, f'License {license.license_number} updated successfully!')
            return redirect('license_detail', pk=license.pk)
            
        except Exception as e:
            messages.error(request, f'Error updating license: {str(e)}')
    
    context = {
        'license': license,
        'businesses': Business.objects.filter(is_active=True).order_by('business_name'),
        'license_types': LicenseType.objects.filter(is_active=True),
    }
    
    return render(request, 'licenses/license_form.html', context)


@login_required
def license_approve(request, pk):
    """Approve a license application"""
    
    license = get_object_or_404(License, pk=pk)
    
    if request.method == 'POST':
        try:
            license.status = 'approved'
            license.approval_date = timezone.now().date()
            license.approved_by = request.user
            
            # Set issue date and expiry date
            license.issue_date = timezone.now().date()
            validity_days = license.license_type.validity_period_days
            license.expiry_date = license.issue_date + timedelta(days=validity_days)
            
            license.save()
            
            messages.success(request, f'License {license.license_number} approved successfully!')
            
        except Exception as e:
            messages.error(request, f'Error approving license: {str(e)}')
    
    return redirect('license_detail', pk=pk)


@login_required
def license_reject(request, pk):
    """Reject a license application"""
    
    license = get_object_or_404(License, pk=pk)
    
    if request.method == 'POST':
        try:
            license.status = 'rejected'
            license.rejection_reason = request.POST.get('rejection_reason', '')
            license.reviewed_by = request.user
            license.save()
            
            messages.success(request, f'License {license.license_number} rejected!')
            
        except Exception as e:
            messages.error(request, f'Error rejecting license: {str(e)}')
    
    return redirect('license_detail', pk=pk)


@login_required
def license_delete(request, pk):
    """Delete license"""
    
    license = get_object_or_404(License, pk=pk)
    
    if request.method == 'POST':
        try:
            license_number = license.license_number
            license.delete()
            messages.success(request, f'License {license_number} deleted successfully!')
            return redirect('license_list')
        except Exception as e:
            messages.error(request, f'Error deleting license: {str(e)}')
            return redirect('license_detail', pk=pk)
    
    return redirect('license_detail', pk=pk)


@login_required
def license_export_excel(request):
    """Export licenses to Excel"""
    
    # Get filtered licenses
    licenses = License.objects.select_related(
        'business', 'license_type', 'business__citizen'
    ).order_by('-application_date')
    
    # Apply same filters as list view
    search_query = request.GET.get('search', '')
    if search_query:
        licenses = licenses.filter(
            Q(license_number__icontains=search_query) |
            Q(business__business_name__icontains=search_query)
        )
    
    status = request.GET.get('status', '')
    if status:
        licenses = licenses.filter(status=status)
    
    license_type = request.GET.get('license_type', '')
    if license_type:
        licenses = licenses.filter(license_type_id=license_type)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Licenses"
    
    # Styling
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'License Number', 'Business', 'License Type', 'Status',
        'Application Date', 'Approval Date', 'Issue Date', 'Expiry Date',
        'Is Renewal', 'Created By'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data
    for row_num, license in enumerate(licenses, 2):
        data = [
            license.license_number,
            license.business.business_name,
            license.license_type.name,
            license.get_status_display(),
            license.application_date.strftime('%Y-%m-%d') if license.application_date else '',
            license.approval_date.strftime('%Y-%m-%d') if license.approval_date else '',
            license.issue_date.strftime('%Y-%m-%d') if license.issue_date else '',
            license.expiry_date.strftime('%Y-%m-%d') if license.expiry_date else '',
            'Yes' if license.is_renewal else 'No',
            license.created_by.get_full_name() if license.created_by else ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    # Adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 18
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=licenses_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


# ============================================================================
# LICENSE TYPE MANAGEMENT
# ============================================================================

@login_required
def license_type_list(request):
    """List all license types"""
    
    license_types = LicenseType.objects.select_related(
        'business_category', 'revenue_stream'
    ).order_by('name')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        license_types = license_types.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    # Filter by category
    category = request.GET.get('category', '')
    if category:
        license_types = license_types.filter(business_category_id=category)
    
    # Statistics
    stats = {
        'total': LicenseType.objects.count(),
        'active': LicenseType.objects.filter(is_active=True).count(),
        'inactive': LicenseType.objects.filter(is_active=False).count(),
    }
    
    # Pagination
    paginator = Paginator(license_types, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'categories': BusinessCategory.objects.filter(is_active=True),
        'current_category': category,
    }
    
    return render(request, 'licenses/license_type_list.html', context)

@login_required
def license_type_export_excel(request):
    """Export license types to Excel"""
    
    # Get queryset with filters
    license_types = LicenseType.objects.select_related(
        'business_category', 'revenue_stream'
    ).order_by('name')
    
    # Apply search filter
    search_query = request.GET.get('search', '')
    if search_query:
        license_types = license_types.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    # Apply category filter
    category = request.GET.get('category', '')
    if category:
        license_types = license_types.filter(business_category_id=category)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "License Types"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    border_side = Side(style='thin', color='E2E8F0')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Title row
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = "LICENSE TYPES REPORT"
    title_cell.font = Font(bold=True, size=16, color="1E293B")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Metadata row
    ws.merge_cells('A2:I2')
    metadata_cell = ws['A2']
    metadata_cell.value = f"Generated on: {datetime.now().strftime('%d %B %Y at %H:%M')}"
    metadata_cell.font = Font(size=10, color="64748B")
    metadata_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add empty row
    ws.append([])
    
    # Headers
    headers = [
        'Code',
        'License Type Name',
        'Business Category',
        'Revenue Stream',
        'Validity Period (Days)',
        'Renewable',
        'Requires Inspection',
        'Status',
        'Description'
    ]
    
    ws.append(headers)
    header_row = ws[4]
    
    # Apply header styling
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Set column widths
    column_widths = {
        'A': 15,  # Code
        'B': 35,  # License Type Name
        'C': 25,  # Business Category
        'D': 30,  # Revenue Stream
        'E': 20,  # Validity Period
        'F': 12,  # Renewable
        'G': 18,  # Requires Inspection
        'H': 12,  # Status
        'I': 50,  # Description
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add data rows
    for license_type in license_types:
        row = [
            license_type.code,
            license_type.name,
            license_type.business_category.name,
            f"{license_type.revenue_stream.name} ({license_type.revenue_stream.code})",
            license_type.validity_period_days,
            'Yes' if license_type.is_renewable else 'No',
            'Yes' if license_type.requires_inspection else 'No',
            'Active' if license_type.is_active else 'Inactive',
            license_type.description,
        ]
        ws.append(row)
    
    # Apply styling to data rows
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            if cell.column in [1, 5, 6, 7, 8]:  # Center align certain columns
                cell.alignment = center_alignment
            else:
                cell.alignment = cell_alignment
            
            # Color coding for status
            if cell.column == 8:  # Status column
                if cell.value == 'Active':
                    cell.font = Font(color="065F46", bold=True)
                    cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                elif cell.value == 'Inactive':
                    cell.font = Font(color="991B1B", bold=True)
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            
            # Highlight renewable and inspection columns
            if cell.column in [6, 7] and cell.value == 'Yes':
                cell.font = Font(color="065F46", bold=True)
    
    # Add summary section
    summary_row = ws.max_row + 2
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    summary_cell = ws[f'A{summary_row}']
    summary_cell.value = "SUMMARY"
    summary_cell.font = Font(bold=True, size=12, color="1E293B")
    summary_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Summary statistics
    total_count = license_types.count()
    active_count = license_types.filter(is_active=True).count()
    inactive_count = license_types.filter(is_active=False).count()
    renewable_count = license_types.filter(is_renewable=True).count()
    inspection_count = license_types.filter(requires_inspection=True).count()
    
    summary_data = [
        ['Total License Types:', total_count],
        ['Active:', active_count],
        ['Inactive:', inactive_count],
        ['Renewable:', renewable_count],
        ['Requires Inspection:', inspection_count],
    ]
    
    for label, value in summary_data:
        summary_row += 1
        ws[f'A{summary_row}'] = label
        ws[f'B{summary_row}'] = value
        ws[f'A{summary_row}'].font = Font(bold=True, color="64748B")
        ws[f'B{summary_row}'].font = Font(bold=True, color="1E293B")
        ws[f'A{summary_row}'].alignment = Alignment(horizontal="left")
        ws[f'B{summary_row}'].alignment = Alignment(horizontal="center")
    
    # Freeze panes (freeze header rows)
    ws.freeze_panes = 'A5'
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"license_types_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from .models import (
    LicenseType, BusinessCategory, RevenueStream, 
    LicenseRequirement
)
from .forms import LicenseTypeForm, LicenseRequirementFormSet


@login_required
def license_type_create(request):
    """Create a new license type"""
    
    if request.method == 'POST':
        form = LicenseTypeForm(request.POST)
        requirement_formset = LicenseRequirementFormSet(request.POST)
        
        if form.is_valid() and requirement_formset.is_valid():
            try:
                with transaction.atomic():
                    license_type = form.save(commit=False)
                    license_type.save()
                    
                    # Save requirements
                    requirements = requirement_formset.save(commit=False)
                    for requirement in requirements:
                        requirement.license_type = license_type
                        requirement.save()
                    
                    # Delete removed requirements
                    for obj in requirement_formset.deleted_objects:
                        obj.delete()
                    
                    messages.success(
                        request, 
                        f'License type "{license_type.name}" created successfully!'
                    )
                    return redirect('license_type_detail', pk=license_type.pk)
            except Exception as e:
                messages.error(request, f'Error creating license type: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = LicenseTypeForm()
        requirement_formset = LicenseRequirementFormSet()
    
    context = {
        'form': form,
        'requirement_formset': requirement_formset,
        'categories': BusinessCategory.objects.filter(is_active=True),
        'revenue_streams': RevenueStream.objects.filter(is_active=True),
        'is_create': True,
    }
    
    return render(request, 'licenses/license_type_form.html', context)


@login_required
def license_type_detail(request, pk):
    """View license type details"""
    
    license_type = get_object_or_404(
        LicenseType.objects.select_related(
            'business_category', 'revenue_stream'
        ),
        pk=pk
    )
    
    # Get requirements
    requirements = license_type.requirements.order_by('display_order')
    
    # Get licenses using this type
    licenses = license_type.license_set.select_related(
        'business', 'business__citizen'
    ).order_by('-created_at')[:10]
    
    # Statistics
    total_licenses = license_type.license_set.count()
    active_licenses = license_type.license_set.filter(status='active').count()
    pending_licenses = license_type.license_set.filter(
        status__in=['submitted', 'under_review']
    ).count()
    
    context = {
        'license_type': license_type,
        'requirements': requirements,
        'recent_licenses': licenses,
        'stats': {
            'total_licenses': total_licenses,
            'active_licenses': active_licenses,
            'pending_licenses': pending_licenses,
        }
    }
    
    return render(request, 'licenses/license_type_detail.html', context)


@login_required
def license_type_update(request, pk):
    """Update license type"""
    
    license_type = get_object_or_404(LicenseType, pk=pk)
    
    if request.method == 'POST':
        form = LicenseTypeForm(request.POST, instance=license_type)
        requirement_formset = LicenseRequirementFormSet(
            request.POST, 
            instance=license_type
        )
        
        if form.is_valid() and requirement_formset.is_valid():
            try:
                with transaction.atomic():
                    license_type = form.save()
                    
                    # Save requirements
                    requirements = requirement_formset.save(commit=False)
                    for requirement in requirements:
                        requirement.license_type = license_type
                        requirement.save()
                    
                    # Delete removed requirements
                    for obj in requirement_formset.deleted_objects:
                        obj.delete()
                    
                    messages.success(
                        request, 
                        f'License type "{license_type.name}" updated successfully!'
                    )
                    return redirect('license_type_detail', pk=license_type.pk)
            except Exception as e:
                messages.error(request, f'Error updating license type: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = LicenseTypeForm(instance=license_type)
        requirement_formset = LicenseRequirementFormSet(instance=license_type)
    
    context = {
        'form': form,
        'requirement_formset': requirement_formset,
        'license_type': license_type,
        'categories': BusinessCategory.objects.filter(is_active=True),
        'revenue_streams': RevenueStream.objects.filter(is_active=True),
        'is_create': False,
    }
    
    return render(request, 'licenses/license_type_form.html', context)


@login_required
def license_type_delete(request, pk):
    """Delete license type"""
    
    license_type = get_object_or_404(LicenseType, pk=pk)
    
    if request.method == 'POST':
        # Check if there are active licenses using this type
        active_licenses = license_type.license_set.filter(
            status__in=['active', 'submitted', 'under_review']
        ).count()
        
        if active_licenses > 0:
            messages.error(
                request,
                f'Cannot delete this license type. There are {active_licenses} active licenses using it. '
                'Please deactivate it instead.'
            )
            return redirect('license_type_detail', pk=pk)
        
        try:
            name = license_type.name
            license_type.delete()
            messages.success(request, f'License type "{name}" deleted successfully!')
            return redirect('license_type_list')
        except Exception as e:
            messages.error(request, f'Error deleting license type: {str(e)}')
            return redirect('license_type_detail', pk=pk)
    
    # GET request - show confirmation
    context = {
        'license_type': license_type,
        'active_licenses_count': license_type.license_set.filter(
            status__in=['active', 'submitted', 'under_review']
        ).count(),
    }
    
    return render(request, 'licenses/license_type_confirm_delete.html', context)



# ============================================================================
# AJAX ENDPOINTS
# ============================================================================

@login_required
def get_wards_by_subcounty(request):
    """Get wards for a specific sub-county (AJAX)"""
    
    sub_county_id = request.GET.get('sub_county_id')
    wards = Ward.objects.filter(
        sub_county_id=sub_county_id,
        is_active=True
    ).values('id', 'name').order_by('name')
    
    return JsonResponse(list(wards), safe=False)


@login_required
def get_license_types_by_category(request):
    """Get license types for a specific business category (AJAX)"""
    
    category_id = request.GET.get('category_id')
    license_types = LicenseType.objects.filter(
        business_category_id=category_id,
        is_active=True
    ).values('id', 'name', 'code').order_by('name')
    
    return JsonResponse(list(license_types), safe=False)


@login_required
def get_license_requirements(request):
    """Get requirements for a specific license type (AJAX)"""
    
    license_type_id = request.GET.get('license_type_id')
    requirements = LicenseRequirement.objects.filter(
        license_type_id=license_type_id
    ).values('id', 'requirement_name', 'description', 'is_mandatory').order_by('display_order')
    
    return JsonResponse(list(requirements), safe=False)




    """
Property & Land Management Views
Handles properties, valuations, development applications, and subdivisions
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, Avg
from django.http import HttpResponse
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .models import (
    Property, PropertyType, LandUseType, PropertyValuation,
    DevelopmentApplication, PropertySubdivision, PropertyAmalgamation,
    PropertyCaveat, PropertyDocument, SubCounty, Ward, Citizen, User
)


# ============================================================================
# PROPERTY MANAGEMENT
# ============================================================================

@login_required
def property_list(request):
    """List all properties with filters and search"""
    properties = Property.objects.select_related(
        'owner', 'property_type', 'land_use_type', 'sub_county', 'ward'
    ).order_by('-created_at')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        properties = properties.filter(
            Q(parcel_number__icontains=search_query) |
            Q(owner__first_name__icontains=search_query) |
            Q(owner__last_name__icontains=search_query) |
            Q(owner__business_name__icontains=search_query) |
            Q(plot_number__icontains=search_query) |
            Q(building_name__icontains=search_query)
        )
    
    # Filters
    property_type = request.GET.get('property_type', '')
    if property_type:
        properties = properties.filter(property_type_id=property_type)
    
    land_use = request.GET.get('land_use', '')
    if land_use:
        properties = properties.filter(land_use_type_id=land_use)
    
    sub_county = request.GET.get('sub_county', '')
    if sub_county:
        properties = properties.filter(sub_county_id=sub_county)
    
    ward = request.GET.get('ward', '')
    if ward:
        properties = properties.filter(ward_id=ward)
    
    status = request.GET.get('status', '')
    if status:
        properties = properties.filter(status=status)
    
    # Statistics
    stats = {
        'total': Property.objects.count(),
        'active': Property.objects.filter(status='active').count(),
        'with_caveat': Property.objects.filter(has_caveat=True).count(),
        'total_area': Property.objects.aggregate(
            total=Sum('area_sqm')
        )['total'] or 0,
        'total_value': Property.objects.aggregate(
            total=Sum('assessed_value')
        )['total'] or 0,
    }
    
    # Pagination
    paginator = Paginator(properties, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'property_types': PropertyType.objects.filter(is_active=True),
        'land_use_types': LandUseType.objects.filter(is_active=True),
        'sub_counties': SubCounty.objects.filter(is_active=True),
        'wards': Ward.objects.filter(is_active=True),
        'current_property_type': property_type,
        'current_land_use': land_use,
        'current_sub_county': sub_county,
        'current_ward': ward,
        'current_status': status,
    }
    
    return render(request, 'property/property_list.html', context)


@login_required
def property_detail(request, pk):
    """Property detail view"""
    property_obj = get_object_or_404(
        Property.objects.select_related(
            'owner', 'property_type', 'land_use_type', 'sub_county', 'ward'
        ),
        pk=pk
    )
    
    # Get related data
    valuations = property_obj.valuations.order_by('-valuation_date')[:5]
    documents = property_obj.documents.order_by('-uploaded_at')[:10]
    caveats = property_obj.caveats.filter(is_active=True)
    development_apps = property_obj.development_applications.order_by('-application_date')[:5]
    
    # Ownership history
    ownership_history = property_obj.ownership_history.select_related(
        'previous_owner', 'new_owner'
    ).order_by('-transfer_date')[:10]
    
    context = {
        'property': property_obj,
        'valuations': valuations,
        'documents': documents,
        'caveats': caveats,
        'development_apps': development_apps,
        'ownership_history': ownership_history,
    }
    
    return render(request, 'property/property_detail.html', context)


@login_required
def property_create(request):
    """Create new property"""
    if request.method == 'POST':
        try:
            # Get form data
            parcel_number = request.POST.get('parcel_number')
            owner_id = request.POST.get('owner')
            property_type_id = request.POST.get('property_type')
            land_use_type_id = request.POST.get('land_use_type')
            area_sqm = request.POST.get('area_sqm')
            assessed_value = request.POST.get('assessed_value')
            sub_county_id = request.POST.get('sub_county')
            ward_id = request.POST.get('ward')
            street = request.POST.get('street', '')
            plot_number = request.POST.get('plot_number', '')
            building_name = request.POST.get('building_name', '')
            annual_rate = request.POST.get('annual_rate', 0)
            registration_date = request.POST.get('registration_date')
            
            # Create property
            property_obj = Property.objects.create(
                parcel_number=parcel_number,
                owner_id=owner_id,
                property_type_id=property_type_id,
                land_use_type_id=land_use_type_id,
                area_sqm=area_sqm,
                assessed_value=assessed_value if assessed_value else None,
                sub_county_id=sub_county_id,
                ward_id=ward_id,
                street=street,
                plot_number=plot_number,
                building_name=building_name,
                annual_rate=annual_rate if annual_rate else 0,
                registration_date=registration_date,
                status='active'
            )
            
            messages.success(request, f'Property {parcel_number} created successfully!')
            return redirect('property_detail', pk=property_obj.pk)
            
        except Exception as e:
            messages.error(request, f'Error creating property: {str(e)}')
    
    context = {
        'property_types': PropertyType.objects.filter(is_active=True),
        'land_use_types': LandUseType.objects.filter(is_active=True),
        'sub_counties': SubCounty.objects.filter(is_active=True),
        'citizens': Citizen.objects.filter(is_active=True),
    }
    
    return render(request, 'property/property_form.html', context)


@login_required
def property_update(request, pk):
    """Update property"""
    property_obj = get_object_or_404(Property, pk=pk)
    
    if request.method == 'POST':
        try:
            property_obj.parcel_number = request.POST.get('parcel_number')
            property_obj.owner_id = request.POST.get('owner')
            property_obj.property_type_id = request.POST.get('property_type')
            property_obj.land_use_type_id = request.POST.get('land_use_type')
            property_obj.area_sqm = request.POST.get('area_sqm')
            
            assessed_value = request.POST.get('assessed_value')
            property_obj.assessed_value = assessed_value if assessed_value else None
            
            property_obj.sub_county_id = request.POST.get('sub_county')
            property_obj.ward_id = request.POST.get('ward')
            property_obj.street = request.POST.get('street', '')
            property_obj.plot_number = request.POST.get('plot_number', '')
            property_obj.building_name = request.POST.get('building_name', '')
            
            annual_rate = request.POST.get('annual_rate')
            property_obj.annual_rate = annual_rate if annual_rate else 0
            
            property_obj.status = request.POST.get('status')
            property_obj.registration_date = request.POST.get('registration_date')
            
            property_obj.save()
            
            messages.success(request, f'Property {property_obj.parcel_number} updated successfully!')
            return redirect('property_detail', pk=property_obj.pk)
            
        except Exception as e:
            messages.error(request, f'Error updating property: {str(e)}')
    
    context = {
        'property': property_obj,
        'property_types': PropertyType.objects.filter(is_active=True),
        'land_use_types': LandUseType.objects.filter(is_active=True),
        'sub_counties': SubCounty.objects.filter(is_active=True),
        'wards': Ward.objects.filter(is_active=True),
        'citizens': Citizen.objects.filter(is_active=True),
    }
    
    return render(request, 'property/property_form.html', context)


@login_required
def property_delete(request, pk):
    """Delete property"""
    property_obj = get_object_or_404(Property, pk=pk)
    parcel_number = property_obj.parcel_number
    
    try:
        property_obj.delete()
        messages.success(request, f'Property {parcel_number} deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error deleting property: {str(e)}')
    
    return redirect('property_list')


@login_required
def property_export_excel(request):
    """Export properties to Excel"""
    # Get filtered queryset
    properties = Property.objects.select_related(
        'owner', 'property_type', 'land_use_type', 'sub_county', 'ward'
    ).order_by('-created_at')
    
    # Apply filters
    search_query = request.GET.get('search', '')
    if search_query:
        properties = properties.filter(
            Q(parcel_number__icontains=search_query) |
            Q(owner__first_name__icontains=search_query) |
            Q(owner__last_name__icontains=search_query)
        )
    
    property_type = request.GET.get('property_type', '')
    if property_type:
        properties = properties.filter(property_type_id=property_type)
    
    status = request.GET.get('status', '')
    if status:
        properties = properties.filter(status=status)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Properties"
    
    # Define styles
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Headers
    headers = [
        'Parcel Number', 'Owner', 'Property Type', 'Land Use', 
        'Area (sqm)', 'Assessed Value', 'Annual Rate',
        'Sub County', 'Ward', 'Plot Number', 'Status', 'Registration Date'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row, prop in enumerate(properties, 2):
        owner_name = prop.owner.first_name + ' ' + prop.owner.last_name if prop.owner.entity_type == 'individual' else prop.owner.business_name
        
        ws.cell(row=row, column=1, value=prop.parcel_number)
        ws.cell(row=row, column=2, value=owner_name)
        ws.cell(row=row, column=3, value=prop.property_type.name)
        ws.cell(row=row, column=4, value=prop.land_use_type.name)
        ws.cell(row=row, column=5, value=float(prop.area_sqm))
        ws.cell(row=row, column=6, value=float(prop.assessed_value) if prop.assessed_value else 0)
        ws.cell(row=row, column=7, value=float(prop.annual_rate))
        ws.cell(row=row, column=8, value=prop.sub_county.name)
        ws.cell(row=row, column=9, value=prop.ward.name)
        ws.cell(row=row, column=10, value=prop.plot_number)
        ws.cell(row=row, column=11, value=prop.get_status_display())
        ws.cell(row=row, column=12, value=prop.registration_date.strftime('%Y-%m-%d'))
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=properties_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response


# ============================================================================
# PROPERTY VALUATIONS
# ============================================================================

@login_required
def valuation_list(request):
    """List all property valuations"""
    valuations = PropertyValuation.objects.select_related(
        'property', 'property__owner', 'created_by'
    ).order_by('-valuation_date')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        valuations = valuations.filter(
            Q(property__parcel_number__icontains=search_query) |
            Q(valuer_name__icontains=search_query)
        )
    
    # Filters
    property_id = request.GET.get('property', '')
    if property_id:
        valuations = valuations.filter(property_id=property_id)
    
    is_current = request.GET.get('is_current', '')
    if is_current:
        valuations = valuations.filter(is_current=is_current == 'true')
    
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        valuations = valuations.filter(valuation_date__gte=date_from)
    if date_to:
        valuations = valuations.filter(valuation_date__lte=date_to)
    
    # Statistics
    stats = {
        'total': PropertyValuation.objects.count(),
        'current': PropertyValuation.objects.filter(is_current=True).count(),
        'avg_value': PropertyValuation.objects.aggregate(
            avg=Avg('total_value')
        )['avg'] or 0,
        'total_value': PropertyValuation.objects.aggregate(
            total=Sum('total_value')
        )['total'] or 0,
    }
    
    # Pagination
    paginator = Paginator(valuations, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'property/valuation_list.html', context)


@login_required
def valuation_create(request):
    """Create new property valuation"""
    if request.method == 'POST':
        try:
            property_id = request.POST.get('property')
            land_value = request.POST.get('land_value')
            improvement_value = request.POST.get('improvement_value', 0)
            total_value = float(land_value) + float(improvement_value)
            
            valuation = PropertyValuation.objects.create(
                property_id=property_id,
                valuation_date=request.POST.get('valuation_date'),
                valuation_method=request.POST.get('valuation_method'),
                land_value=land_value,
                improvement_value=improvement_value,
                total_value=total_value,
                valuer_name=request.POST.get('valuer_name'),
                is_current=request.POST.get('is_current') == 'on',
                created_by=request.user
            )
            
            # If set as current, unset others
            if valuation.is_current:
                PropertyValuation.objects.filter(
                    property_id=property_id
                ).exclude(pk=valuation.pk).update(is_current=False)
            
            messages.success(request, 'Valuation created successfully!')
            return redirect('valuation_list')
            
        except Exception as e:
            messages.error(request, f'Error creating valuation: {str(e)}')
    
    context = {
        'properties': Property.objects.filter(status='active'),
    }
    
    return render(request, 'property/valuation_form.html', context)


@login_required
def valuation_update(request, pk):
    """Update property valuation"""
    valuation = get_object_or_404(PropertyValuation, pk=pk)
    
    if request.method == 'POST':
        try:
            land_value = request.POST.get('land_value')
            improvement_value = request.POST.get('improvement_value', 0)
            total_value = float(land_value) + float(improvement_value)
            
            valuation.valuation_date = request.POST.get('valuation_date')
            valuation.valuation_method = request.POST.get('valuation_method')
            valuation.land_value = land_value
            valuation.improvement_value = improvement_value
            valuation.total_value = total_value
            valuation.valuer_name = request.POST.get('valuer_name')
            valuation.is_current = request.POST.get('is_current') == 'on'
            
            valuation.save()
            
            # If set as current, unset others
            if valuation.is_current:
                PropertyValuation.objects.filter(
                    property_id=valuation.property_id
                ).exclude(pk=valuation.pk).update(is_current=False)
            
            messages.success(request, 'Valuation updated successfully!')
            return redirect('valuation_list')
            
        except Exception as e:
            messages.error(request, f'Error updating valuation: {str(e)}')
    
    context = {
        'valuation': valuation,
        'properties': Property.objects.filter(status='active'),
    }
    
    return render(request, 'property/valuation_form.html', context)


@login_required
def valuation_delete(request, pk):
    """Delete property valuation"""
    valuation = get_object_or_404(PropertyValuation, pk=pk)
    
    try:
        valuation.delete()
        messages.success(request, 'Valuation deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error deleting valuation: {str(e)}')
    
    return redirect('valuation_list')


# ============================================================================
# DEVELOPMENT APPLICATIONS
# ============================================================================

@login_required
def development_list(request):
    """List all development applications"""
    applications = DevelopmentApplication.objects.select_related(
        'applicant', 'property', 'reviewed_by', 'approved_by'
    ).order_by('-application_date')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        applications = applications.filter(
            Q(application_number__icontains=search_query) |
            Q(applicant__first_name__icontains=search_query) |
            Q(applicant__last_name__icontains=search_query) |
            Q(property__parcel_number__icontains=search_query)
        )
    
    # Filters
    app_type = request.GET.get('application_type', '')
    if app_type:
        applications = applications.filter(application_type=app_type)
    
    status = request.GET.get('status', '')
    if status:
        applications = applications.filter(status=status)
    
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        applications = applications.filter(application_date__gte=date_from)
    if date_to:
        applications = applications.filter(application_date__lte=date_to)
    
    # Statistics
    stats = {
        'total': DevelopmentApplication.objects.count(),
        'submitted': DevelopmentApplication.objects.filter(status='submitted').count(),
        'under_review': DevelopmentApplication.objects.filter(status='under_review').count(),
        'approved': DevelopmentApplication.objects.filter(status='approved').count(),
        'rejected': DevelopmentApplication.objects.filter(status='rejected').count(),
    }
    
    # Pagination
    paginator = Paginator(applications, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'current_type': app_type,
        'current_status': status,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'property/development_list.html', context)


@login_required
def development_create(request):
    """Create new development application"""
    if request.method == 'POST':
        try:
            # Generate application number
            last_app = DevelopmentApplication.objects.order_by('-id').first()
            if last_app:
                last_num = int(last_app.application_number.split('-')[-1])
                app_number = f"DEV-{datetime.now().year}-{last_num + 1:05d}"
            else:
                app_number = f"DEV-{datetime.now().year}-00001"
            
            application = DevelopmentApplication.objects.create(
                application_number=app_number,
                applicant_id=request.POST.get('applicant'),
                property_id=request.POST.get('property'),
                application_type=request.POST.get('application_type'),
                description=request.POST.get('description'),
                proposed_use=request.POST.get('proposed_use'),
                estimated_cost=request.POST.get('estimated_cost') or None,
                floor_area=request.POST.get('floor_area') or None,
                application_date=request.POST.get('application_date'),
                status='submitted'
            )
            
            messages.success(request, f'Application {app_number} created successfully!')
            return redirect('development_list')
            
        except Exception as e:
            messages.error(request, f'Error creating application: {str(e)}')
    
    context = {
        'properties': Property.objects.filter(status='active'),
        'citizens': Citizen.objects.filter(is_active=True),
    }
    
    return render(request, 'property/development_form.html', context)


@login_required
def development_update(request, pk):
    """Update development application"""
    application = get_object_or_404(DevelopmentApplication, pk=pk)
    
    if request.method == 'POST':
        try:
            application.applicant_id = request.POST.get('applicant')
            application.property_id = request.POST.get('property')
            application.application_type = request.POST.get('application_type')
            application.description = request.POST.get('description')
            application.proposed_use = request.POST.get('proposed_use')
            application.estimated_cost = request.POST.get('estimated_cost') or None
            application.floor_area = request.POST.get('floor_area') or None
            application.application_date = request.POST.get('application_date')
            application.status = request.POST.get('status')
            
            if request.POST.get('conditions'):
                application.conditions = request.POST.get('conditions')
            
            if request.POST.get('rejection_reason'):
                application.rejection_reason = request.POST.get('rejection_reason')
            
            application.save()
            
            messages.success(request, 'Application updated successfully!')
            return redirect('development_list')
            
        except Exception as e:
            messages.error(request, f'Error updating application: {str(e)}')
    
    context = {
        'development': application,
        'properties': Property.objects.filter(status='active'),
        'citizens': Citizen.objects.filter(is_active=True),
    }
    
    return render(request, 'property/development_form.html', context)


@login_required
def development_delete(request, pk):
    """Delete development application"""
    application = get_object_or_404(DevelopmentApplication, pk=pk)
    app_number = application.application_number
    
    try:
        application.delete()
        messages.success(request, f'Application {app_number} deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error deleting application: {str(e)}')
    
    return redirect('development_list')


# ============================================================================
# PROPERTY SUBDIVISIONS
# ============================================================================

@login_required
def subdivision_list(request):
    """List all property subdivisions"""
    subdivisions = PropertySubdivision.objects.select_related(
        'parent_property', 'parent_property__owner', 'created_by'
    ).prefetch_related('child_properties').order_by('-subdivision_date')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        subdivisions = subdivisions.filter(
            Q(parent_property__parcel_number__icontains=search_query) |
            Q(approval_number__icontains=search_query) |
            Q(surveyor__icontains=search_query)
        )
    
    # Filters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        subdivisions = subdivisions.filter(subdivision_date__gte=date_from)
    if date_to:
        subdivisions = subdivisions.filter(subdivision_date__lte=date_to)
    
    # Statistics
    stats = {
        'total': PropertySubdivision.objects.count(),
        'this_year': PropertySubdivision.objects.filter(
            subdivision_date__year=datetime.now().year
        ).count(),
    }
    
    # Pagination
    paginator = Paginator(subdivisions, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'property/subdivision_list.html', context)


@login_required
def subdivision_create(request):
    """Create new property subdivision"""
    if request.method == 'POST':
        try:
            subdivision = PropertySubdivision.objects.create(
                parent_property_id=request.POST.get('parent_property'),
                subdivision_date=request.POST.get('subdivision_date'),
                approval_number=request.POST.get('approval_number'),
                surveyor=request.POST.get('surveyor'),
                notes=request.POST.get('notes', ''),
                created_by=request.user
            )
            
            # Add child properties
            child_properties = request.POST.getlist('child_properties')
            subdivision.child_properties.set(child_properties)
            
            messages.success(request, 'Subdivision created successfully!')
            return redirect('subdivision_list')
            
        except Exception as e:
            messages.error(request, f'Error creating subdivision: {str(e)}')
    
    context = {
        'properties': Property.objects.filter(status='active'),
    }
    
    return render(request, 'property/subdivision_form.html', context)


@login_required
def subdivision_update(request, pk):
    """Update property subdivision"""
    subdivision = get_object_or_404(PropertySubdivision, pk=pk)
    
    if request.method == 'POST':
        try:
            subdivision.parent_property_id = request.POST.get('parent_property')
            subdivision.subdivision_date = request.POST.get('subdivision_date')
            subdivision.approval_number = request.POST.get('approval_number')
            subdivision.surveyor = request.POST.get('surveyor')
            subdivision.notes = request.POST.get('notes', '')
            
            subdivision.save()
            
            # Update child properties
            child_properties = request.POST.getlist('child_properties')
            subdivision.child_properties.set(child_properties)
            
            messages.success(request, 'Subdivision updated successfully!')
            return redirect('subdivision_list')
            
        except Exception as e:
            messages.error(request, f'Error updating subdivision: {str(e)}')
    
    context = {
        'subdivision': subdivision,
        'properties': Property.objects.filter(status='active'),
    }
    
    return render(request, 'property/subdivision_form.html', context)


@login_required
def subdivision_delete(request, pk):
    """Delete property subdivision"""
    subdivision = get_object_or_404(PropertySubdivision, pk=pk)
    
    try:
        subdivision.delete()
        messages.success(request, 'Subdivision deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error deleting subdivision: {str(e)}')
    
    return redirect('subdivision_list')


"""
Wajir County ERP - Parking Management and Fines & Penalties Views
Complete CRUD operations with search, filter, and Excel export
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, Avg, F
from django.http import HttpResponse
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    ParkingZone, Sacco, Vehicle, ParkingPayment, ClampingRecord,
    FineCategory, Fine, FinePayment, Citizen, Payment, SubCounty, Ward,
    User, PaymentMethod, RevenueStream
)


# ============================================================================
# PARKING ZONE MANAGEMENT
# ============================================================================

@login_required
def parking_zone_list(request):
    """List all parking zones with search and filters"""
    zones = ParkingZone.objects.select_related(
        'sub_county', 'ward'
    ).annotate(
        total_payments=Count('parkingpayment'),
        active_vehicles=Count('parkingpayment', filter=Q(parkingpayment__end_date__gte=datetime.now().date()))
    ).order_by('name')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        zones = zones.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(zone_type__icontains=search_query)
        )
    
    # Filters
    zone_type = request.GET.get('zone_type', '')
    sub_county_id = request.GET.get('sub_county', '')
    status = request.GET.get('status', '')
    
    if zone_type:
        zones = zones.filter(zone_type=zone_type)
    if sub_county_id:
        zones = zones.filter(sub_county_id=sub_county_id)
    if status:
        zones = zones.filter(is_active=(status == 'active'))
    
    # Statistics
    stats = {
        'total': ParkingZone.objects.count(),
        'active': ParkingZone.objects.filter(is_active=True).count(),
        'total_capacity': ParkingZone.objects.aggregate(Sum('capacity'))['capacity__sum'] or 0,
        'total_vehicles': ParkingPayment.objects.filter(end_date__gte=datetime.now().date()).count(),
    }
    
    # Pagination
    paginator = Paginator(zones, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'zone_type': zone_type,
        'sub_county_id': sub_county_id,
        'status': status,
        'sub_counties': SubCounty.objects.filter(is_active=True),
    }
    
    return render(request, 'parking/zone_list.html', context)


@login_required
def parking_zone_export(request):
    """Export parking zones to Excel"""
    zones = ParkingZone.objects.select_related('sub_county', 'ward').order_by('code')
    
    # Apply same filters as list view
    search_query = request.GET.get('search', '')
    if search_query:
        zones = zones.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parking Zones"
    
    # Styling
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Zone Code', 'Zone Name', 'Type', 'Sub County', 'Ward', 'Capacity', 'Status']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_num, zone in enumerate(zones, 2):
        ws.cell(row=row_num, column=1).value = zone.code
        ws.cell(row=row_num, column=2).value = zone.name
        ws.cell(row=row_num, column=3).value = zone.zone_type
        ws.cell(row=row_num, column=4).value = zone.sub_county.name
        ws.cell(row=row_num, column=5).value = zone.ward.name
        ws.cell(row=row_num, column=6).value = zone.capacity
        ws.cell(row=row_num, column=7).value = "Active" if zone.is_active else "Inactive"
        
        # Apply borders
        for col_num in range(1, 8):
            ws.cell(row=row_num, column=col_num).border = border
    
    # Adjust column widths
    for col_num in range(1, 8):
        ws.column_dimensions[get_column_letter(col_num)].width = 18
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=parking_zones_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response


# ============================================================================
# CLAMPING RECORDS
# ============================================================================

@login_required
def clamping_list(request):
    """List all clamping records"""
    clampings = ClampingRecord.objects.select_related(
        'vehicle', 'vehicle__owner', 'parking_zone', 'clamped_by', 'released_by'
    ).order_by('-clamped_date')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        clampings = clampings.filter(
            Q(clamping_number__icontains=search_query) |
            Q(vehicle__registration_number__icontains=search_query) |
            Q(reason__icontains=search_query)
        )
    
    # Filters
    status = request.GET.get('status', '')
    zone_id = request.GET.get('zone', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if status:
        clampings = clampings.filter(status=status)
    if zone_id:
        clampings = clampings.filter(parking_zone_id=zone_id)
    if date_from:
        clampings = clampings.filter(clamped_date__date__gte=date_from)
    if date_to:
        clampings = clampings.filter(clamped_date__date__lte=date_to)
    
    # Statistics
    stats = {
        'total': ClampingRecord.objects.count(),
        'clamped': ClampingRecord.objects.filter(status='clamped').count(),
        'released': ClampingRecord.objects.filter(status='released').count(),
        'total_fees': ClampingRecord.objects.aggregate(Sum('total_fee'))['total_fee__sum'] or 0,
        'today': ClampingRecord.objects.filter(clamped_date__date=datetime.now().date()).count(),
    }
    
    # Pagination
    paginator = Paginator(clampings, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'status': status,
        'zone_id': zone_id,
        'date_from': date_from,
        'date_to': date_to,
        'zones': ParkingZone.objects.filter(is_active=True),
    }
    
    return render(request, 'parking/clamping_list.html', context)


@login_required
def clamping_create(request):
    """Create new clamping record"""
    if request.method == 'POST':
        try:
            # Generate clamping number
            last_clamping = ClampingRecord.objects.order_by('-id').first()
            if last_clamping:
                last_num = int(last_clamping.clamping_number.split('-')[-1])
                clamping_number = f"CLAMP-{datetime.now().year}-{last_num + 1:05d}"
            else:
                clamping_number = f"CLAMP-{datetime.now().year}-00001"
            
            clamping_fee = float(request.POST.get('clamping_fee', 0))
            towing_fee = float(request.POST.get('towing_fee', 0))
            storage_fee = float(request.POST.get('storage_fee', 0))
            total_fee = clamping_fee + towing_fee + storage_fee
            
            clamping = ClampingRecord.objects.create(
                clamping_number=clamping_number,
                vehicle_id=request.POST.get('vehicle'),
                parking_zone_id=request.POST.get('parking_zone'),
                reason=request.POST.get('reason'),
                clamped_date=datetime.now(),
                clamping_fee=clamping_fee,
                towing_fee=towing_fee,
                storage_fee=storage_fee,
                total_fee=total_fee,
                status='clamped',
                clamped_by=request.user
            )
            
            messages.success(request, f'Vehicle clamped successfully! Clamping No: {clamping_number}')
            return redirect('clamping_list')
            
        except Exception as e:
            messages.error(request, f'Error creating clamping record: {str(e)}')
    
    context = {
        'vehicles': Vehicle.objects.filter(is_active=True),
        'zones': ParkingZone.objects.filter(is_active=True),
    }
    
    return render(request, 'parking/clamping_form.html', context)


@login_required
def clamping_release(request, pk):
    """Release clamped vehicle"""
    clamping = get_object_or_404(ClampingRecord, pk=pk)
    
    if request.method == 'POST':
        try:
            clamping.status = 'released'
            clamping.released_by = request.user
            clamping.released_date = datetime.now()
            clamping.payment_id = request.POST.get('payment_id') or None
            clamping.save()
            
            messages.success(request, f'Vehicle {clamping.vehicle.registration_number} released successfully!')
            return redirect('clamping_list')
            
        except Exception as e:
            messages.error(request, f'Error releasing vehicle: {str(e)}')
    
    context = {
        'clamping': clamping,
    }
    
    return render(request, 'parking/clamping_release.html', context)


@login_required
def clamping_export(request):
    """Export clamping records to Excel"""
    clampings = ClampingRecord.objects.select_related(
        'vehicle', 'parking_zone', 'clamped_by'
    ).order_by('-clamped_date')
    
    # Apply filters
    status = request.GET.get('status', '')
    if status:
        clampings = clampings.filter(status=status)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clamping Records"
    
    # Styling
    header_fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Clamping No.', 'Vehicle', 'Parking Zone', 'Reason', 'Clamped Date', 
               'Clamping Fee', 'Towing Fee', 'Storage Fee', 'Total Fee', 'Status', 'Clamped By']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_num, clamping in enumerate(clampings, 2):
        ws.cell(row=row_num, column=1).value = clamping.clamping_number
        ws.cell(row=row_num, column=2).value = clamping.vehicle.registration_number
        ws.cell(row=row_num, column=3).value = clamping.parking_zone.name
        ws.cell(row=row_num, column=4).value = clamping.reason
        ws.cell(row=row_num, column=5).value = clamping.clamped_date.strftime('%Y-%m-%d %H:%M')
        ws.cell(row=row_num, column=6).value = float(clamping.clamping_fee)
        ws.cell(row=row_num, column=7).value = float(clamping.towing_fee)
        ws.cell(row=row_num, column=8).value = float(clamping.storage_fee)
        ws.cell(row=row_num, column=9).value = float(clamping.total_fee)
        ws.cell(row=row_num, column=10).value = clamping.get_status_display()
        ws.cell(row=row_num, column=11).value = clamping.clamped_by.get_full_name()
        
        for col_num in range(1, 12):
            ws.cell(row=row_num, column=col_num).border = border
    
    # Adjust widths
    for col_num in range(1, 12):
        ws.column_dimensions[get_column_letter(col_num)].width = 16
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=clamping_records_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response


# ============================================================================
# FINES & PENALTIES
# ============================================================================

@login_required
def fine_list(request):
    """List all fines"""
    fines = Fine.objects.select_related(
        'category', 'offender', 'issued_by', 'payment'
    ).order_by('-issued_date')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        fines = fines.filter(
            Q(fine_number__icontains=search_query) |
            Q(offender__first_name__icontains=search_query) |
            Q(offender__last_name__icontains=search_query) |
            Q(offender__business_name__icontains=search_query) |
            Q(offense_description__icontains=search_query)
        )
    
    # Filters
    category_id = request.GET.get('category', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if category_id:
        fines = fines.filter(category_id=category_id)
    if status:
        fines = fines.filter(status=status)
    if date_from:
        fines = fines.filter(issued_date__gte=date_from)
    if date_to:
        fines = fines.filter(issued_date__lte=date_to)
    
    # Statistics
    stats = {
        'total': Fine.objects.count(),
        'issued': Fine.objects.filter(status='issued').count(),
        'paid': Fine.objects.filter(status='paid').count(),
        'overdue': Fine.objects.filter(due_date__lt=datetime.now().date(), status='issued').count(),
        'total_amount': Fine.objects.aggregate(Sum('fine_amount'))['fine_amount__sum'] or 0,
        'collected': Fine.objects.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
        'outstanding': Fine.objects.aggregate(Sum('balance'))['balance__sum'] or 0,
    }
    
    # Pagination
    paginator = Paginator(fines, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'category_id': category_id,
        'status': status,
        'date_from': date_from,
        'date_to': date_to,
        'categories': FineCategory.objects.filter(is_active=True),
    }
    
    return render(request, 'fines/fine_list.html', context)


@login_required
def fine_create(request):
    """Create new fine"""
    if request.method == 'POST':
        try:
            # Generate fine number
            last_fine = Fine.objects.order_by('-id').first()
            if last_fine:
                last_num = int(last_fine.fine_number.split('-')[-1])
                fine_number = f"FINE-{datetime.now().year}-{last_num + 1:05d}"
            else:
                fine_number = f"FINE-{datetime.now().year}-00001"
            
            fine_amount = float(request.POST.get('fine_amount'))
            
            fine = Fine.objects.create(
                fine_number=fine_number,
                category_id=request.POST.get('category'),
                offender_id=request.POST.get('offender'),
                offense_description=request.POST.get('offense_description'),
                offense_date=request.POST.get('offense_date'),
                fine_amount=fine_amount,
                balance=fine_amount,
                due_date=request.POST.get('due_date'),
                status='issued',
                issued_by=request.user,
                issued_date=datetime.now().date(),
                notes=request.POST.get('notes', '')
            )
            
            messages.success(request, f'Fine {fine_number} issued successfully!')
            return redirect('fine_list')
            
        except Exception as e:
            messages.error(request, f'Error creating fine: {str(e)}')
    
    context = {
        'categories': FineCategory.objects.filter(is_active=True),
        'citizens': Citizen.objects.filter(is_active=True),
    }
    
    return render(request, 'fines/fine_form.html', context)


@login_required
def fine_update(request, pk):
    """Update fine"""
    fine = get_object_or_404(Fine, pk=pk)
    
    if request.method == 'POST':
        try:
            fine.category_id = request.POST.get('category')
            fine.offender_id = request.POST.get('offender')
            fine.offense_description = request.POST.get('offense_description')
            fine.offense_date = request.POST.get('offense_date')
            fine.fine_amount = float(request.POST.get('fine_amount'))
            fine.due_date = request.POST.get('due_date')
            fine.status = request.POST.get('status')
            fine.notes = request.POST.get('notes', '')
            
            # Recalculate balance
            fine.balance = fine.fine_amount - fine.amount_paid
            
            fine.save()
            
            messages.success(request, 'Fine updated successfully!')
            return redirect('fine_list')
            
        except Exception as e:
            messages.error(request, f'Error updating fine: {str(e)}')
    
    context = {
        'fine': fine,
        'categories': FineCategory.objects.filter(is_active=True),
        'citizens': Citizen.objects.filter(is_active=True),
    }
    
    return render(request, 'fines/fine_form.html', context)


@login_required
def fine_waive(request, pk):
    """Waive a fine"""
    fine = get_object_or_404(Fine, pk=pk)
    
    if request.method == 'POST':
        try:
            fine.status = 'waived'
            fine.notes = f"{fine.notes}\n\nWaived by {request.user.get_full_name()} on {datetime.now().strftime('%Y-%m-%d %H:%M')}. Reason: {request.POST.get('waive_reason')}"
            fine.save()
            
            messages.success(request, f'Fine {fine.fine_number} waived successfully!')
            return redirect('fine_list')
            
        except Exception as e:
            messages.error(request, f'Error waiving fine: {str(e)}')
    
    context = {
        'fine': fine,
    }
    
    return render(request, 'fines/fine_waive.html', context)


@login_required
def fine_export(request):
    """Export fines to Excel"""
    fines = Fine.objects.select_related(
        'category', 'offender', 'issued_by'
    ).order_by('-issued_date')
    
    # Apply filters
    category_id = request.GET.get('category', '')
    status = request.GET.get('status', '')
    
    if category_id:
        fines = fines.filter(category_id=category_id)
    if status:
        fines = fines.filter(status=status)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fines & Penalties"
    
    # Styling
    header_fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Fine No.', 'Category', 'Offender', 'Offense', 'Offense Date', 
               'Fine Amount', 'Amount Paid', 'Balance', 'Due Date', 'Status', 'Issued By', 'Issued Date']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_num, fine in enumerate(fines, 2):
        offender_name = (f"{fine.offender.first_name} {fine.offender.last_name}" 
                        if fine.offender.entity_type == 'individual' 
                        else fine.offender.business_name)
        
        ws.cell(row=row_num, column=1).value = fine.fine_number
        ws.cell(row=row_num, column=2).value = fine.category.name
        ws.cell(row=row_num, column=3).value = offender_name
        ws.cell(row=row_num, column=4).value = fine.offense_description
        ws.cell(row=row_num, column=5).value = fine.offense_date.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=6).value = float(fine.fine_amount)
        ws.cell(row=row_num, column=7).value = float(fine.amount_paid)
        ws.cell(row=row_num, column=8).value = float(fine.balance)
        ws.cell(row=row_num, column=9).value = fine.due_date.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=10).value = fine.get_status_display()
        ws.cell(row=row_num, column=11).value = fine.issued_by.get_full_name() if fine.issued_by else "N/A"
        ws.cell(row=row_num, column=12).value = fine.issued_date.strftime('%Y-%m-%d')
        
        for col_num in range(1, 13):
            ws.cell(row=row_num, column=col_num).border = border
    
    # Adjust widths
    for col_num in range(1, 13):
        ws.column_dimensions[get_column_letter(col_num)].width = 16
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=fines_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response


# ============================================================================
# PARKING PAYMENTS
# ============================================================================

@login_required
def parking_payment_list(request):
    """List all parking payments"""
    payments = ParkingPayment.objects.select_related(
        'vehicle', 'vehicle__owner', 'parking_zone', 'payment'
    ).order_by('-created_at')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        payments = payments.filter(
            Q(vehicle__registration_number__icontains=search_query) |
            Q(payment__receipt_number__icontains=search_query)
        )
    
    # Filters
    payment_type = request.GET.get('payment_type', '')
    zone_id = request.GET.get('zone', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if payment_type:
        payments = payments.filter(payment_type=payment_type)
    if zone_id:
        payments = payments.filter(parking_zone_id=zone_id)
    if date_from:
        payments = payments.filter(start_date__gte=date_from)
    if date_to:
        payments = payments.filter(end_date__lte=date_to)
    
    # Statistics
    stats = {
        'total': ParkingPayment.objects.count(),
        'daily': ParkingPayment.objects.filter(payment_type='daily').count(),
        'monthly': ParkingPayment.objects.filter(payment_type='monthly').count(),
        'yearly': ParkingPayment.objects.filter(payment_type='yearly').count(),
        'total_revenue': ParkingPayment.objects.aggregate(Sum('amount'))['amount__sum'] or 0,
        'active': ParkingPayment.objects.filter(end_date__gte=datetime.now().date()).count(),
    }
    
    # Pagination
    paginator = Paginator(payments, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'payment_type': payment_type,
        'zone_id': zone_id,
        'date_from': date_from,
        'date_to': date_to,
        'zones': ParkingZone.objects.filter(is_active=True),
    }
    
    return render(request, 'parking/payment_list.html', context)


@login_required
def parking_payment_export(request):
    """Export parking payments to Excel"""
    payments = ParkingPayment.objects.select_related(
        'vehicle', 'parking_zone', 'payment'
    ).order_by('-created_at')
    
    # Apply filters
    payment_type = request.GET.get('payment_type', '')
    if payment_type:
        payments = payments.filter(payment_type=payment_type)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parking Payments"
    
    # Styling
    header_fill = PatternFill(start_color="9b59b6", end_color="9b59b6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Vehicle Reg', 'Parking Zone', 'Payment Type', 'Start Date', 
               'End Date', 'Amount', 'Receipt No.', 'Payment Date']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_num, payment in enumerate(payments, 2):
        ws.cell(row=row_num, column=1).value = payment.vehicle.registration_number
        ws.cell(row=row_num, column=2).value = payment.parking_zone.name
        ws.cell(row=row_num, column=3).value = payment.get_payment_type_display()
        ws.cell(row=row_num, column=4).value = payment.start_date.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=5).value = payment.end_date.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=6).value = float(payment.amount)
        ws.cell(row=row_num, column=7).value = payment.payment.receipt_number
        ws.cell(row=row_num, column=8).value = payment.payment.payment_date.strftime('%Y-%m-%d')
        
        for col_num in range(1, 9):
            ws.cell(row=row_num, column=col_num).border = border
    
    # Adjust widths
    for col_num in range(1, 9):
        ws.column_dimensions[get_column_letter(col_num)].width = 16
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=parking_payments_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response


# ============================================================================
# SACCO MANAGEMENT
# ============================================================================

@login_required
def sacco_list(request):
    """List all vehicle saccos"""
    saccos = Sacco.objects.select_related('citizen').annotate(
        vehicle_count=Count('vehicles')
    ).order_by('name')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        saccos = saccos.filter(
            Q(name__icontains=search_query) |
            Q(registration_number__icontains=search_query)
        )
    
    # Filter by status
    status = request.GET.get('status', '')
    if status:
        saccos = saccos.filter(is_active=(status == 'active'))
    
    # Statistics
    stats = {
        'total': Sacco.objects.count(),
        'active': Sacco.objects.filter(is_active=True).count(),
        'total_vehicles': Vehicle.objects.filter(sacco__isnull=False).count(),
    }
    
    # Pagination
    paginator = Paginator(saccos, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'status': status,
    }
    
    return render(request, 'parking/sacco_list.html', context)


@login_required
def sacco_create(request):
    """Create new sacco"""
    if request.method == 'POST':
        try:
            sacco = Sacco.objects.create(
                name=request.POST.get('name'),
                registration_number=request.POST.get('registration_number'),
                citizen_id=request.POST.get('citizen'),
                phone=request.POST.get('phone'),
                email=request.POST.get('email', ''),
                physical_address=request.POST.get('physical_address'),
                is_active=True
            )
            
            messages.success(request, f'Sacco "{sacco.name}" registered successfully!')
            return redirect('sacco_list')
            
        except Exception as e:
            messages.error(request, f'Error registering sacco: {str(e)}')
    
    context = {
        'citizens': Citizen.objects.filter(is_active=True),
    }
    
    return render(request, 'parking/sacco_form.html', context)


@login_required
def sacco_update(request, pk):
    """Update sacco information"""
    sacco = get_object_or_404(Sacco, pk=pk)
    
    if request.method == 'POST':
        try:
            sacco.name = request.POST.get('name')
            sacco.registration_number = request.POST.get('registration_number')
            sacco.citizen_id = request.POST.get('citizen')
            sacco.phone = request.POST.get('phone')
            sacco.email = request.POST.get('email', '')
            sacco.physical_address = request.POST.get('physical_address')
            sacco.is_active = request.POST.get('is_active') == 'on'
            
            sacco.save()
            
            messages.success(request, 'Sacco updated successfully!')
            return redirect('sacco_list')
            
        except Exception as e:
            messages.error(request, f'Error updating sacco: {str(e)}')
    
    context = {
        'sacco': sacco,
        'citizens': Citizen.objects.filter(is_active=True),
    }
    
    return render(request, 'parking/sacco_form.html', context)


@login_required
def sacco_export(request):
    """Export saccos to Excel"""
    saccos = Sacco.objects.select_related('citizen').annotate(
        vehicle_count=Count('vehicles')
    ).order_by('name')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicle Saccos"
    
    # Styling
    header_fill = PatternFill(start_color="1abc9c", end_color="1abc9c", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Sacco Name', 'Registration No.', 'Contact Person', 'Phone', 
               'Email', 'Address', 'Vehicles', 'Status']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_num, sacco in enumerate(saccos, 2):
        contact_person = (f"{sacco.citizen.first_name} {sacco.citizen.last_name}" 
                         if sacco.citizen.entity_type == 'individual' 
                         else sacco.citizen.business_name)
        
        ws.cell(row=row_num, column=1).value = sacco.name
        ws.cell(row=row_num, column=2).value = sacco.registration_number
        ws.cell(row=row_num, column=3).value = contact_person
        ws.cell(row=row_num, column=4).value = sacco.phone
        ws.cell(row=row_num, column=5).value = sacco.email
        ws.cell(row=row_num, column=6).value = sacco.physical_address
        ws.cell(row=row_num, column=7).value = sacco.vehicle_count
        ws.cell(row=row_num, column=8).value = "Active" if sacco.is_active else "Inactive"
        
        for col_num in range(1, 9):
            ws.cell(row=row_num, column=col_num).border = border
    
    # Adjust widths
    for col_num in range(1, 9):
        ws.column_dimensions[get_column_letter(col_num)].width = 18
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=saccos_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response


# ============================================================================
# FINE CATEGORY MANAGEMENT
# ============================================================================

@login_required
def fine_category_list(request):
    """List all fine categories"""
    categories = FineCategory.objects.select_related('revenue_stream').annotate(
        fine_count=Count('fines')
    ).order_by('name')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        categories = categories.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    # Filter by status
    status = request.GET.get('status', '')
    if status:
        categories = categories.filter(is_active=(status == 'active'))
    
    # Statistics
    stats = {
        'total': FineCategory.objects.count(),
        'active': FineCategory.objects.filter(is_active=True).count(),
    }
    
    # Pagination
    paginator = Paginator(categories, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'search_query': search_query,
        'status': status,
    }
    
    return render(request, 'fines/category_list.html', context)


@login_required
def fine_category_create(request):
    """Create new fine category"""
    if request.method == 'POST':
        try:
            category = FineCategory.objects.create(
                name=request.POST.get('name'),
                code=request.POST.get('code').upper(),
                description=request.POST.get('description'),
                revenue_stream_id=request.POST.get('revenue_stream'),
                is_active=True
            )
            
            messages.success(request, f'Fine category "{category.name}" created successfully!')
            return redirect('fine_category_list')
            
        except Exception as e:
            messages.error(request, f'Error creating fine category: {str(e)}')
    
    context = {
        'revenue_streams': RevenueStream.objects.filter(is_active=True),
    }
    
    return render(request, 'fines/category_form.html', context)


@login_required
def fine_category_update(request, pk):
    """Update fine category"""
    category = get_object_or_404(FineCategory, pk=pk)
    
    if request.method == 'POST':
        try:
            category.name = request.POST.get('name')
            category.code = request.POST.get('code').upper()
            category.description = request.POST.get('description')
            category.revenue_stream_id = request.POST.get('revenue_stream')
            category.is_active = request.POST.get('is_active') == 'on'
            
            category.save()
            
            messages.success(request, 'Fine category updated successfully!')
            return redirect('fine_category_list')
            
        except Exception as e:
            messages.error(request, f'Error updating fine category: {str(e)}')
    
    context = {
        'category': category,
        'revenue_streams': RevenueStream.objects.filter(is_active=True),
    }
    
    return render(request, 'fines/category_form.html', context)


@login_required
def fine_category_delete(request, pk):
    """Delete fine category"""
    category = get_object_or_404(FineCategory, pk=pk)
    
    if request.method == 'POST':
        try:
            category_name = category.name
            category.delete()
            messages.success(request, f'Fine category "{category_name}" deleted successfully!')
            return redirect('fine_category_list')
        except Exception as e:
            messages.error(request, f'Error deleting fine category: {str(e)}')
            return redirect('fine_category_list')
    
    context = {
        'category': category,
    }
    
    return render(request, 'fines/category_confirm_delete.html', context)


# ============================================================================
# DELETE VIEWS
# ============================================================================

from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages

@login_required
def clamping_delete(request, pk):
    clamping = get_object_or_404(ClampingRecord, pk=pk)
    try:
        clamping_number = clamping.clamping_number
        clamping.delete()
        messages.success(request, f'Clamping record {clamping_number} deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error deleting clamping record: {str(e)}')

    return redirect('clamping_list')


@login_required
def fine_delete(request, pk):
    fine = get_object_or_404(Fine, pk=pk)
    try:
        fine_number = fine.fine_number
        fine.delete()
        messages.success(request, f'Fine {fine_number} deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error deleting fine: {str(e)}')

    return redirect('fine_list')


@login_required
def sacco_delete(request, pk):
    sacco = get_object_or_404(Sacco, pk=pk)
    try:
        sacco_name = sacco.name
        sacco.delete()
        messages.success(request, f'Sacco "{sacco_name}" deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error deleting sacco: {str(e)}')

    return redirect('sacco_list')

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from .models import (
    Patient, HealthFacility, Triage, Visit, HospitalWard, Admission, 
    LabTest, Imaging, Prescription, MorgueRecord
)

def is_admin_user(user):
    """Check if user has admin privileges"""
    return user.is_authenticated and (user.is_superuser or user.is_staff)

# ============================================================================
# PATIENT MANAGEMENT VIEWS
# ============================================================================

class PatientListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Patient
    template_name = 'heallth/patient_list.html'
    context_object_name = 'patients'
    paginate_by = 20
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def get_queryset(self):
        queryset = Patient.objects.select_related('citizen').all()
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(patient_number__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(phone__icontains=search_query)
            )
        return queryset.order_by('-registered_at')

class PatientDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Patient
    template_name = 'heallth/patient_detail.html'
    context_object_name = 'patient'
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['visits'] = self.object.visits.select_related('facility').all()[:10]
        context['admissions'] = self.object.admissions.select_related('ward').all()[:5]
        return context

class PatientCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Patient
    template_name = 'heallth/patient_form.html'
    fields = [
        'first_name', 'middle_name', 'last_name', 'date_of_birth', 'gender',
        'id_number', 'phone', 'email', 'address', 'next_of_kin_name',
        'next_of_kin_phone', 'next_of_kin_relationship', 'blood_group', 'allergies'
    ]
    success_url = reverse_lazy('health:patient_list')
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def form_valid(self, form):
        # Generate patient number
        last_patient = Patient.objects.order_by('-id').first()
        if last_patient:
            last_number = int(last_patient.patient_number.split('-')[-1])
            new_number = f"PAT-{last_number + 1:06d}"
        else:
            new_number = "PAT-000001"
        
        form.instance.patient_number = new_number
        messages.success(self.request, 'Patient registered successfully!')
        return super().form_valid(form)

class PatientUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Patient
    template_name = 'heallth/patient_form.html'
    fields = [
        'first_name', 'middle_name', 'last_name', 'date_of_birth', 'gender',
        'id_number', 'phone', 'email', 'address', 'next_of_kin_name',
        'next_of_kin_phone', 'next_of_kin_relationship', 'blood_group', 'allergies', 'is_active'
    ]
    success_url = reverse_lazy('health:patient_list')
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, 'Patient updated successfully!')
        return super().form_valid(form)

# ============================================================================
# VISIT MANAGEMENT VIEWS
# ============================================================================

class VisitListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Visit
    template_name = 'heallth/visit_list.html'
    context_object_name = 'visits'
    paginate_by = 20
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def get_queryset(self):
        queryset = Visit.objects.select_related('patient', 'facility', 'attended_by').all()
        status_filter = self.request.GET.get('status', '')
        facility_filter = self.request.GET.get('facility', '')
        date_filter = self.request.GET.get('date', '')
        
        if status_filter:
            queryset = queryset.filter(is_complete=(status_filter == 'completed'))
        if facility_filter:
            queryset = queryset.filter(facility_id=facility_filter)
        if date_filter:
            queryset = queryset.filter(visit_date__date=date_filter)
            
        return queryset.order_by('-visit_date')

class VisitCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Visit
    template_name = 'heallth/visit_form.html'
    fields = ['patient', 'facility', 'visit_type', 'diagnosis', 'treatment', 'notes']
    success_url = reverse_lazy('health:visit_list')
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def form_valid(self, form):
        # Generate visit number
        last_visit = Visit.objects.order_by('-id').first()
        if last_visit:
            last_number = int(last_visit.visit_number.split('-')[-1])
            new_number = f"VIS-{last_number + 1:06d}"
        else:
            new_number = "VIS-000001"
        
        form.instance.visit_number = new_number
        form.instance.visit_date = timezone.now()
        form.instance.attended_by = self.request.user
        messages.success(self.request, 'Visit recorded successfully!')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['facilities'] = HealthFacility.objects.filter(is_active=True)
        return context

# ============================================================================
# ADMISSION MANAGEMENT VIEWS
# ============================================================================

class AdmissionListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Admission
    template_name = 'heallth/admission_list.html'
    context_object_name = 'admissions'
    paginate_by = 20
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def get_queryset(self):
        queryset = Admission.objects.select_related(
            'patient', 'ward', 'ward__facility', 'admitted_by'
        ).all()
        status_filter = self.request.GET.get('status', '')
        ward_filter = self.request.GET.get('ward', '')
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if ward_filter:
            queryset = queryset.filter(ward_id=ward_filter)
            
        return queryset.order_by('-admission_date')

class AdmissionCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Admission
    template_name = 'heallth/admission_form.html'
    fields = ['visit', 'patient', 'ward', 'bed_number', 'admission_reason']
    success_url = reverse_lazy('health:admission_list')
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def form_valid(self, form):
        # Generate admission number
        last_admission = Admission.objects.order_by('-id').first()
        if last_admission:
            last_number = int(last_admission.admission_number.split('-')[-1])
            new_number = f"ADM-{last_number + 1:06d}"
        else:
            new_number = "ADM-000001"
        
        form.instance.admission_number = new_number
        form.instance.admission_date = timezone.now()
        form.instance.admitted_by = self.request.user
        messages.success(self.request, 'Patient admitted successfully!')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['wards'] = HospitalWard.objects.select_related('facility').filter(is_active=True)
        return context

# ============================================================================
# FACILITY MANAGEMENT VIEWS
# ============================================================================

class HealthFacilityListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = HealthFacility
    template_name = 'heallth/facility_list.html'
    context_object_name = 'facilities'
    
    def test_func(self):
        return is_admin_user(self.request.user)

class HealthFacilityCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = HealthFacility
    template_name = 'heallth/facility_form.html'
    fields = [
        'name', 'code', 'facility_level', 'location', 'sub_county', 'ward',
        'phone', 'email', 'bed_capacity'
    ]
    success_url = reverse_lazy('health:facility_list')
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, 'Health facility created successfully!')
        return super().form_valid(form)

# ============================================================================
# LAB TEST MANAGEMENT VIEWS
# ============================================================================

class LabTestListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = LabTest
    template_name = 'heallth/labtest_list.html'
    context_object_name = 'lab_tests'
    paginate_by = 20
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def get_queryset(self):
        queryset = LabTest.objects.select_related('patient', 'visit').all()
        status_filter = self.request.GET.get('status', '')
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        return queryset.order_by('-requested_date')

class LabTestUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = LabTest
    template_name = 'heallth/labtest_form.html'
    fields = ['results', 'remarks', 'status', 'test_cost']
    success_url = reverse_lazy('health:labtest_list')
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def form_valid(self, form):
        if form.instance.status == 'completed' and not form.instance.completed_date:
            form.instance.completed_date = timezone.now()
            form.instance.processed_by = self.request.user
        messages.success(self.request, 'Lab test updated successfully!')
        return super().form_valid(form)

# ============================================================================
# IMAGING MANAGEMENT VIEWS
# ============================================================================

class ImagingListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Imaging
    template_name = 'heallth/imaging_list.html'
    context_object_name = 'imaging_tests'
    paginate_by = 20
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def get_queryset(self):
        queryset = Imaging.objects.select_related('patient', 'visit').all()
        status_filter = self.request.GET.get('status', '')
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        return queryset.order_by('-requested_date')

# ============================================================================
# PHARMACY MANAGEMENT VIEWS
# ============================================================================

class PrescriptionListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Prescription
    template_name = 'heallth/prescription_list.html'
    context_object_name = 'prescriptions'
    paginate_by = 20
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def get_queryset(self):
        queryset = Prescription.objects.select_related('patient', 'visit').all()
        status_filter = self.request.GET.get('status', '')
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        return queryset.order_by('-created_at')

class PrescriptionUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Prescription
    template_name = 'heallth/prescription_form.html'
    fields = ['status', 'dispensed_by', 'dispensed_date']
    success_url = reverse_lazy('health:prescription_list')
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def form_valid(self, form):
        if form.instance.status == 'dispensed' and not form.instance.dispensed_date:
            form.instance.dispensed_date = timezone.now()
        messages.success(self.request, 'Prescription updated successfully!')
        return super().form_valid(form)

# ============================================================================
# MORGUE MANAGEMENT VIEWS
# ============================================================================

class MorgueRecordListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = MorgueRecord
    template_name = 'heallth/morgue_list.html'
    context_object_name = 'morgue_records'
    paginate_by = 20
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def get_queryset(self):
        queryset = MorgueRecord.objects.select_related('facility').all()
        status_filter = self.request.GET.get('status', '')
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        return queryset.order_by('-admission_date')

class MorgueRecordCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = MorgueRecord
    template_name = 'heallth/morgue_form.html'
    fields = [
        'deceased_name', 'age', 'gender', 'date_of_death', 'cause_of_death',
        'next_of_kin_name', 'next_of_kin_phone', 'next_of_kin_relationship',
        'compartment_number', 'facility'
    ]
    success_url = reverse_lazy('health:morgue_list')
    
    def test_func(self):
        return is_admin_user(self.request.user)
    
    def form_valid(self, form):
        # Generate morgue number
        last_record = MorgueRecord.objects.order_by('-id').first()
        if last_record:
            last_number = int(last_record.morgue_number.split('-')[-1])
            new_number = f"MOR-{last_number + 1:06d}"
        else:
            new_number = "MOR-000001"
        
        form.instance.morgue_number = new_number
        form.instance.admission_date = timezone.now()
        form.instance.admitted_by = self.request.user
        messages.success(self.request, 'Morgue record created successfully!')
        return super().form_valid(form)

# ============================================================================
# DASHBOARD & REPORTS
# ============================================================================

@login_required
@user_passes_test(is_admin_user)
def health_dashboard(request):
    """Health Services Dashboard"""
    total_patients = Patient.objects.filter(is_active=True).count()
    total_facilities = HealthFacility.objects.filter(is_active=True).count()
    today_visits = Visit.objects.filter(visit_date__date=timezone.now().date()).count()
    current_admissions = Admission.objects.filter(status='admitted').count()
    
    # Recent activities
    recent_visits = Visit.objects.select_related('patient', 'facility').order_by('-visit_date')[:5]
    recent_admissions = Admission.objects.select_related('patient', 'ward').order_by('-admission_date')[:5]
    
    context = {
        'total_patients': total_patients,
        'total_facilities': total_facilities,
        'today_visits': today_visits,
        'current_admissions': current_admissions,
        'recent_visits': recent_visits,
        'recent_admissions': recent_admissions,
    }
    return render(request, 'heallth/dashboard.html', context)

@login_required
@user_passes_test(is_admin_user)
def health_reports(request):
    """Health Services Reports"""
    facility_stats = HealthFacility.objects.annotate(
        total_visits=Count('visits'),
        total_admissions=Count('admissions')
    )
    
    context = {
        'facility_stats': facility_stats,
    }
    return render(request, 'heallth/reports.html', context)


"""
Administration Management Views
Handles Geographic Setup, Departments, Case Management, and Documents
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Import models
from .models import (
    County, SubCounty, Ward, Department, Case, CaseCategory, 
    CaseDocument, CaseHearing, ElectronicDocument, DocumentCategory,
    DocumentAccess, User
)


# ============================================================================
# GEOGRAPHIC SETUP VIEWS
# ============================================================================

@login_required
def geographic_setup_dashboard(request):
    """Geographic setup overview"""
    context = {
        'counties': County.objects.count(),
        'sub_counties': SubCounty.objects.count(),
        'wards': Ward.objects.count(),
        'active_sub_counties': SubCounty.objects.filter(is_active=True).count(),
        'active_wards': Ward.objects.filter(is_active=True).count(),
    }
    return render(request, 'administration/geographic_setup.html', context)


@login_required
def sub_county_list(request):
    """List all sub-counties with search and filters"""
    sub_counties = SubCounty.objects.select_related('county').all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        sub_counties = sub_counties.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(headquarters__icontains=search_query)
        )
    
    # Filters
    county_filter = request.GET.get('county', '')
    if county_filter:
        sub_counties = sub_counties.filter(county_id=county_filter)
    
    status_filter = request.GET.get('status', '')
    if status_filter:
        is_active = status_filter == 'active'
        sub_counties = sub_counties.filter(is_active=is_active)
    
    # Pagination
    paginator = Paginator(sub_counties, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    stats = {
        'total': SubCounty.objects.count(),
        'active': SubCounty.objects.filter(is_active=True).count(),
        'inactive': SubCounty.objects.filter(is_active=False).count(),
    }
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'counties': County.objects.all(),
        'current_county': county_filter,
        'current_status': status_filter,
        'stats': stats,
    }
    return render(request, 'administration/sub_county_list.html', context)


@login_required
def sub_county_detail(request, code):
    """Sub-county detail view"""
    sub_county = get_object_or_404(SubCounty, code=code)
    
    context = {
        'sub_county': sub_county,
        'wards': sub_county.wards.all(),
        'departments': Department.objects.filter(
            Q(headed_departments__sub_county=sub_county) |
            Q(id__in=User.objects.filter(sub_county=sub_county).values('department_id'))
        ).distinct(),
        'stats': {
            'wards': sub_county.wards.count(),
            'active_wards': sub_county.wards.filter(is_active=True).count(),
            'staff': User.objects.filter(sub_county=sub_county).count(),
        }
    }
    return render(request, 'administration/sub_county_detail.html', context)


@login_required
def sub_county_create(request):
    """Create new sub-county"""
    if request.method == 'POST':
        try:
            SubCounty.objects.create(
                county_id=request.POST.get('county'),
                name=request.POST.get('name'),
                code=request.POST.get('code'),
                headquarters=request.POST.get('headquarters'),
                is_active=request.POST.get('is_active') == 'on'
            )
            messages.success(request, 'Sub-county created successfully!')
            return redirect('sub_county_list')
        except Exception as e:
            messages.error(request, f'Error creating sub-county: {str(e)}')
    
    context = {
        'counties': County.objects.all(),
    }
    return render(request, 'administration/sub_county_form.html', context)


@login_required
def sub_county_update(request, code):
    """Update sub-county"""
    sub_county = get_object_or_404(SubCounty, code=code)
    
    if request.method == 'POST':
        try:
            sub_county.county_id = request.POST.get('county')
            sub_county.name = request.POST.get('name')
            sub_county.headquarters = request.POST.get('headquarters')
            sub_county.is_active = request.POST.get('is_active') == 'on'
            sub_county.save()
            
            messages.success(request, 'Sub-county updated successfully!')
            return redirect('sub_county_detail', code=code)
        except Exception as e:
            messages.error(request, f'Error updating sub-county: {str(e)}')
    
    context = {
        'sub_county': sub_county,
        'counties': County.objects.all(),
    }
    return render(request, 'administration/sub_county_form.html', context)


@login_required
def ward_list(request):
    """List all wards with search and filters"""
    wards = Ward.objects.select_related('sub_county', 'sub_county__county').all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        wards = wards.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    # Filters
    sub_county_filter = request.GET.get('sub_county', '')
    if sub_county_filter:
        wards = wards.filter(sub_county_id=sub_county_filter)
    
    status_filter = request.GET.get('status', '')
    if status_filter:
        is_active = status_filter == 'active'
        wards = wards.filter(is_active=is_active)
    
    # Pagination
    paginator = Paginator(wards, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    stats = {
        'total': Ward.objects.count(),
        'active': Ward.objects.filter(is_active=True).count(),
        'inactive': Ward.objects.filter(is_active=False).count(),
    }
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'sub_counties': SubCounty.objects.all(),
        'current_sub_county': sub_county_filter,
        'current_status': status_filter,
        'stats': stats,
    }
    return render(request, 'administration/ward_list.html', context)


@login_required
def ward_detail(request, code):
    """Ward detail view"""
    ward = get_object_or_404(Ward, code=code)
    
    context = {
        'ward': ward,
        'stats': {
            'population': ward.population or 0,
            'citizens': ward.citizen_set.count(),
            'properties': ward.property_set.count(),
        }
    }
    return render(request, 'administration/ward_detail.html', context)


@login_required
def ward_create(request):
    """Create new ward"""
    if request.method == 'POST':
        try:
            Ward.objects.create(
                sub_county_id=request.POST.get('sub_county'),
                name=request.POST.get('name'),
                code=request.POST.get('code'),
                population=request.POST.get('population') or None,
                is_active=request.POST.get('is_active') == 'on'
            )
            messages.success(request, 'Ward created successfully!')
            return redirect('ward_list')
        except Exception as e:
            messages.error(request, f'Error creating ward: {str(e)}')
    
    context = {
        'sub_counties': SubCounty.objects.filter(is_active=True),
    }
    return render(request, 'administration/ward_form.html', context)


@login_required
def ward_update(request, code):
    """Update ward"""
    ward = get_object_or_404(Ward, code=code)
    
    if request.method == 'POST':
        try:
            ward.sub_county_id = request.POST.get('sub_county')
            ward.name = request.POST.get('name')
            ward.population = request.POST.get('population') or None
            ward.is_active = request.POST.get('is_active') == 'on'
            ward.save()
            
            messages.success(request, 'Ward updated successfully!')
            return redirect('ward_detail', code=code)
        except Exception as e:
            messages.error(request, f'Error updating ward: {str(e)}')
    
    context = {
        'ward': ward,
        'sub_counties': SubCounty.objects.filter(is_active=True),
    }
    return render(request, 'administration/ward_form.html', context)


# ============================================================================
# DEPARTMENT VIEWS
# ============================================================================

@login_required
def department_list(request):
    """List all departments with search and filters"""
    departments = Department.objects.select_related(
        'parent_department', 'head_of_department'
    ).prefetch_related('user_set')
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        departments = departments.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Filters
    status_filter = request.GET.get('status', '')
    if status_filter:
        is_active = status_filter == 'active'
        departments = departments.filter(is_active=is_active)
    
    parent_filter = request.GET.get('parent', '')
    if parent_filter:
        departments = departments.filter(parent_department_id=parent_filter)
    
    # Pagination
    paginator = Paginator(departments, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    stats = {
        'total': Department.objects.count(),
        'active': Department.objects.filter(is_active=True).count(),
        'inactive': Department.objects.filter(is_active=False).count(),
        'with_head': Department.objects.exclude(head_of_department=None).count(),
    }
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'current_status': status_filter,
        'current_parent': parent_filter,
        'parent_departments': Department.objects.filter(parent_department=None),
        'stats': stats,
    }
    return render(request, 'administration/department_list.html', context)


@login_required
def department_detail(request, code):
    """Department detail view"""
    department = get_object_or_404(Department, code=code)
    
    staff = User.objects.filter(department=department)
    
    context = {
        'department': department,
        'staff': staff[:10],  # Latest 10 staff
        'sub_departments': department.department_set.all(),
        'stats': {
            'total_staff': staff.count(),
            'active_staff': staff.filter(is_active_staff=True).count(),
            'sub_departments': department.department_set.count(),
            'revenue_streams': department.revenue_streams.count(),
        }
    }
    return render(request, 'administration/department_detail.html', context)


@login_required
def department_create(request):
    """Create new department"""
    if request.method == 'POST':
        try:
            Department.objects.create(
                name=request.POST.get('name'),
                code=request.POST.get('code'),
                description=request.POST.get('description'),
                parent_department_id=request.POST.get('parent_department') or None,
                head_of_department_id=request.POST.get('head_of_department') or None,
                is_active=request.POST.get('is_active') == 'on'
            )
            messages.success(request, 'Department created successfully!')
            return redirect('department_list')
        except Exception as e:
            messages.error(request, f'Error creating department: {str(e)}')
    
    context = {
        'departments': Department.objects.filter(is_active=True),
        'users': User.objects.filter(is_active_staff=True),
    }
    return render(request, 'administration/department_form.html', context)


@login_required
def department_update(request, code):
    """Update department"""
    department = get_object_or_404(Department, code=code)
    
    if request.method == 'POST':
        try:
            department.name = request.POST.get('name')
            department.description = request.POST.get('description')
            department.parent_department_id = request.POST.get('parent_department') or None
            department.head_of_department_id = request.POST.get('head_of_department') or None
            department.is_active = request.POST.get('is_active') == 'on'
            department.save()
            
            messages.success(request, 'Department updated successfully!')
            return redirect('department_detail', code=code)
        except Exception as e:
            messages.error(request, f'Error updating department: {str(e)}')
    
    context = {
        'department': department,
        'departments': Department.objects.filter(is_active=True).exclude(id=department.id),
        'users': User.objects.filter(is_active_staff=True),
    }
    return render(request, 'administration/department_form.html', context)


@login_required
def department_export_excel(request):
    """Export departments to Excel"""
    departments = Department.objects.select_related(
        'parent_department', 'head_of_department'
    ).all()
    
    # Apply filters
    search_query = request.GET.get('search', '')
    if search_query:
        departments = departments.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Departments"
    
    # Headers
    headers = ['Code', 'Name', 'Parent Department', 'Head of Department', 
               'Total Staff', 'Status', 'Created Date']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for dept in departments:
        ws.append([
            dept.code,
            dept.name,
            dept.parent_department.name if dept.parent_department else '-',
            dept.head_of_department.get_full_name() if dept.head_of_department else '-',
            dept.user_set.count(),
            'Active' if dept.is_active else 'Inactive',
            dept.created_at.strftime('%Y-%m-%d %H:%M'),
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=departments_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


# ============================================================================
# CASE MANAGEMENT VIEWS
# ============================================================================

@login_required
def case_list(request):
    """List all cases with search and filters"""
    cases = Case.objects.select_related(
        'category', 'complainant', 'respondent', 'case_officer'
    ).all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        cases = cases.filter(
            Q(case_number__icontains=search_query) |
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Filters
    category_filter = request.GET.get('category', '')
    if category_filter:
        cases = cases.filter(category_id=category_filter)
    
    status_filter = request.GET.get('status', '')
    if status_filter:
        cases = cases.filter(status=status_filter)
    
    officer_filter = request.GET.get('officer', '')
    if officer_filter:
        cases = cases.filter(case_officer_id=officer_filter)
    
    # Pagination
    paginator = Paginator(cases, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    stats = {
        'total': Case.objects.count(),
        'open': Case.objects.filter(status='open').count(),
        'under_review': Case.objects.filter(status='under_review').count(),
        'resolved': Case.objects.filter(status='resolved').count(),
        'pending_hearing': Case.objects.filter(status='pending_hearing').count(),
    }
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'categories': CaseCategory.objects.filter(is_active=True),
        'officers': User.objects.filter(is_active_staff=True),
        'current_category': category_filter,
        'current_status': status_filter,
        'current_officer': officer_filter,
        'stats': stats,
    }
    return render(request, 'administration/case_list.html', context)


@login_required
def case_detail(request, case_number):
    """Case detail view"""
    case = get_object_or_404(Case, case_number=case_number)
    
    context = {
        'case': case,
        'documents': case.documents.all(),
        'hearings': case.hearings.order_by('-hearing_date'),
        'stats': {
            'documents': case.documents.count(),
            'hearings': case.hearings.count(),
            'days_open': (datetime.now().date() - case.filing_date).days,
        }
    }
    return render(request, 'administration/case_detail.html', context)


@login_required
def case_create(request):
    """Create new case"""
    if request.method == 'POST':
        try:
            # Generate case number
            last_case = Case.objects.order_by('-id').first()
            if last_case:
                last_number = int(last_case.case_number.split('-')[-1])
                case_number = f"CASE-{datetime.now().year}-{last_number + 1:05d}"
            else:
                case_number = f"CASE-{datetime.now().year}-00001"
            
            Case.objects.create(
                case_number=case_number,
                category_id=request.POST.get('category'),
                title=request.POST.get('title'),
                description=request.POST.get('description'),
                complainant_id=request.POST.get('complainant') or None,
                respondent_id=request.POST.get('respondent') or None,
                filing_date=request.POST.get('filing_date'),
                case_officer=request.user,
                status='open'
            )
            messages.success(request, f'Case {case_number} created successfully!')
            return redirect('case_list')
        except Exception as e:
            messages.error(request, f'Error creating case: {str(e)}')
    
    context = {
        'categories': CaseCategory.objects.filter(is_active=True),
        'officers': User.objects.filter(is_active_staff=True),
    }
    return render(request, 'administration/case_form.html', context)


@login_required
def case_update(request, case_number):
    """Update case"""
    case = get_object_or_404(Case, case_number=case_number)
    
    if request.method == 'POST':
        try:
            case.category_id = request.POST.get('category')
            case.title = request.POST.get('title')
            case.description = request.POST.get('description')
            case.status = request.POST.get('status')
            case.case_officer_id = request.POST.get('case_officer') or None
            
            if request.POST.get('hearing_date'):
                case.hearing_date = request.POST.get('hearing_date')
            
            if request.POST.get('resolution'):
                case.resolution = request.POST.get('resolution')
                if request.POST.get('resolution_date'):
                    case.resolution_date = request.POST.get('resolution_date')
            
            case.save()
            
            messages.success(request, 'Case updated successfully!')
            return redirect('case_detail', case_number=case_number)
        except Exception as e:
            messages.error(request, f'Error updating case: {str(e)}')
    
    context = {
        'case': case,
        'categories': CaseCategory.objects.filter(is_active=True),
        'officers': User.objects.filter(is_active_staff=True),
    }
    return render(request, 'administration/case_form.html', context)


@login_required
def case_add_hearing(request, case_number):
    """Add hearing to case"""
    case = get_object_or_404(Case, case_number=case_number)
    
    if request.method == 'POST':
        try:
            CaseHearing.objects.create(
                case=case,
                hearing_date=request.POST.get('hearing_date'),
                venue=request.POST.get('venue'),
                proceedings=request.POST.get('proceedings', ''),
                decision=request.POST.get('decision', ''),
                presiding_officer=request.user
            )
            
            # Update case hearing date
            case.hearing_date = request.POST.get('hearing_date')
            case.status = 'pending_hearing'
            case.save()
            
            messages.success(request, 'Hearing scheduled successfully!')
            return redirect('case_detail', case_number=case_number)
        except Exception as e:
            messages.error(request, f'Error scheduling hearing: {str(e)}')
    
    context = {
        'case': case,
    }
    return render(request, 'administration/case_hearing_form.html', context)


@login_required
def case_export_excel(request):
    """Export cases to Excel"""
    cases = Case.objects.select_related(
        'category', 'complainant', 'respondent', 'case_officer'
    ).all()
    
    # Apply filters
    search_query = request.GET.get('search', '')
    if search_query:
        cases = cases.filter(
            Q(case_number__icontains=search_query) |
            Q(title__icontains=search_query)
        )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cases"
    
    # Headers
    headers = ['Case Number', 'Title', 'Category', 'Complainant', 'Respondent',
               'Filing Date', 'Status', 'Case Officer', 'Days Open']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for case in cases:
        days_open = (datetime.now().date() - case.filing_date).days
        ws.append([
            case.case_number,
            case.title,
            case.category.name,
            str(case.complainant) if case.complainant else '-',
            str(case.respondent) if case.respondent else '-',
            case.filing_date.strftime('%Y-%m-%d'),
            case.get_status_display(),
            case.case_officer.get_full_name() if case.case_officer else '-',
            days_open,
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=cases_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


# ============================================================================
# DOCUMENT MANAGEMENT VIEWS
# ============================================================================

@login_required
def document_list(request):
    """List all electronic documents with search and filters"""
    documents = ElectronicDocument.objects.select_related(
        'category', 'department', 'created_by'
    ).all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        documents = documents.filter(
            Q(document_number__icontains=search_query) |
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Filters
    category_filter = request.GET.get('category', '')
    if category_filter:
        documents = documents.filter(category_id=category_filter)
    
    department_filter = request.GET.get('department', '')
    if department_filter:
        documents = documents.filter(department_id=department_filter)
    
    status_filter = request.GET.get('status', '')
    if status_filter:
        documents = documents.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(documents, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    stats = {
        'total': ElectronicDocument.objects.count(),
        'draft': ElectronicDocument.objects.filter(status='draft').count(),
        'active': ElectronicDocument.objects.filter(status='active').count(),
        'archived': ElectronicDocument.objects.filter(status='archived').count(),
    }
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'categories': DocumentCategory.objects.filter(is_active=True),
        'departments': Department.objects.filter(is_active=True),
        'current_category': category_filter,
        'current_department': department_filter,
        'current_status': status_filter,
        'stats': stats,
    }
    return render(request, 'administration/document_list.html', context)


@login_required
def document_detail(request, document_number):
    """Document detail view"""
    document = get_object_or_404(ElectronicDocument, document_number=document_number)
    
    # Log access
    DocumentAccess.objects.create(
        document=document,
        user=request.user,
        access_type='view',
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    context = {
        'document': document,
        'access_logs': document.access_logs.select_related('user').order_by('-access_date')[:20],
        'stats': {
            'views': document.access_logs.filter(access_type='view').count(),
            'downloads': document.access_logs.filter(access_type='download').count(),
            'edits': document.access_logs.filter(access_type='edit').count(),
        }
    }
    return render(request, 'administration/document_detail.html', context)


@login_required
def document_create(request):
    """Create new document"""
    if request.method == 'POST':
        try:
            # Generate document number
            last_doc = ElectronicDocument.objects.order_by('-id').first()
            if last_doc:
                last_number = int(last_doc.document_number.split('-')[-1])
                doc_number = f"DOC-{datetime.now().year}-{last_number + 1:05d}"
            else:
                doc_number = f"DOC-{datetime.now().year}-00001"
            
            file = request.FILES.get('file')
            
            document = ElectronicDocument.objects.create(
                document_number=doc_number,
                title=request.POST.get('title'),
                category_id=request.POST.get('category'),
                description=request.POST.get('description', ''),
                file=file,
                file_size=file.size if file else 0,
                file_type=file.content_type if file else '',
                department_id=request.POST.get('department'),
                document_date=request.POST.get('document_date'),
                status='draft',
                created_by=request.user
            )
            
            messages.success(request, f'Document {doc_number} created successfully!')
            return redirect('document_detail', document_number=doc_number)
        except Exception as e:
            messages.error(request, f'Error creating document: {str(e)}')
    
    context = {
        'categories': DocumentCategory.objects.filter(is_active=True),
        'departments': Department.objects.filter(is_active=True),
    }
    return render(request, 'administration/document_form.html', context)


@login_required
def document_update(request, document_number):
    """Update document"""
    document = get_object_or_404(ElectronicDocument, document_number=document_number)
    
    if request.method == 'POST':
        try:
            document.title = request.POST.get('title')
            document.category_id = request.POST.get('category')
            document.description = request.POST.get('description', '')
            document.department_id = request.POST.get('department')
            document.document_date = request.POST.get('document_date')
            document.status = request.POST.get('status')
            
            if request.FILES.get('file'):
                file = request.FILES['file']
                document.file = file
                document.file_size = file.size
                document.file_type = file.content_type
            
            document.save()
            
            # Log edit access
            DocumentAccess.objects.create(
                document=document,
                user=request.user,
                access_type='edit',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Document updated successfully!')
            return redirect('document_detail', document_number=document_number)
        except Exception as e:
            messages.error(request, f'Error updating document: {str(e)}')
    
    context = {
        'document': document,
        'categories': DocumentCategory.objects.filter(is_active=True),
        'departments': Department.objects.filter(is_active=True),
    }
    return render(request, 'administration/document_form.html', context)


@login_required
def document_download(request, document_number):
    """Download document"""
    document = get_object_or_404(ElectronicDocument, document_number=document_number)
    
    # Log download access
    DocumentAccess.objects.create(
        document=document,
        user=request.user,
        access_type='download',
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    response = HttpResponse(document.file, content_type=document.file_type)
    response['Content-Disposition'] = f'attachment; filename="{document.title}"'
    return response


@login_required
def document_export_excel(request):
    """Export documents to Excel"""
    documents = ElectronicDocument.objects.select_related(
        'category', 'department', 'created_by'
    ).all()
    
    # Apply filters
    search_query = request.GET.get('search', '')
    if search_query:
        documents = documents.filter(
            Q(document_number__icontains=search_query) |
            Q(title__icontains=search_query)
        )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"
    
    # Headers
    headers = ['Document Number', 'Title', 'Category', 'Department', 
               'Document Date', 'File Type', 'File Size (KB)', 'Status', 
               'Created By', 'Created Date']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for doc in documents:
        ws.append([
            doc.document_number,
            doc.title,
            doc.category.name,
            doc.department.name,
            doc.document_date.strftime('%Y-%m-%d'),
            doc.file_type,
            round(doc.file_size / 1024, 2),
            doc.get_status_display(),
            doc.created_by.get_full_name() if doc.created_by else '-',
            doc.created_at.strftime('%Y-%m-%d %H:%M'),
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=documents_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


# ============================================================================
# CASE CATEGORY MANAGEMENT
# ============================================================================

@login_required
def case_category_list(request):
    """List all case categories"""
    categories = CaseCategory.objects.all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        categories = categories.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    context = {
        'categories': categories,
        'search_query': search_query,
    }
    return render(request, 'administration/case_category_list.html', context)


@login_required
def case_category_create(request):
    """Create new case category"""
    if request.method == 'POST':
        try:
            CaseCategory.objects.create(
                name=request.POST.get('name'),
                code=request.POST.get('code'),
                description=request.POST.get('description'),
                is_active=request.POST.get('is_active') == 'on'
            )
            messages.success(request, 'Case category created successfully!')
            return redirect('case_category_list')
        except Exception as e:
            messages.error(request, f'Error creating category: {str(e)}')
    
    return render(request, 'administration/case_category_form.html')


# ============================================================================
# DOCUMENT CATEGORY MANAGEMENT
# ============================================================================

@login_required
def document_category_list(request):
    """List all document categories"""
    categories = DocumentCategory.objects.all()
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        categories = categories.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    context = {
        'categories': categories,
        'search_query': search_query,
    }
    return render(request, 'administration/document_category_list.html', context)


@login_required
def document_category_create(request):
    """Create new document category"""
    if request.method == 'POST':
        try:
            DocumentCategory.objects.create(
                name=request.POST.get('name'),
                code=request.POST.get('code'),
                description=request.POST.get('description'),
                retention_period_years=request.POST.get('retention_period_years', 0),
                is_active=request.POST.get('is_active') == 'on'
            )
            messages.success(request, 'Document category created successfully!')
            return redirect('document_category_list')
        except Exception as e:
            messages.error(request, f'Error creating category: {str(e)}')
    
    return render(request, 'administration/document_category_form.html')